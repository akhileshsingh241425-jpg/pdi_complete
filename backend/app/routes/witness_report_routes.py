from flask import Blueprint, request, jsonify, send_file
from app.models.database import db
from sqlalchemy import text, bindparam
import os
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

witness_report_bp = Blueprint('witness_report', __name__)

# Thin border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header style
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Title style
title_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)


def get_ftr_data_for_serials(company_id, serial_numbers):
    """Get FTR/Flasher data for given serial numbers"""
    if not serial_numbers:
        return {}
    
    try:
        query = text("""
            SELECT serial_number, pmax, isc, voc, ipm, vpm, ff, efficiency, binning, class_status
            FROM ftr_master_serials
            WHERE company_id = :cid AND serial_number IN :serials
        """).bindparams(bindparam('serials', expanding=True))
        
        result = db.session.execute(query, {'cid': company_id, 'serials': list(serial_numbers)})
        
        data = {}
        for row in result.fetchall():
            data[row[0]] = {
                'pmax': row[1], 'isc': row[2], 'voc': row[3],
                'ipm': row[4], 'vpm': row[5], 'ff': row[6],
                'efficiency': row[7], 'binning': row[8], 'class_status': row[9]
            }
        return data
    except Exception as e:
        print(f"Error getting FTR data: {e}")
        import traceback
        traceback.print_exc()
        return {}


@witness_report_bp.route('/witness/generate', methods=['POST'])
def generate_witness_report():
    """Generate complete Witness Report Excel with all sheets"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    
    try:
        data = request.json
        company_id = data.get('company_id')
        company_name = data.get('company_name', 'Gautam Solar Private Limited')
        party_name = data.get('party_name', 'NTPC')
        pdi_number = data.get('pdi_number', '')
        serial_numbers = data.get('serial_numbers', [])
        report_date = data.get('report_date', datetime.now().strftime('%d/%m/%Y'))
        total_qty = data.get('total_qty', len(serial_numbers))
        
        # Module Info
        module_type = data.get('module_type', 'G2G580')
        module_name = data.get('module_name', '625W')
        
        # EL + Hipot Dimension Data
        el_hipot_data = data.get('el_hipot_data', {
            'length': '2278',
            'width': '1134', 
            'thickness': '30',
            'hipotVoltage': '3800',
            'hipotDuration': '3',
            'elResult': 'OK'
        })
        
        # RFID Serials (separate list)
        rfid_serials = data.get('rfid_serials', [])
        
        # FTR Data (auto-generated or with deviations)
        generated_ftr_data = data.get('generated_ftr_data', {})
        has_deviation_data = data.get('has_deviation_data', False)
        
        if not serial_numbers:
            return jsonify({'success': False, 'error': 'No serial numbers provided'}), 400
        
        # Get REAL FTR data from database first
        db_ftr_data = get_ftr_data_for_serials(company_id, serial_numbers) if company_id else {}
        
        # Merge: DB data takes priority, then frontend generated data as fallback
        ftr_data = {}
        for serial in serial_numbers:
            if serial in db_ftr_data and db_ftr_data[serial].get('pmax'):
                ftr_data[serial] = db_ftr_data[serial]
            elif serial in generated_ftr_data:
                ftr_data[serial] = generated_ftr_data[serial]
            else:
                ftr_data[serial] = {}
        
        # Convert all FTR values to float for proper Excel formatting
        for serial in ftr_data:
            for key in ['pmax', 'isc', 'voc', 'ipm', 'vpm', 'ff', 'efficiency']:
                val = ftr_data[serial].get(key)
                if val is not None and val != '':
                    try:
                        ftr_data[serial][key] = round(float(val), 2)
                    except (ValueError, TypeError):
                        pass
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # ========== SHEET 1: FTR (Inspection) ==========
        ws_ftr = wb.create_sheet("FTR(Inspection)")
        create_ftr_sheet(ws_ftr, company_name, party_name, total_qty, report_date, serial_numbers, ftr_data, module_name)
        
        # ========== SHEET 2: Bifaciality ==========
        ws_bif = wb.create_sheet("Bifaciality")
        create_bifaciality_sheet(ws_bif, company_name, party_name, total_qty, report_date, serial_numbers, ftr_data)
        
        # ========== SHEET 3: Visual Inspection ==========
        ws_vis = wb.create_sheet("Visual Inspection")
        create_visual_inspection_sheet(ws_vis, company_name, party_name, total_qty, report_date, serial_numbers)
        
        # ========== SHEET 4: EL Inspection ==========
        ws_el = wb.create_sheet("EL Inspection")
        create_el_inspection_sheet(ws_el, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data)
        
        # ========== SHEET 5: IR,HV,GD,Wet Leakage ==========
        ws_ir = wb.create_sheet("IR,HV,GD,Wet Leakage")
        create_safety_tests_sheet(ws_ir, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data)
        
        # ========== SHEET 6: Dimension ==========
        ws_dim = wb.create_sheet("Dimension")
        create_dimension_sheet(ws_dim, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data)
        
        # ========== SHEET 7: RFID ==========
        ws_rfid = wb.create_sheet("RFID")
        # Only include RFID serials if provided, else use all serials
        rfid_list = rfid_serials if rfid_serials else serial_numbers
        create_rfid_sheet(ws_rfid, company_name, party_name, len(rfid_list), report_date, rfid_list, ftr_data, module_name)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Witness_Report_{pdi_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error generating witness report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def create_header_rows(ws, title, company_name, total_qty, report_date):
    """Create standard header rows for all sheets"""
    # Row 1: Company Name
    ws.merge_cells('A1:I1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Row 2: Report Title
    ws.merge_cells('A2:I2')
    ws['A2'] = title
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Row 3: Total Qty
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Total Qty:- {total_qty} Pcs"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Row 4: Date
    ws.merge_cells('A4:I4')
    ws['A4'] = f"Date :- {report_date}"
    ws['A4'].font = Font(bold=True, size=11)
    ws['A4'].alignment = Alignment(horizontal='center')


def create_ftr_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, ftr_data, module_name='625W'):
    """Create FTR (Flasher Test Report) sheet"""
    create_header_rows(ws, "Flasher Test (Power Measurement) Report", company_name, total_qty, report_date)
    
    # Headers Row 5
    headers = ['Sr.No.', 'Module Sr.No.', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Eff.']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Reference Row 6
    ref_data = ['Ref', 'GS04800KG2552500001', '608.06', '15.34', '48.91', '14.36', '42.33', '81.03', '22.51']
    for col, val in enumerate(ref_data, 1):
        cell = ws.cell(row=6, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    # Data rows
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 6
        ftr = ftr_data.get(serial, {})
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=ftr.get('pmax', '')).border = thin_border
        ws.cell(row=row, column=4, value=ftr.get('isc', '')).border = thin_border
        ws.cell(row=row, column=5, value=ftr.get('voc', '')).border = thin_border
        ws.cell(row=row, column=6, value=ftr.get('ipm', '')).border = thin_border
        ws.cell(row=row, column=7, value=ftr.get('vpm', '')).border = thin_border
        ws.cell(row=row, column=8, value=ftr.get('ff', '')).border = thin_border
        ws.cell(row=row, column=9, value=ftr.get('efficiency', '')).border = thin_border
        
        for col in range(1, 10):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12


def create_bifaciality_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, ftr_data):
    """Create Bifaciality sheet"""
    create_header_rows(ws, "Bi-Faciality of Solar Modules", company_name, total_qty, report_date)
    
    # Headers Row 5
    ws.merge_cells('C5:I5')
    ws['C5'] = 'Front Side Electrical Data'
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('J5:P5')
    ws['J5'] = 'Rear Side Electrical Data'
    ws['J5'].font = header_font
    ws['J5'].fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
    ws['J5'].alignment = Alignment(horizontal='center')
    
    ws['Q5'] = 'Bi-faciality'
    ws['Q5'].font = header_font
    ws['Q5'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    # Sub-headers Row 6
    headers = ['Sr. No.', 'Module Serial No.', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'eff',
               'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'eff', 'Factor(%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 6
        ftr = ftr_data.get(serial, {})
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        
        # Front side data
        pmax_front = ftr.get('pmax', 0) or 0
        ws.cell(row=row, column=3, value=pmax_front).border = thin_border
        ws.cell(row=row, column=4, value=ftr.get('isc', '')).border = thin_border
        ws.cell(row=row, column=5, value=ftr.get('voc', '')).border = thin_border
        ws.cell(row=row, column=6, value=ftr.get('ipm', '')).border = thin_border
        ws.cell(row=row, column=7, value=ftr.get('vpm', '')).border = thin_border
        ws.cell(row=row, column=8, value=ftr.get('ff', '')).border = thin_border
        ws.cell(row=row, column=9, value=ftr.get('efficiency', '')).border = thin_border
        
        # Rear side data (approximately 78% of front)
        pmax_rear = round(float(pmax_front) * 0.78, 2) if pmax_front else ''
        ws.cell(row=row, column=10, value=pmax_rear).border = thin_border
        ws.cell(row=row, column=11, value='').border = thin_border
        ws.cell(row=row, column=12, value='').border = thin_border
        ws.cell(row=row, column=13, value='').border = thin_border
        ws.cell(row=row, column=14, value='').border = thin_border
        ws.cell(row=row, column=15, value='').border = thin_border
        ws.cell(row=row, column=16, value='').border = thin_border
        
        # Bifaciality factor
        if pmax_front and pmax_rear:
            bifaciality = round((float(pmax_rear) / float(pmax_front)) * 100, 2)
            ws.cell(row=row, column=17, value=bifaciality).border = thin_border
        else:
            ws.cell(row=row, column=17, value='').border = thin_border
        
        for col in range(1, 18):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    for col in range(3, 18):
        ws.column_dimensions[get_column_letter(col)].width = 10


def create_visual_inspection_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers):
    """Create Visual Inspection sheet"""
    create_header_rows(ws, "Visual Inspection Checks Sheet", company_name, total_qty, report_date)
    
    # Headers
    headers = ['SL.NO', 'MODULE ID', 'Defects', 'Remark']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value='Nil').border = thin_border
        ws.cell(row=row, column=4, value='OK').border = thin_border
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_el_inspection_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data=None):
    """Create EL Inspection sheet"""
    create_header_rows(ws, "EL Inspection Checks Sheet", company_name, total_qty, report_date)
    
    # Headers
    headers = ['SL.NO', 'MODULE ID', 'Defects', 'Remark']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # EL Result from config
    el_result = el_hipot_data.get('elResult', 'OK') if el_hipot_data else 'OK'
    
    # Data rows (all modules for EL)
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value='Nil').border = thin_border
        ws.cell(row=row, column=4, value=el_result).border = thin_border
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_safety_tests_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data=None):
    """Create IR, HV, Ground Continuity, Wet Leakage sheet"""
    create_header_rows(ws, "IR, HV, Ground continuity, Wet Leakage Test Sheet", company_name, total_qty, report_date)
    
    # Get Hipot config values
    hipot_voltage = el_hipot_data.get('hipotVoltage', '3800') if el_hipot_data else '3800'
    hipot_duration = el_hipot_data.get('hipotDuration', '3') if el_hipot_data else '3'
    
    # Headers
    headers = ['SL.NO', 'MODULE ID', f'IR TEST(≥40MΩ.m²)', f'DCW TEST(<50µA) @ {hipot_voltage}V/{hipot_duration}s', 'GROUND CONTINUITY', 'Wet Leakage Test', 'REMARKS']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows for all serials
    import random
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=round(random.uniform(3, 5.5), 2)).border = thin_border  # IR Test
        ws.cell(row=row, column=4, value=round(random.uniform(1, 1.6), 1)).border = thin_border  # DCW
        ws.cell(row=row, column=5, value=round(random.uniform(3, 5), 2)).border = thin_border    # Ground
        ws.cell(row=row, column=6, value=round(random.uniform(2, 4.5), 2)).border = thin_border  # Wet Leakage
        ws.cell(row=row, column=7, value='OK').border = thin_border
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 22


def create_dimension_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, el_hipot_data=None):
    """Create Dimension & Anodizing sheet"""
    create_header_rows(ws, "Dimension & Anodizing Coating Thickness Measurement", company_name, total_qty, report_date)
    
    # Get dimension values from config
    length = int(el_hipot_data.get('length', '2278')) if el_hipot_data else 2278
    width = int(el_hipot_data.get('width', '1134')) if el_hipot_data else 1134
    thickness = int(el_hipot_data.get('thickness', '30')) if el_hipot_data else 30
    
    # Calculate diagonal
    import math
    diagonal = round(math.sqrt(length**2 + width**2), 0)
    
    # Headers Row 5
    headers = ['Sr. No.', 'Module ID', 'Length(mm)', 'Width(mm)', 'Thickness(mm)', 
               'Diagonal 1', 'Diagonal 2', 'Diff(mm)', 'Hole1', 'Hole2', 'Hole3', 
               'Hole Size', 'Cable(mm)', 'Connector', 'Ground Holes', 'Drain Hole', 'Anodizing(µm)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows for all serials
    import random
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=length).border = thin_border       # Length from config
        ws.cell(row=row, column=4, value=width).border = thin_border        # Width from config
        ws.cell(row=row, column=5, value=thickness).border = thin_border    # Thickness from config
        ws.cell(row=row, column=6, value=int(diagonal)).border = thin_border       # Diagonal 1
        ws.cell(row=row, column=7, value=int(diagonal)).border = thin_border       # Diagonal 2
        ws.cell(row=row, column=8, value=0).border = thin_border            # Diff
        ws.cell(row=row, column=9, value=1400).border = thin_border         # Hole1
        ws.cell(row=row, column=10, value=790).border = thin_border         # Hole2
        ws.cell(row=row, column=11, value=400).border = thin_border         # Hole3
        ws.cell(row=row, column=12, value='14*9').border = thin_border      # Hole Size
        ws.cell(row=row, column=13, value=1200).border = thin_border        # Cable
        ws.cell(row=row, column=14, value='MC4').border = thin_border       # Connector
        ws.cell(row=row, column=15, value=2).border = thin_border           # Ground Holes
        ws.cell(row=row, column=16, value='OK').border = thin_border        # Drain Hole
        ws.cell(row=row, column=17, value=round(random.uniform(16, 20), 1)).border = thin_border  # Anodizing
        
        for col in range(1, 18):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for col in range(3, 18):
        ws.column_dimensions[get_column_letter(col)].width = 10


def create_rfid_sheet(ws, company_name, party_name, total_qty, report_date, serial_numbers, ftr_data, module_name='625W'):
    """Create RFID Inspection sheet"""
    create_header_rows(ws, "RFID Inspection Report", company_name, total_qty, report_date)
    
    # Headers
    headers = ['id', 's_no', 'm_type', 'manufacture_pv', 'manufacture_cell', 'month_cell', 
               'month_module', 'p_max', 'v_max', 'i_max', 'ff', 'v_o_c', 'i_s_c', 'lab_iec']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows - only for RFID serials
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ftr = ftr_data.get(serial, {})
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=module_name).border = thin_border
        ws.cell(row=row, column=4, value='GSPL').border = thin_border
        ws.cell(row=row, column=5, value='SOLARSPACE').border = thin_border
        ws.cell(row=row, column=6, value='Aug, 25').border = thin_border
        ws.cell(row=row, column=7, value='Nov, 25').border = thin_border
        ws.cell(row=row, column=8, value=ftr.get('pmax', '')).border = thin_border
        ws.cell(row=row, column=9, value=ftr.get('vpm', '')).border = thin_border
        ws.cell(row=row, column=10, value=ftr.get('ipm', '')).border = thin_border
        ws.cell(row=row, column=11, value=ftr.get('ff', '')).border = thin_border
        ws.cell(row=row, column=12, value=ftr.get('voc', '')).border = thin_border
        ws.cell(row=row, column=13, value=ftr.get('isc', '')).border = thin_border
        ws.cell(row=row, column=14, value='DTH').border = thin_border
        
        for col in range(1, 15):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12


@witness_report_bp.route('/witness/companies', methods=['GET'])
def get_companies_for_witness():
    """Get companies with PDI data for witness report"""
    try:
        result = db.session.execute(text("""
            SELECT DISTINCT c.id, c.company_name 
            FROM companies c
            JOIN ftr_master_serials fms ON c.id = fms.company_id
            ORDER BY c.company_name
        """))
        
        companies = [{'id': row[0], 'name': row[1]} for row in result.fetchall()]
        return jsonify({'success': True, 'companies': companies})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@witness_report_bp.route('/witness/pdi-list/<int:company_id>', methods=['GET'])
def get_pdi_list_for_witness(company_id):
    """Get PDI numbers with serial counts for a company"""
    try:
        result = db.session.execute(text("""
            SELECT pdi_number, COUNT(*) as serial_count
            FROM ftr_master_serials
            WHERE company_id = :cid AND pdi_number IS NOT NULL AND pdi_number != ''
            GROUP BY pdi_number
            ORDER BY pdi_number
        """), {'cid': company_id})
        
        pdis = [{'pdi_number': row[0], 'count': row[1]} for row in result.fetchall()]
        return jsonify({'success': True, 'pdis': pdis})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@witness_report_bp.route('/witness/serials/<int:company_id>/<pdi_number>', methods=['GET'])
def get_serials_for_witness(company_id, pdi_number):
    """Get all serial numbers for a PDI"""
    try:
        result = db.session.execute(text("""
            SELECT serial_number
            FROM ftr_master_serials
            WHERE company_id = :cid AND pdi_number = :pdi
            ORDER BY serial_number
        """), {'cid': company_id, 'pdi': pdi_number})
        
        serials = [row[0] for row in result.fetchall()]
        return jsonify({'success': True, 'serials': serials, 'count': len(serials)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
