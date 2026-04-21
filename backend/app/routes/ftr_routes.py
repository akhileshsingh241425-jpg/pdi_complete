"""
FTR (Field Test Report) Routes
"""

from flask import Blueprint, request, jsonify, send_file
from app.services.ftr_pdf_generator import create_ftr_report
from config import Config
import os
import pymysql
import requests as http_requests
import re
import json
import pandas as pd
from datetime import datetime
import time

ftr_bp = Blueprint('ftr', __name__, url_prefix='/api/ftr')

# Global cache for dispatch data (per party_id)
# Structure: {party_id: {'data': {serial: details}, 'set': set(), 'timestamp': time}}
DISPATCH_CACHE = {}
DISPATCH_CACHE_TTL = 600  # 10 minutes


def get_db_connection():
    """Get database connection using Config"""
    return pymysql.connect(
        host=Config.MYSQL_HOST,
        user=Config.MYSQL_USER,
        password=Config.MYSQL_PASSWORD,
        database=Config.MYSQL_DB,
        cursorclass=pymysql.cursors.DictCursor
    )


@ftr_bp.route('/generate-report', methods=['POST'])
def generate_ftr_report():
    """
    Generate FTR PDF report from test data
    
    Expected JSON payload:
    {
        "producer": "Gautam Solar",
        "moduleType": "630W",
        "serialNumber": "GS04890KG2582504241",
        "testDate": "2025/12/19",
        "testTime": "15:34:39",
        "irradiance": 1001.09,
        "moduleTemp": 24.88,
        "ambientTemp": 23.62,
        "moduleArea": 2.70,
        "modulePower": 630,
        "results": {
            "pmax": 629.96,
            "vpm": 45.39,
            "ipm": 13.90,
            "voc": 53.82,
            "isc": 14.70,
            "fillFactor": 79.60,
            "rs": 0.12,
            "rsh": 2461.33,
            "efficiency": 23.32
        }
    }
    """
    try:
        data = request.json
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        # Get template path
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        template_path = os.path.join(backend_dir, '..', 'frontend', 'public', 'IV curve template.pdf')
        
        if not os.path.exists(template_path):
            return jsonify({"error": "Template PDF not found"}), 404
        
        # Get graph image path based on module power
        module_power = data.get('modulePower')
        graph_image_path = None
        
        if module_power:
            graph_image_path = os.path.join(
                backend_dir, '..', 'frontend', 'public', 'iv_curves', 
                f'{module_power}.png'
            )
            
            # Check if graph exists
            if not os.path.exists(graph_image_path):
                print(f"Graph image not found: {graph_image_path}")
                graph_image_path = None
        
        # Generate PDF
        pdf_output = create_ftr_report(template_path, data, graph_image_path)
        
        # Generate filename
        serial_number = data.get('serialNumber', 'unknown')
        filename = f"FTR_Report_{serial_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Send file
        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error generating FTR report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@ftr_bp.route('/test', methods=['GET'])
def test_ftr():
    """Test endpoint to verify FTR routes are working"""
    return jsonify({
        "status": "ok",
        "message": "FTR routes are working"
    })


@ftr_bp.route('/available-serial-numbers', methods=['GET'])
def get_available_serial_numbers():
    """
    Get all serial numbers from Master FTR that are not yet assigned to any PDI batch
    Returns list of available serial numbers
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all serial numbers from master_ftr
        cursor.execute("""
            SELECT 
                serial_number as serialNumber,
                module_wattage as wattage,
                created_at as uploadDate
            FROM master_ftr
            WHERE serial_number NOT IN (
                SELECT DISTINCT serial_number 
                FROM pdi_serial_numbers 
                WHERE serial_number IS NOT NULL
            )
            ORDER BY created_at DESC
        """)
        
        available_serials = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "available_serials": available_serials,
            "count": len(available_serials)
        })
        
    except Exception as e:
        print(f"Error fetching available serial numbers: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@ftr_bp.route('/upload-master-ftr', methods=['POST'])
def upload_master_ftr():
    """
    Upload Master FTR serial numbers to database (Company-level, not PDI-specific)
    
    Expected JSON payload:
    {
        "serialNumbers": [
            {"serialNumber": "GS123", "wattage": "625"},
            ...
        ],
        "companyId": 1
    }
    """
    try:
        data = request.json
        serial_numbers = data.get('serialNumbers', [])
        company_id = data.get('companyId')
        
        print(f"Received {len(serial_numbers)} serial numbers for company {company_id}")
        
        if not serial_numbers:
            return jsonify({"success": False, "error": "No serial numbers provided"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        inserted_count = 0
        duplicate_count = 0
        
        for serial in serial_numbers:
            serial_number = serial.get('serialNumber')
            wattage = serial.get('wattage')
            
            if not serial_number:
                continue
            
            try:
                # Check if serial number already exists
                cursor.execute("""
                    SELECT COUNT(*) as count FROM master_ftr 
                    WHERE serial_number = %s
                """, (serial_number,))
                
                result = cursor.fetchone()
                if result[0] > 0:
                    duplicate_count += 1
                    continue
                
                # Insert new serial number (Company-level, not PDI-specific)
                cursor.execute("""
                    INSERT INTO master_ftr 
                    (serial_number, module_wattage, company_id, created_at)
                    VALUES (%s, %s, %s, NOW())
                """, (serial_number, wattage, company_id))
                
                inserted_count += 1
                print(f"Inserted: {serial_number}")
                
            except Exception as e:
                print(f"Error inserting serial {serial_number}: {e}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        message = f"Uploaded {inserted_count} serial numbers"
        if duplicate_count > 0:
            message += f" ({duplicate_count} duplicates skipped)"
        
        print(f"Upload complete: {message}")
        
        return jsonify({
            "success": True,
            "count": inserted_count,
            "duplicates": duplicate_count,
            "message": message
        })
    
    except Exception as e:
        print(f"Error uploading Master FTR: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@ftr_bp.route('/assign-pdi-serials', methods=['POST'])
def assign_pdi_serials():
    """
    Assign serial numbers to a specific PDI batch
    This marks these serials as "used" for this PDI
    
    Expected JSON payload:
    {
        "serialNumbers": [
            {"serialNumber": "GS123", "wattage": "625"},
            ...
        ],
        "companyId": 1,
        "pdiNumber": "PDI-1"
    }
    """
    try:
        data = request.json
        serial_numbers = data.get('serialNumbers', [])
        company_id = data.get('companyId')
        pdi_number = data.get('pdiNumber')
        
        print(f"Assigning {len(serial_numbers)} serials to {pdi_number}")
        
        if not serial_numbers or not pdi_number:
            return jsonify({"success": False, "error": "Serial numbers and PDI number required"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        inserted_count = 0
        duplicate_count = 0
        
        for serial in serial_numbers:
            serial_number = serial.get('serialNumber')
            
            if not serial_number:
                continue
            
            try:
                # Check if already assigned to this or another PDI
                cursor.execute("""
                    SELECT pdi_number FROM pdi_serial_numbers 
                    WHERE serial_number = %s
                """, (serial_number,))
                
                result = cursor.fetchone()
                if result:
                    duplicate_count += 1
                    print(f"Serial {serial_number} already assigned to {result[0]}")
                    continue
                
                # Insert into pdi_serial_numbers
                cursor.execute("""
                    INSERT INTO pdi_serial_numbers 
                    (pdi_number, serial_number, company_id, created_at)
                    VALUES (%s, %s, %s, NOW())
                """, (pdi_number, serial_number, company_id))
                
                inserted_count += 1
                print(f"Assigned: {serial_number} to {pdi_number}")
                
            except Exception as e:
                print(f"Error assigning serial {serial_number}: {e}")
                continue
        
        # Update PDI records ftr_uploaded flag
        if inserted_count > 0:
            try:
                cursor.execute("""
                    UPDATE production_records
                    SET ftr_uploaded = TRUE
                    WHERE pdi = %s AND company_id = %s
                """, (pdi_number, company_id))
            except Exception as e:
                print(f"Error updating ftr_uploaded flag: {e}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        message = f"Assigned {inserted_count} serial numbers to {pdi_number}"
        if duplicate_count > 0:
            message += f" ({duplicate_count} already assigned)"
        
        print(f"Assignment complete: {message}")
        
        return jsonify({
            "success": True,
            "count": inserted_count,
            "duplicates": duplicate_count,
            "message": message
        })
        
    except Exception as e:
        print(f"Error assigning PDI serials: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@ftr_bp.route('/pdi-dashboard/<company_id>', methods=['GET'])
def get_pdi_dashboard(company_id):
    """
    Get PDI Dashboard data - shows barcode tracking status
    - Total assigned barcodes
    - Packed (in stock)
    - Dispatched
    - Remaining
    """
    import requests
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if table exists FIRST
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM information_schema.tables 
            WHERE table_schema = %s AND table_name = 'pdi_serial_numbers'
        """, (Config.MYSQL_DB,))
        
        table_check = cursor.fetchone()
        if not table_check or table_check['count'] == 0:
            # Table doesn't exist, create it
            print("Creating pdi_serial_numbers table...")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS pdi_serial_numbers (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pdi_number VARCHAR(50) NOT NULL,
                    serial_number VARCHAR(100) NOT NULL,
                    company_id INT,
                    production_record_id INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_pdi (pdi_number),
                    INDEX idx_serial (serial_number),
                    INDEX idx_company (company_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            conn.commit()
            print("âœ… pdi_serial_numbers table created successfully")
            
            # Return empty data since table was just created
            cursor.close()
            conn.close()
            return jsonify({
                "success": True,
                "summary": {
                    "total_assigned": 0,
                    "total_tracked": 0,
                    "packed": 0,
                    "dispatched": 0,
                    "pending": 0,
                    "unknown": 0,
                    "packed_percent": 0,
                    "dispatched_percent": 0,
                    "pending_percent": 0
                },
                "details": {
                    "packed": [],
                    "dispatched": [],
                    "pending": []
                },
                "message": "PDI tracking initialized. No serial numbers assigned yet."
            }), 200
        
        # Get all PDI numbers and their serial counts for this company
        cursor.execute("""
            SELECT 
                pdi_number,
                COUNT(*) as serial_count,
                MIN(created_at) as assigned_date
            FROM pdi_serial_numbers 
            WHERE company_id = %s
            GROUP BY pdi_number
            ORDER BY MIN(created_at) DESC
        """, (company_id,))
        
        pdi_summary = cursor.fetchall()
        
        # Get all serial numbers for this company
        cursor.execute("""
            SELECT serial_number, pdi_number, created_at
            FROM pdi_serial_numbers 
            WHERE company_id = %s
            ORDER BY created_at DESC
        """, (company_id,))
        
        all_serials = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # If no serials assigned yet, return empty data
        if not all_serials or len(all_serials) == 0:
            return jsonify({
                "success": True,
                "summary": {
                    "total_assigned": 0,
                    "total_tracked": 0,
                    "packed": 0,
                    "dispatched": 0,
                    "pending": 0,
                    "unknown": 0,
                    "packed_percent": 0,
                    "dispatched_percent": 0,
                    "pending_percent": 0
                },
                "details": {
                    "packed": [],
                    "dispatched": [],
                    "pending": []
                },
                "pdi_wise": [],
                "recent_dispatched": [],
                "recent_packed": [],
                "recent_pending": [],
                "message": "No serial numbers assigned to this company yet."
            }), 200
        
        # Now track each serial using external API
        BARCODE_TRACKING_API = 'https://umanmrp.in/api/get_barcode_tracking.php'
        
        total_assigned = len(all_serials)
        packed_count = 0
        dispatched_count = 0
        pending_count = 0
        unknown_count = 0
        
        packed_serials = []
        dispatched_serials = []
        pending_serials = []
        
        # Track status for each serial (limit to avoid timeout)
        serials_to_track = all_serials[:500]  # Limit to 500 for performance
        
        for serial_data in serials_to_track:
            serial = serial_data['serial_number']
            try:
                response = requests.post(
                    BARCODE_TRACKING_API,
                    data={'barcode': serial},
                    timeout=5
                )
                
                if response.status_code == 200:
                    data = response.json()
                    
                    if data.get('success') and data.get('data'):
                        tracking_data = data['data']
                        
                        # Check dispatch status
                        dispatch_info = tracking_data.get('dispatch', {})
                        packing_info = tracking_data.get('packing', {})
                        
                        if dispatch_info.get('dispatch_date'):
                            dispatched_count += 1
                            dispatched_serials.append({
                                'serial': serial,
                                'pdi': serial_data['pdi_number'],
                                'dispatch_date': dispatch_info.get('dispatch_date'),
                                'vehicle_no': dispatch_info.get('vehicle_no', ''),
                                'party': dispatch_info.get('party_name', '')
                            })
                        elif packing_info.get('packing_date'):
                            packed_count += 1
                            packed_serials.append({
                                'serial': serial,
                                'pdi': serial_data['pdi_number'],
                                'packing_date': packing_info.get('packing_date'),
                                'box_no': packing_info.get('box_no', '')
                            })
                        else:
                            pending_count += 1
                            pending_serials.append({
                                'serial': serial,
                                'pdi': serial_data['pdi_number']
                            })
                    else:
                        pending_count += 1
                        pending_serials.append({
                            'serial': serial,
                            'pdi': serial_data['pdi_number']
                        })
                else:
                    unknown_count += 1
                    
            except Exception as e:
                unknown_count += 1
                continue
        
        # Calculate percentages
        tracked_total = packed_count + dispatched_count + pending_count
        
        return jsonify({
            "success": True,
            "summary": {
                "total_assigned": total_assigned,
                "total_tracked": len(serials_to_track),
                "packed": packed_count,
                "dispatched": dispatched_count,
                "pending": pending_count,
                "unknown": unknown_count,
                "packed_percent": round((packed_count / tracked_total * 100), 1) if tracked_total > 0 else 0,
                "dispatched_percent": round((dispatched_count / tracked_total * 100), 1) if tracked_total > 0 else 0,
                "pending_percent": round((pending_count / tracked_total * 100), 1) if tracked_total > 0 else 0
            },
            "details": {
                "packed": packed_serials,
                "dispatched": dispatched_serials,
                "pending": pending_serials
            },
            "pdi_wise": pdi_summary,
            "recent_dispatched": dispatched_serials[:20],
            "recent_packed": packed_serials[:20],
            "recent_pending": pending_serials[:20]
        })
        
    except Exception as e:
        print(f"Error getting PDI dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@ftr_bp.route('/pdi-dashboard-quick/<company_id>', methods=['GET'])
def get_pdi_dashboard_quick(company_id):
    """
    Quick PDI Dashboard - just database counts without external API calls
    For fast loading
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get total assigned
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM pdi_serial_numbers 
            WHERE company_id = %s
        """, (company_id,))
        total = cursor.fetchone()['total']
        
        # Get PDI wise summary
        cursor.execute("""
            SELECT 
                pdi_number,
                COUNT(*) as serial_count,
                DATE(MIN(created_at)) as assigned_date
            FROM pdi_serial_numbers 
            WHERE company_id = %s
            GROUP BY pdi_number
            ORDER BY MIN(created_at) DESC
            LIMIT 50
        """, (company_id,))
        
        pdi_summary = cursor.fetchall()
        
        # Get recent serials
        cursor.execute("""
            SELECT serial_number, pdi_number, created_at
            FROM pdi_serial_numbers 
            WHERE company_id = %s
            ORDER BY created_at DESC
            LIMIT 100
        """, (company_id,))
        
        recent_serials = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "summary": {
                "total_assigned": total,
                "pdi_count": len(pdi_summary)
            },
            "pdi_wise": pdi_summary,
            "recent_serials": recent_serials
        })
        
    except Exception as e:
        print(f"Error getting quick PDI dashboard: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


# ===== COMPANY NAME MAPPING (Local â†’ MRP) - Copied from ai_assistant_routes =====
COMPANY_NAME_MAP = {
    'Rays Power': 'RAYS POWER INFRA LIMITED',
    'rays power': 'RAYS POWER INFRA LIMITED',
    'RAYS POWER': 'RAYS POWER INFRA LIMITED',
    
    'Larsen & Toubro': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'larsen & toubro': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'LARSEN & TOUBRO': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'L&T': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'l&t': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    
    'Sterlin and Wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'sterlin and wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'STERLIN AND WILSON': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'Sterling and Wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'sterling and wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'S&W': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    's&w': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    
    'KPI Green Energy': 'KPI GREEN ENERGY LIMITED',
    'KPI GREEN ENERGY': 'KPI GREEN ENERGY LIMITED',
    'kpi green energy': 'KPI GREEN ENERGY LIMITED',
    'KPI': 'KPI GREEN ENERGY LIMITED',
}

# Party IDs for dispatch history API - exact mapping from MRP system
PARTY_IDS = {
    # Rays Power
    'rays power': '931db2c5-b016-4914-b378-69e9f22562a7',
    'rays power infra': '931db2c5-b016-4914-b378-69e9f22562a7',
    'rays power infra limited': '931db2c5-b016-4914-b378-69e9f22562a7',
    'rays': '931db2c5-b016-4914-b378-69e9f22562a7',
    # Larsen & Toubro
    'larsen & toubro': 'a005562f-568a-46e9-bf2e-700affb171e8',
    'larsen and toubro': 'a005562f-568a-46e9-bf2e-700affb171e8',
    'larsen & toubro limited': 'a005562f-568a-46e9-bf2e-700affb171e8',
    'l&t': 'a005562f-568a-46e9-bf2e-700affb171e8',
    'lnt': 'a005562f-568a-46e9-bf2e-700affb171e8',
    # Sterling and Wilson
    'sterling and wilson': '141b81a0-2bab-4790-b825-3c8734d41484',
    'sterlin and wilson': '141b81a0-2bab-4790-b825-3c8734d41484',
    'sterling & wilson': '141b81a0-2bab-4790-b825-3c8734d41484',
    'sterling and wilson renewable energy limited': '141b81a0-2bab-4790-b825-3c8734d41484',
    's&w': '141b81a0-2bab-4790-b825-3c8734d41484',
    'sw': '141b81a0-2bab-4790-b825-3c8734d41484',
    # KPI Green Energy
    'kpi green energy': 'kpi-green-energy-party-id',
    'kpi': 'kpi-green-energy-party-id',
}


def get_mrp_party_name_ftr(local_name):
    """Map local company name to MRP full party name"""
    if local_name in COMPANY_NAME_MAP:
        return COMPANY_NAME_MAP[local_name]
    lower_name = local_name.strip().lower()
    for key, value in COMPANY_NAME_MAP.items():
        if key.lower() == lower_name:
            return value
    return local_name


def normalize_company_name(name):
    """Normalize company name for matching - remove special chars, extra spaces"""
    import re
    name = name.strip().lower()
    name = re.sub(r'[&]+', ' and ', name)
    name = re.sub(r'[^a-z0-9\s]', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


# Party Dispatch History API
DISPATCH_HISTORY_API = 'https://umanmrp.in/api/party-dispatch-history.php'


def auto_sync_mrp_cache(matched_company, party_id):
    """
    Automatically sync MRP dispatch data to local cache.
    Loops through pages 1-1000 to fetch ALL data.
    Called when cache is empty or stale (older than 1 hour).
    """
    from datetime import datetime, timedelta
    
    print(f"[Auto Sync] Starting sync for {matched_company}...")
    
    # Date range - last 1 year
    to_date = datetime.now().strftime('%Y-%m-%d')
    from_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    
    all_barcodes = []
    limit = 100
    empty_pages = 0
    max_empty = 3  # Stop after 3 consecutive empty pages
    
    for page in range(1, 1001):  # Loop pages 1 to 1000
        try:
            payload = {
                "party_id": party_id,
                "from_date": from_date,
                "to_date": to_date,
                "page": page,
                "limit": limit
            }
            
            response = http_requests.post(
                "https://umanmrp.in/api/party-dispatch-history.php",
                json=payload,
                timeout=60
            )
            
            if response.status_code != 200:
                print(f"[Auto Sync] API error on page {page}: {response.status_code}")
                break
                
            result = response.json()
            dispatch_summary = result.get('dispatch_summary', [])
            
            if not dispatch_summary:
                empty_pages += 1
                if empty_pages >= max_empty:
                    print(f"[Auto Sync] Page {page}: Empty - stopping (3 consecutive empty)")
                    break
                continue
            else:
                empty_pages = 0
            
            page_count = 0
            for item in dispatch_summary:
                pallet_nos = item.get('pallet_nos', {})
                status = item.get('status', 'Packed')
                dispatch_party = item.get('dispatch_party', '')
                vehicle_no = item.get('vehicle_no', '')
                dispatch_date = item.get('dispatch_date') or item.get('date', '')
                invoice_no = item.get('invoice_no', '')
                
                if isinstance(pallet_nos, dict):
                    for pallet_no, serials_str in pallet_nos.items():
                        if serials_str and isinstance(serials_str, str):
                            serials = serials_str.strip().split()
                            for serial in serials:
                                serial = serial.strip()
                                if serial:
                                    all_barcodes.append({
                                        'serial_number': serial.upper(),
                                        'pallet_no': pallet_no,
                                        'status': status,
                                        'dispatch_party': dispatch_party,
                                        'vehicle_no': vehicle_no,
                                        'dispatch_date': dispatch_date,
                                        'invoice_no': invoice_no,
                                        'company': matched_company,
                                        'party_id': party_id
                                    })
                                    page_count += 1
            
            print(f"[Auto Sync] Page {page}: {len(dispatch_summary)} dispatches, {page_count} serials (Total: {len(all_barcodes)})")
            
        except Exception as e:
            print(f"[Auto Sync] Error on page {page}: {e}")
            break
    
    print(f"[Auto Sync] Fetched {len(all_barcodes)} barcodes from MRP API")
    
    # Save to database
    if all_barcodes:
        try:
            conn = get_db_connection()
            with conn.cursor() as cursor:
                for barcode in all_barcodes:
                    cursor.execute("""
                        INSERT INTO mrp_dispatch_cache 
                        (serial_number, pallet_no, status, dispatch_party, vehicle_no, dispatch_date, invoice_no, company, party_id, synced_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON DUPLICATE KEY UPDATE
                        pallet_no = VALUES(pallet_no),
                        status = VALUES(status),
                        dispatch_party = VALUES(dispatch_party),
                        vehicle_no = VALUES(vehicle_no),
                        dispatch_date = VALUES(dispatch_date),
                        invoice_no = VALUES(invoice_no),
                        synced_at = NOW()
                    """, (
                        barcode['serial_number'],
                        barcode['pallet_no'],
                        barcode['status'],
                        barcode['dispatch_party'],
                        barcode['vehicle_no'],
                        barcode['dispatch_date'] if barcode['dispatch_date'] else None,
                        barcode['invoice_no'],
                        barcode['company'],
                        barcode['party_id']
                    ))
                conn.commit()
            conn.close()
            print(f"[Auto Sync] Saved {len(all_barcodes)} barcodes to local cache")
        except Exception as e:
            print(f"[Auto Sync] DB save error: {e}")
    
    return len(all_barcodes)


def fetch_dispatch_history(company_name):
    """
    Fetch dispatch history from LOCAL CACHE (mrp_dispatch_cache table)
    Auto-syncs from MRP API if cache is empty or older than 1 hour.
    Returns mrp_lookup dict: barcode â†’ {status, pallet_no, dispatch_date, vehicle_no, etc.}
    """
    from datetime import datetime, timedelta
    
    mrp_party_name = get_mrp_party_name_ftr(company_name)
    lower_name = company_name.strip().lower()
    
    print(f"[Dispatch History] Company: {company_name}, lower_name: {lower_name}")
    
    # Get party_id - keyword based matching
    party_id = None
    matched_company = None
    if 'rays' in lower_name:
        party_id = '931db2c5-b016-4914-b378-69e9f22562a7'
        matched_company = 'Rays Power'
        print(f"[Dispatch History] Matched: Rays Power")
    elif 'larsen' in lower_name or 'l&t' in lower_name or 'lnt' in lower_name:
        party_id = 'a005562f-568a-46e9-bf2e-700affb171e8'
        matched_company = 'L&T'
        print(f"[Dispatch History] Matched: Larsen & Toubro")
    elif 'sterling' in lower_name or 'sterlin' in lower_name or 's&w' in lower_name:
        party_id = '141b81a0-2bab-4790-b825-3c8734d41484'
        matched_company = 'Sterling'
        print(f"[Dispatch History] Matched: Sterling and Wilson")
    elif 'kpi' in lower_name:
        party_id = 'kpi-green-energy-party-id'
        matched_company = 'KPI'
        print(f"[Dispatch History] Matched: KPI Green Energy")
    
    if not party_id:
        print(f"[Dispatch History] No party_id found for: {company_name}")
        return {}, mrp_party_name, None
    
    print(f"[Dispatch History] Using party_id: {party_id}, matched_company: {matched_company}")
    
    mrp_lookup = {}
    
    # ===== Check if cache needs refresh (empty or older than 1 hour) =====
    needs_sync = False
    cache_count = 0
    last_sync = None
    
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Check cache count and last sync time
            cursor.execute("""
                SELECT COUNT(*) as cnt, MAX(synced_at) as last_sync 
                FROM mrp_dispatch_cache 
                WHERE company = %s
            """, (matched_company,))
            result = cursor.fetchone()
            cache_count = result['cnt'] or 0
            last_sync = result['last_sync']
        conn.close()
        
        if cache_count == 0:
            needs_sync = True
            print(f"[Dispatch History] Cache EMPTY - will sync")
        elif last_sync:
            # Check if older than 1 hour
            time_diff = datetime.now() - last_sync
            if time_diff > timedelta(hours=1):
                needs_sync = True
                print(f"[Dispatch History] Cache STALE ({time_diff}) - will sync")
            else:
                print(f"[Dispatch History] Cache FRESH (synced {time_diff} ago)")
        else:
            needs_sync = True
            
    except Exception as e:
        print(f"[Dispatch History] Cache check error: {e}")
        needs_sync = True
    
    # ===== Auto sync if needed =====
    if needs_sync:
        auto_sync_mrp_cache(matched_company, party_id)
    
    # ===== Fetch from local cache =====
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT serial_number, pallet_no, status, dispatch_party, vehicle_no, 
                       dispatch_date, invoice_no
                FROM mrp_dispatch_cache 
                WHERE company = %s
            """, (matched_company,))
            cache_rows = cursor.fetchall()
        conn.close()
        
        print(f"[Dispatch History] LOCAL CACHE: Found {len(cache_rows)} serials for {matched_company}")
        
        for row in cache_rows:
            barcode = row['serial_number'].strip().upper()
            mrp_lookup[barcode] = {
                'status': row['status'] or 'Dispatched',
                'pallet_no': row['pallet_no'] or '',
                'dispatch_party': row['dispatch_party'] or '',
                'vehicle_no': row['vehicle_no'] or '',
                'dispatch_date': str(row['dispatch_date']) if row['dispatch_date'] else '',
                'invoice_no': row['invoice_no'] or '',
                'date': str(row['dispatch_date']) if row['dispatch_date'] else ''
            }
        
        print(f"[Dispatch History] Returning {len(mrp_lookup)} barcodes from cache")
        return mrp_lookup, mrp_party_name, party_id
            
    except Exception as e:
        print(f"[Dispatch History] Cache read error: {e}")
    
    # ===== FALLBACK: Direct fetch from MRP API =====
    print(f"[Dispatch History] Fallback to direct MRP API fetch...")
    
    page = 1
    limit = 50
    total_barcodes = 0
    total_dispatches = 0
    
    # Date range - 1 year back to today
    today = datetime.now().strftime('%Y-%m-%d')
    from_date = '2025-01-01'  # Start from 2025
    
    while True:
        try:
            payload = {
                'party_id': party_id,
                'from_date': from_date,
                'to_date': today,
                'page': page,
                'limit': limit
            }
            print(f"[Dispatch History] API call page {page}: {payload}")
            
            response = http_requests.post(
                DISPATCH_HISTORY_API,
                json=payload,
                timeout=60
            )
            
            print(f"[Dispatch History] Response status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                
                if data.get('status') == 'success':
                    dispatch_summary = data.get('dispatch_summary', [])
                    pagination = data.get('pagination', {})
                    
                    print(f"[Dispatch History] Page {page}: {len(dispatch_summary)} dispatches, pagination: {pagination}")
                    
                    for dispatch in dispatch_summary:
                        dispatch_id = dispatch.get('dispatch_id', '')
                        dispatch_date = dispatch.get('dispatch_date', '')
                        vehicle_no = dispatch.get('vehicle_no', '')
                        invoice_no = dispatch.get('invoice_no', '')
                        factory_name = dispatch.get('factory_name', '')
                        pallet_nos = dispatch.get('pallet_nos', {})
                        
                        total_dispatches += 1
                        
                        # Parse pallet_nos - each pallet has space-separated serial numbers
                        if isinstance(pallet_nos, dict):
                            for pallet_no, barcodes_str in pallet_nos.items():
                                if isinstance(barcodes_str, str):
                                    # Split by space to get individual barcodes
                                    barcodes = barcodes_str.strip().split()
                                    for barcode in barcodes:
                                        barcode = barcode.strip().upper()
                                        if barcode:
                                            mrp_lookup[barcode] = {
                                                'status': 'Dispatched',
                                                'pallet_no': str(pallet_no),
                                                'dispatch_party': factory_name,
                                                'vehicle_no': vehicle_no,
                                                'dispatch_date': dispatch_date,
                                                'invoice_no': invoice_no,
                                                'dispatch_id': dispatch_id,
                                                'date': dispatch_date
                                            }
                                            total_barcodes += 1
                    
                    print(f"[Dispatch History] Running total: {total_barcodes} barcodes from {total_dispatches} dispatches")
                    
                    # Check for next page
                    has_next = pagination.get('has_next_page', False)
                    if has_next:
                        page += 1
                    else:
                        break
                else:
                    print(f"[Dispatch History] API status not success: {data.get('status')}, message: {data.get('message', 'N/A')}")
                    break
            else:
                print(f"[Dispatch History] HTTP error: {response.status_code}")
                break
                
        except Exception as e:
            print(f"[Dispatch History] Error on page {page}: {str(e)}")
            import traceback
            traceback.print_exc()
            break
    
    print(f"[Dispatch History] FINAL: {len(mrp_lookup)} unique barcodes from {total_dispatches} dispatches")
    return mrp_lookup, mrp_party_name, party_id


@ftr_bp.route('/dispatch-tracking/<company_id>', methods=['GET'])
def get_dispatch_tracking(company_id):
    """
    Proxy endpoint to fetch dispatch tracking data from MRP API.
    Solves CORS issues and adds company name mapping.
    Uses same logic as ai_assistant_routes.py get_all_mrp_data().
    """
    try:
        # Get company from database
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()
        cursor.close()
        conn.close()

        if not company:
            return jsonify({"success": False, "error": "Company not found"}), 404

        company_name = company.get('company_name') or company.get('companyName', '')
        
        print(f"\n[Dispatch Tracking] Company ID: {company_id}, Name: {company_name}")

        # Fetch dispatch history from new MRP API
        mrp_lookup, mrp_party_name, _ = fetch_dispatch_history(company_name)

        print(f"[Dispatch Tracking] Dispatch barcodes found: {len(mrp_lookup)}")

        if not mrp_lookup:
            return jsonify({
                "success": True,
                "company_name": company_name,
                "mrp_party_name": mrp_party_name,
                "summary": {
                    "total_assigned": 0,
                    "packed": 0,
                    "dispatched": 0,
                    "pending": 0,
                    "packed_percent": 0,
                    "dispatched_percent": 0,
                    "pending_percent": 0
                },
                "pdi_groups": [],
                "vehicle_groups": [],
                "dispatch_groups": [],
                "message": "No dispatch data found in MRP system"
            })

        # All barcodes in dispatch history are dispatched
        total = len(mrp_lookup)
        dispatched = total
        packed = 0
        pending = 0

        # Group by vehicle_no (dispatch_party)
        vehicle_map = {}
        pallet_map = {}

        for barcode, info in mrp_lookup.items():
            vehicle = info.get('vehicle_no', '') or info.get('dispatch_party', '') or 'Unknown'
            pallet_no = info.get('pallet_no', '')
            dispatch_date = info.get('dispatch_date', '') or info.get('date', '')
            invoice_no = info.get('invoice_no', '')
            factory_name = info.get('factory_name', '')

            # Group by vehicle
            if vehicle not in vehicle_map:
                vehicle_map[vehicle] = {
                    'dispatch_party': vehicle,
                    'vehicle_no': vehicle,
                    'dispatch_date': dispatch_date,
                    'invoice_no': invoice_no,
                    'factory_name': factory_name,
                    'module_count': 0,
                    'pallets': set(),
                    'serials': []
                }
            vehicle_map[vehicle]['module_count'] += 1
            if pallet_no:
                vehicle_map[vehicle]['pallets'].add(pallet_no)
            if len(vehicle_map[vehicle]['serials']) < 50:
                vehicle_map[vehicle]['serials'].append(barcode)

            # Group by pallet
            if pallet_no:
                if pallet_no not in pallet_map:
                    pallet_map[pallet_no] = {
                        'pallet_no': pallet_no,
                        'module_count': 0,
                        'vehicle_no': vehicle,
                        'dispatch_date': dispatch_date,
                        'serials': []
                    }
                pallet_map[pallet_no]['module_count'] += 1
                if len(pallet_map[pallet_no]['serials']) < 20:
                    pallet_map[pallet_no]['serials'].append(barcode)

        # Build dispatch groups
        dispatch_groups = []
        for v_name, v_data in vehicle_map.items():
            dispatch_groups.append({
                'dispatch_party': v_name,
                'vehicle_no': v_data['vehicle_no'],
                'dispatch_date': v_data['dispatch_date'],
                'invoice_no': v_data['invoice_no'],
                'factory_name': v_data['factory_name'],
                'module_count': v_data['module_count'],
                'pallet_count': len(v_data['pallets']),
                'pallets': sorted(list(v_data['pallets'])),
                'serials': v_data['serials']
            })
        dispatch_groups.sort(key=lambda x: x['module_count'], reverse=True)

        pallet_groups = sorted(pallet_map.values(), key=lambda x: str(x['pallet_no']))

        print(f"[Dispatch Tracking] Result: dispatched={dispatched}, vehicles={len(vehicle_map)}, pallets={len(pallet_map)}")

        return jsonify({
            "success": True,
            "company_name": company_name,
            "mrp_party_name": mrp_party_name,
            "summary": {
                "total_assigned": total,
                "packed": packed,
                "dispatched": dispatched,
                "pending": pending,
                "packed_percent": 0,
                "dispatched_percent": 100 if total > 0 else 0,
                "pending_percent": 0
            },
            "pallet_groups": pallet_groups,
            "dispatch_groups": dispatch_groups
        })

    except http_requests.exceptions.Timeout:
        print(f"[Dispatch Tracking] MRP API timeout for company_id={company_id}")
        return jsonify({
            "success": False,
            "error": "MRP API timed out. Please try again."
        }), 504

    except Exception as e:
        print(f"[Dispatch Tracking] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@ftr_bp.route('/dispatch-tracking-pdi/<company_id>', methods=['GET'])
def get_dispatch_tracking_pdi_wise(company_id):
    """
    PDI-wise dispatch tracking â€” cross-references ftr_master_serials with MRP data.
    Same logic as ai_assistant_routes.py check_pdi_dispatch_status().
    
    1. Gets all PDI assignments from ftr_master_serials
    2. Fetches MRP barcode data for the company
    3. Cross-references each serial to classify: Dispatched / Packed / Pending
    4. Returns PDI-wise breakdown
    """
    try:
        # Step 1: Get company info
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()

        if not company:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "error": "Company not found"}), 404

        company_name = company.get('company_name') or company.get('companyName', '')

        # Step 2: Get all PDI assignments from pdi_serial_numbers
        cursor.execute("""
            SELECT pdi_number, COUNT(*) as total, MIN(created_at) as assigned_date
            FROM pdi_serial_numbers
            WHERE company_id = %s AND pdi_number IS NOT NULL
            GROUP BY pdi_number
            ORDER BY assigned_date DESC
        """, (company_id,))
        pdi_assignments = cursor.fetchall()

        if not pdi_assignments:
            cursor.close()
            conn.close()
            return jsonify({
                "success": True,
                "company_name": company_name,
                "message": "No PDI assignments found for this company",
                "summary": {"total": 0, "dispatched": 0, "packed": 0, "pending": 0},
                "pdi_wise": []
            })

        # Step 3: Get ALL serials for this company from pdi_serial_numbers
        cursor.execute("""
            SELECT serial_number, pdi_number
            FROM pdi_serial_numbers
            WHERE company_id = %s AND pdi_number IS NOT NULL
        """, (company_id,))
        all_serials_rows = cursor.fetchall()
        cursor.close()
        conn.close()

        # Build PDI â†’ serials mapping
        pdi_serials_map = {}
        for row in all_serials_rows:
            serial = row['serial_number']
            pdi = row['pdi_number']
            if pdi not in pdi_serials_map:
                pdi_serials_map[pdi] = []
            pdi_serials_map[pdi].append(serial)

        print(f"\n[PDI Dispatch] Company: {company_name}, PDIs: {list(pdi_serials_map.keys())}, Total serials: {len(all_serials_rows)}")

        # Step 4: Fetch dispatch history from new MRP API
        mrp_lookup, mrp_party_name, _ = fetch_dispatch_history(company_name)
        print(f"[PDI Dispatch] Dispatch lookup built: {len(mrp_lookup)} barcodes")

        # Step 6: Cross-reference each PDI's serials with MRP data
        overall_dispatched = 0
        overall_packed = 0
        overall_pending = 0
        overall_total = 0
        pdi_wise_results = []

        for pdi_info in pdi_assignments:
            pdi_number = pdi_info['pdi_number']
            pdi_total = pdi_info['total']
            assigned_date = pdi_info['assigned_date']
            serials = pdi_serials_map.get(pdi_number, [])

            dispatched = 0
            packed = 0
            pending = 0
            dispatch_parties = {}

            # Collect serial details for click-through
            dispatched_serials = []
            packed_serials = []
            pending_serials = []

            for serial in serials:
                if serial in mrp_lookup:
                    info = mrp_lookup[serial]
                    if info['status'] == 'Dispatched':
                        dispatched += 1
                        dp = info['dispatch_party']
                        if dp not in dispatch_parties:
                            dispatch_parties[dp] = 0
                        dispatch_parties[dp] += 1
                        if len(dispatched_serials) < 500:
                            dispatched_serials.append({
                                'serial': serial,
                                'pallet_no': info['pallet_no'],
                                'dispatch_party': info['dispatch_party']
                            })
                    elif info['status'] == 'Packed':
                        packed += 1
                        if len(packed_serials) < 500:
                            packed_serials.append({
                                'serial': serial,
                                'pallet_no': info['pallet_no']
                            })
                    else:
                        pending += 1
                        if len(pending_serials) < 200:
                            pending_serials.append({'serial': serial})
                else:
                    pending += 1
                    if len(pending_serials) < 200:
                        pending_serials.append({'serial': serial})

            total = dispatched + packed + pending
            overall_dispatched += dispatched
            overall_packed += packed
            overall_pending += pending
            overall_total += total

            pdi_wise_results.append({
                'pdi_number': pdi_number,
                'total': total,
                'dispatched': dispatched,
                'packed': packed,
                'pending': pending,
                'dispatched_percent': round((dispatched / total) * 100) if total > 0 else 0,
                'packed_percent': round((packed / total) * 100) if total > 0 else 0,
                'pending_percent': round((pending / total) * 100) if total > 0 else 0,
                'assigned_date': str(assigned_date) if assigned_date else '',
                'dispatch_parties': [{'party': k, 'count': v} for k, v in sorted(dispatch_parties.items(), key=lambda x: x[1], reverse=True)],
                'dispatched_serials': dispatched_serials,
                'packed_serials': packed_serials,
                'pending_serials': pending_serials
            })

        # Calculate overall percentages
        overall_dispatched_pct = round((overall_dispatched / overall_total) * 100) if overall_total > 0 else 0
        overall_packed_pct = round((overall_packed / overall_total) * 100) if overall_total > 0 else 0
        overall_pending_pct = round((overall_pending / overall_total) * 100) if overall_total > 0 else 0

        print(f"[PDI Dispatch] Result: total={overall_total}, dispatched={overall_dispatched}, packed={overall_packed}, pending={overall_pending}")

        return jsonify({
            "success": True,
            "company_name": company_name,
            "mrp_party_name": mrp_party_name,
            "summary": {
                "total": overall_total,
                "dispatched": overall_dispatched,
                "packed": overall_packed,
                "pending": overall_pending,
                "dispatched_percent": overall_dispatched_pct,
                "packed_percent": overall_packed_pct,
                "pending_percent": overall_pending_pct
            },
            "pdi_wise": pdi_wise_results
        })

    except http_requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "MRP API timed out. Please try again."}), 504
    except Exception as e:
        print(f"[PDI Dispatch] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================================
# PDI Production Status â€” kitne ban gaye, kitne pending
# ============================================================
@ftr_bp.route('/pdi-production-status/<company_id>', methods=['GET'])
def get_pdi_production_status(company_id):
    """
    Returns PDI-wise production status for a company:
    - FTR tested serials per PDI (from ftr_master_serials)
    - Production output per PDI (from production_records)
    - Planned qty per PDI (from pdi_batches + master_orders)
    - Pending = planned - produced
    
    Query params:
    - force_refresh=true: Bypass cache and fetch fresh data
    """
    # Check if force refresh requested
    force_refresh = request.args.get('force_refresh', '').lower() == 'true'
    print(f"[PDI Production] === API CALLED === force_refresh={force_refresh}, time={datetime.now().strftime('%H:%M:%S')}")
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 1. Get company info
        cursor.execute("SELECT id, company_name FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()
        if not company:
            conn.close()
            return jsonify({"success": False, "error": "Company not found"}), 404

        # 2. PDI Serial Numbers â€” PDI-wise count (modules assigned via Excel upload)
        ftr_pdi_counts = {}
        try:
            cursor.execute("""
                SELECT pdi_number, COUNT(*) as count, MIN(created_at) as assigned_date
                FROM pdi_serial_numbers
                WHERE company_id = %s AND pdi_number IS NOT NULL
                GROUP BY pdi_number
                ORDER BY pdi_number
            """, (company_id,))
            for row in cursor.fetchall():
                ftr_pdi_counts[row['pdi_number']] = {
                    'ftr_count': row['count'],
                    'assigned_date': str(row['assigned_date']) if row['assigned_date'] else None
                }
        except Exception as e:
            print(f"[PDI Production] pdi_serial_numbers query error: {e}")

        # 3. Production Records â€” PDI-wise total production (day + night)
        production_pdi_counts = {}
        try:
            cursor.execute("""
                SELECT pdi, 
                       SUM(COALESCE(day_production, 0) + COALESCE(night_production, 0)) as total_production,
                       COUNT(*) as record_count,
                       MIN(date) as start_date,
                       MAX(date) as last_date
                FROM production_records
                WHERE company_id = %s AND pdi IS NOT NULL AND pdi != ''
                GROUP BY pdi
                ORDER BY pdi
            """, (company_id,))
            for row in cursor.fetchall():
                production_pdi_counts[row['pdi']] = {
                    'total_production': int(row['total_production'] or 0),
                    'record_count': row['record_count'],
                    'start_date': str(row['start_date']) if row['start_date'] else None,
                    'last_date': str(row['last_date']) if row['last_date'] else None
                }
        except Exception as e:
            print(f"[PDI Production] production_records query error: {e}")

        # 4. PDI Batches â€” planned modules per PDI (from master_orders)
        planned_pdi = {}
        total_order_qty = 0
        order_number = None
        try:
            cursor.execute("""
                SELECT mo.id as order_id, mo.order_number, mo.total_modules,
                       pb.pdi_number, pb.planned_modules, pb.actual_modules, pb.status as batch_status
                FROM master_orders mo
                JOIN pdi_batches pb ON pb.order_id = mo.id
                WHERE mo.company_id = %s
                ORDER BY pb.batch_sequence
            """, (company_id,))
            rows = cursor.fetchall()
            for row in rows:
                total_order_qty = row['total_modules'] or 0
                order_number = row['order_number']
                planned_pdi[row['pdi_number']] = {
                    'planned_modules': row['planned_modules'] or 0,
                    'actual_modules': row['actual_modules'] or 0,
                    'batch_status': row['batch_status']
                }
        except Exception as e:
            print(f"[PDI Production] pdi_batches query error (table may not exist): {e}")

        # 5. Total FTR count (all serials, including available)
        total_ftr = 0
        total_ftr_ok = 0
        total_rejected = 0
        total_available = 0
        try:
            cursor.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN class_status = 'OK' OR class_status IS NULL THEN 1 ELSE 0 END) as ok_count,
                    SUM(CASE WHEN class_status = 'REJECTED' THEN 1 ELSE 0 END) as rejected,
                    SUM(CASE WHEN status = 'available' THEN 1 ELSE 0 END) as available
                FROM ftr_master_serials
                WHERE company_id = %s
            """, (company_id,))
            row = cursor.fetchone()
            if row:
                total_ftr = row['total'] or 0
                total_ftr_ok = row['ok_count'] or 0
                total_rejected = row['rejected'] or 0
                total_available = row['available'] or 0
        except Exception as e:
            print(f"[PDI Production] total FTR query error: {e}")

        # 6. Get ALL serials per PDI for dispatch cross-reference from pdi_serial_numbers
        pdi_serials_map = {}
        try:
            cursor.execute("""
                SELECT serial_number, pdi_number
                FROM pdi_serial_numbers
                WHERE company_id = %s AND pdi_number IS NOT NULL AND serial_number IS NOT NULL
            """, (company_id,))
            for row in cursor.fetchall():
                pdi = row['pdi_number']
                serial = row['serial_number']
                if pdi and serial and not serial.strip().startswith('20'):
                    if pdi not in pdi_serials_map:
                        pdi_serials_map[pdi] = []
                    pdi_serials_map[pdi].append(serial.strip())
            print(f"[PDI Production] Total serials from pdi_serial_numbers: {sum(len(s) for s in pdi_serials_map.values())} across {len(pdi_serials_map)} PDIs")
            # Print sample serial for debug
            if pdi_serials_map:
                first_pdi = list(pdi_serials_map.keys())[0]
                if pdi_serials_map[first_pdi]:
                    print(f"[PDI Production] Sample local serial: {pdi_serials_map[first_pdi][0]}")
        except Exception as e:
            print(f"[PDI Production] serial fetch error: {e}")

        # 6b. Get Packing API data (get_barcode_tracking.php)
        company_name = company['company_name']
        lower_name = company_name.strip().lower()
        
        # Map to MRP party name for packing API â€” fetch ALL sub-parties for comparison
        packing_party_names = []
        if 'rays' in lower_name:
            packing_party_names = ['RAYS POWER INFRA PRIVATE LIMITED', 'Rays', 'Rays-NTPC', 'Rays-NTPC-Barethi']
        elif 'larsen' in lower_name or 'l&t' in lower_name or 'lnt' in lower_name:
            packing_party_names = ['LARSEN & TOUBRO LIMITED, CONSTRUCTION', 'L&T', 'LARSEN & TOUBRO LIMITED', 'LARSEN AND TOUBRO']
        elif 'sterling' in lower_name or 'sterlin' in lower_name or 's&w' in lower_name:
            packing_party_names = ['STERLING AND WILSON RENEWABLE ENERGY LIMITED', 'S&W', 'S&W - NTPC']
        
        # Fetch packed serials from packing API
        packed_lookup = {}  # serial -> {pallet_no, running_order, ...}
        party_fetch_counts = {}  # party_name -> count of barcodes fetched
        print(f"[PDI Production] Fetching packing data for: {packing_party_names}")
        
        for party_name in packing_party_names:
            try:
                response = http_requests.post(
                    'https://umanmrp.in/api/get_barcode_tracking.php',
                    json={'party_name': party_name},
                    timeout=120
                )
                print(f"[PDI Production] Packing API response for {party_name}: status={response.status_code}")
                if response.status_code == 200:
                    data = response.json()
                    print(f"[PDI Production] Packing API data: status={data.get('status')}, count={len(data.get('data', []))}")
                    if data.get('status') == 'success' or data.get('data'):
                        items = data.get('data', [])
                        for item in items:
                            barcode = item.get('barcode', '').strip().upper()
                            if barcode:
                                packed_party = item.get('packed_party', '') or party_name
                                packed_lookup[barcode] = {
                                    'pallet_no': item.get('pallet_no', ''),
                                    'running_order': item.get('running_order', ''),
                                    'party_name': packed_party,
                                    'status': 'Packed'
                                }
                                party_fetch_counts[packed_party] = party_fetch_counts.get(packed_party, 0) + 1
                        if items:
                            print(f"[PDI Production] Sample MRP barcode: {items[0].get('barcode', 'N/A')}")
                    else:
                        party_fetch_counts[party_name] = 0
                else:
                    party_fetch_counts[party_name] = 0
            except Exception as e:
                print(f"[PDI Production] Packing API error ({party_name}): {e}")
                party_fetch_counts[party_name] = 0
        
        print(f"[PDI Production] Total packed serials: {len(packed_lookup)}")

        # 6c. Get LIVE dispatch data from MRP Dispatch API (party-dispatch-history.php)
        pdi_dispatch_data = {}  # pdi -> {dispatched: count, packed: count, serials: [...]}
        
        # Party ID mapping for dispatch API
        PARTY_IDS = {
            'rays': '931db2c5-b016-4914-b378-69e9f22562a7',
            'l&t': 'a005562f-568a-46e9-bf2e-700affb171e8',
            'larsen': 'a005562f-568a-46e9-bf2e-700affb171e8',
            'lnt': 'a005562f-568a-46e9-bf2e-700affb171e8',
            'sterling': '141b81a0-2bab-4790-b825-3c8734d41484',
            'sterlin': '141b81a0-2bab-4790-b825-3c8734d41484',
            's&w': '141b81a0-2bab-4790-b825-3c8734d41484'
        }
        
        # Find party_id for this company
        party_id = None
        matched_company = None
        for key, pid in PARTY_IDS.items():
            if key in lower_name:
                party_id = pid
                matched_company = key.upper()
                break
        
        # 6c. Fetch LIVE dispatched serials using BOTH APIs for maximum freshness
        # OLD API = real-time with pallet/vehicle/date details (limit=10000)
        # NEW API = historical backup for older data (barcodes_only)
        dispatched_serials_set = set()
        dispatched_details = {}  # serial -> {pallet_no, dispatch_party, vehicle_no, date}
        dispatch_api_error = None
        
        if party_id:
            try:
                from datetime import timedelta
                to_date = datetime.now().strftime('%Y-%m-%d')
                from_date = (datetime.now() - timedelta(days=730)).strftime('%Y-%m-%d')
                
                # STEP 1: OLD API (LIVE, real-time) â€” fetch ALL PAGES for complete detailed data
                print(f"[PDI Production] STEP 1: Fetching ALL dispatch pages from OLD API (detailed)...")
                try:
                    page = 1
                    max_pages = 20  # Safety limit
                    total_old_entries = 0
                    
                    while page <= max_pages:
                        old_response = http_requests.post(
                            'https://umanmrp.in/api/party-dispatch-history.php',
                            json={
                                'party_id': party_id,
                                'from_date': from_date,
                                'to_date': to_date,
                                'page': page,
                                'limit': 10000
                            },
                            timeout=120,
                            headers={'Cache-Control': 'no-cache', 'Pragma': 'no-cache'}
                        )
                        
                        if old_response.status_code == 200:
                            old_data = old_response.json()
                            dispatch_summary = old_data.get('dispatch_summary', [])
                            
                            if not dispatch_summary:
                                print(f"[PDI Production] OLD API page {page}: 0 entries, stopping pagination")
                                break
                            
                            total_old_entries += len(dispatch_summary)
                            print(f"[PDI Production] OLD API page {page}: {len(dispatch_summary)} entries")
                            
                            for dispatch in dispatch_summary:
                                dispatch_date = dispatch.get('dispatch_date', '')
                                vehicle_no = dispatch.get('vehicle_no', '')
                                invoice_no = dispatch.get('invoice_no', '')
                                pallet_nos = dispatch.get('pallet_nos', {})
                                
                                if isinstance(pallet_nos, dict):
                                    for pallet_no, barcodes_str in pallet_nos.items():
                                        if isinstance(barcodes_str, str):
                                            serials = barcodes_str.strip().split()
                                            for serial in serials:
                                                serial = serial.strip().upper()
                                                if serial:
                                                    dispatched_serials_set.add(serial)
                                                    dispatched_details[serial] = {
                                                        'pallet_no': pallet_no,
                                                        'dispatch_party': invoice_no,
                                                        'vehicle_no': vehicle_no,
                                                        'date': dispatch_date
                                                    }
                            
                            page += 1
                        else:
                            print(f"[PDI Production] OLD API page {page} HTTP error: {old_response.status_code}")
                            break
                    
                    print(f"[PDI Production] OLD API TOTAL: {total_old_entries} entries, {len(dispatched_serials_set)} serials with full details")
                except Exception as e:
                    print(f"[PDI Production] OLD API error: {e}")
                
                # STEP 2: NEW API (barcodes_only) â€” safety backup for any serials missed by OLD API
                old_api_count = len(dispatched_serials_set)
                print(f"[PDI Production] STEP 2: Fetching NEW API (barcodes_only) as backup...")
                try:
                    new_response = http_requests.post(
                        'https://umanmrp.in/api/party-dispatch-history1.php',
                        json={
                            'party_id': party_id,
                            'from_date': from_date,
                            'to_date': to_date,
                            'barcodes_only': True
                        },
                        timeout=300,
                        headers={'Cache-Control': 'no-cache', 'Pragma': 'no-cache'}
                    )
                    
                    if new_response.status_code == 200:
                        new_data = new_response.json()
                        if new_data.get('status') == 'success':
                            barcodes = new_data.get('barcodes', [])
                            new_count = 0
                            for barcode_str in barcodes:
                                if barcode_str and isinstance(barcode_str, str):
                                    individual_serials = barcode_str.strip().split()
                                    for serial in individual_serials:
                                        serial = serial.strip().upper()
                                        if serial and serial not in dispatched_serials_set:
                                            dispatched_serials_set.add(serial)
                                            dispatched_details[serial] = {
                                                'pallet_no': '',
                                                'dispatch_party': '',
                                                'vehicle_no': '',
                                                'date': ''
                                            }
                                            new_count += 1
                            print(f"[PDI Production] NEW API added {new_count} extra serials (backup)")
                        else:
                            print(f"[PDI Production] NEW API error: {new_data.get('message', 'unknown')}")
                    else:
                        print(f"[PDI Production] NEW API HTTP error: {new_response.status_code}")
                except Exception as e:
                    print(f"[PDI Production] NEW API error: {e}")
                
                print(f"[PDI Production] FINAL TOTAL: {len(dispatched_serials_set)} dispatched serials (OLD: {old_api_count}, NEW backup: {len(dispatched_serials_set) - old_api_count})")
                
            except Exception as e:
                dispatch_api_error = str(e)
                print(f"[PDI Production] Dispatch API error: {e}")
                import traceback
                traceback.print_exc()
        
        # 6d. Now categorize each PDI's serials into 3 categories
        for pdi, serials in pdi_serials_map.items():
            pdi_dispatch_data[pdi] = {
                'dispatched': 0,
                'packed': 0,
                'not_packed': 0,
                'dispatched_serials': [],
                'packed_serials': [],
                'not_packed_serials': [],
                'pallet_groups': {}
            }
            
            for serial in serials:
                serial_upper = serial.strip().upper()
                
                if serial_upper in dispatched_serials_set:
                    # DISPATCHED - in live dispatch API
                    pdi_dispatch_data[pdi]['dispatched'] += 1
                    dispatch_info = dispatched_details.get(serial_upper, {})
                    packing_info = packed_lookup.get(serial_upper, {})
                    pallet_no = dispatch_info.get('pallet_no') or packing_info.get('pallet_no') or ''
                    pdi_dispatch_data[pdi]['dispatched_serials'].append({
                        'serial': serial,
                        'pallet_no': pallet_no,
                        'dispatch_party': dispatch_info.get('dispatch_party', ''),
                        'vehicle_no': dispatch_info.get('vehicle_no', ''),
                        'date': dispatch_info.get('date', ''),
                        'sub_party': packing_info.get('party_name', ''),
                        'status': 'Dispatched'
                    })
                    # Add to pallet group
                    pallet_key = pallet_no or 'Unknown'
                    if pallet_key not in pdi_dispatch_data[pdi]['pallet_groups']:
                        pdi_dispatch_data[pdi]['pallet_groups'][pallet_key] = {
                            'pallet_no': pallet_key, 'status': 'Dispatched', 'count': 0, 'serials': []
                        }
                    pdi_dispatch_data[pdi]['pallet_groups'][pallet_key]['count'] += 1
                    if len(pdi_dispatch_data[pdi]['pallet_groups'][pallet_key]['serials']) < 50:
                        pdi_dispatch_data[pdi]['pallet_groups'][pallet_key]['serials'].append(serial)
                        
                elif serial_upper in packed_lookup:
                    # PACKED (NOT DISPATCHED) - in packing API but not dispatch cache
                    pdi_dispatch_data[pdi]['packed'] += 1
                    packing_info = packed_lookup[serial_upper]
                    pdi_dispatch_data[pdi]['packed_serials'].append({
                        'serial': serial,
                        'pallet_no': packing_info.get('pallet_no', ''),
                        'party_name': packing_info.get('party_name', ''),
                        'sub_party': packing_info.get('party_name', ''),
                        'status': 'Packed'
                    })
                    # Add to pallet group
                    pallet_no = packing_info.get('pallet_no') or 'Unknown'
                    if pallet_no not in pdi_dispatch_data[pdi]['pallet_groups']:
                        pdi_dispatch_data[pdi]['pallet_groups'][pallet_no] = {
                            'pallet_no': pallet_no, 'status': 'Packed', 'count': 0, 'serials': []
                        }
                    pdi_dispatch_data[pdi]['pallet_groups'][pallet_no]['count'] += 1
                    if len(pdi_dispatch_data[pdi]['pallet_groups'][pallet_no]['serials']) < 50:
                        pdi_dispatch_data[pdi]['pallet_groups'][pallet_no]['serials'].append(serial)
                        
                else:
                    # NOT PACKED - not in packing API at all
                    pdi_dispatch_data[pdi]['not_packed'] += 1
                    pdi_dispatch_data[pdi]['not_packed_serials'].append({
                        'serial': serial,
                        'pallet_no': '',
                        'status': 'Not Packed'
                    })
        
        total_dispatched = sum(d['dispatched'] for d in pdi_dispatch_data.values())
        total_packed = sum(d['packed'] for d in pdi_dispatch_data.values())
        total_not_packed = sum(d['not_packed'] for d in pdi_dispatch_data.values())
        print(f"[PDI Production] Final: Dispatched={total_dispatched}, Packed={total_packed}, Not Packed={total_not_packed}")

        conn.close()

        # 7. Build debug info
        total_dispatched_in_result = sum(d.get('dispatched', 0) + d.get('packed', 0) for d in pdi_dispatch_data.values())
        mrp_error = None
        
        # Get last refresh timestamp
        current_time_stamp = time.time()
        server_current_time = datetime.fromtimestamp(current_time_stamp).strftime('%Y-%m-%d %H:%M:%S')
        last_refresh_time = server_current_time
        
        print(f"[PDI Production] Debug: LIVE mode, last_refresh={last_refresh_time}")
        
        # Collect sample serials for debugging format mismatches
        all_local_serials = []
        for pdi, serials in pdi_serials_map.items():
            all_local_serials.extend([s.strip().upper() for s in serials])
        
        sample_mrp_barcodes = list(dispatched_serials_set)[:5] if dispatched_serials_set else []
        sample_packed_barcodes = list(packed_lookup.keys())[:5] if packed_lookup else []
        sample_local_serials = all_local_serials[:5] if all_local_serials else []
        
        # Count exact matches
        local_set = set(all_local_serials)
        dispatch_matches = len(local_set.intersection(dispatched_serials_set))
        packed_matches = len(local_set.intersection(set(packed_lookup.keys())))
        
        # 7a. Extra Dispatched â€” serials dispatched to party but NOT in any local PDI
        extra_dispatched_set = dispatched_serials_set - local_set
        extra_dispatched_serials = []
        extra_pallet_groups = {}
        for serial in sorted(extra_dispatched_set):
            detail = dispatched_details.get(serial, {})
            pallet_no = detail.get('pallet_no', '')
            packing_info = packed_lookup.get(serial, {})
            if not pallet_no:
                pallet_no = packing_info.get('pallet_no', '')
            extra_dispatched_serials.append({
                'serial': serial,
                'pallet_no': pallet_no,
                'dispatch_party': detail.get('dispatch_party', ''),
                'vehicle_no': detail.get('vehicle_no', ''),
                'date': detail.get('date', ''),
                'sub_party': packing_info.get('party_name', ''),
                'status': 'Extra Dispatched'
            })
            pallet_key = pallet_no or 'Unknown'
            if pallet_key not in extra_pallet_groups:
                extra_pallet_groups[pallet_key] = {
                    'pallet_no': pallet_key, 'status': 'Extra Dispatched', 'count': 0, 'serials': []
                }
            extra_pallet_groups[pallet_key]['count'] += 1
            if len(extra_pallet_groups[pallet_key]['serials']) < 50:
                extra_pallet_groups[pallet_key]['serials'].append(serial)
        
        extra_dispatched_count = len(extra_dispatched_set)
        extra_pallet_list = sorted(extra_pallet_groups.values(), key=lambda x: str(x['pallet_no']))
        print(f"[PDI Production] Extra Dispatched (not in any PDI): {extra_dispatched_count} serials, {len(extra_pallet_list)} pallets")
        
        # 7b. Extra Packed â€” serials packed but NOT in any local PDI (and not dispatched)
        packed_serials_set = set(packed_lookup.keys())
        extra_packed_set = packed_serials_set - local_set - dispatched_serials_set
        extra_packed_serials = []
        extra_packed_pallet_groups = {}
        for serial in sorted(extra_packed_set):
            packing_info = packed_lookup.get(serial, {})
            pallet_no = packing_info.get('pallet_no', '')
            extra_packed_serials.append({
                'serial': serial,
                'pallet_no': pallet_no,
                'party_name': packing_info.get('party_name', ''),
                'sub_party': packing_info.get('party_name', ''),
                'running_order': packing_info.get('running_order', ''),
                'status': 'Extra Packed'
            })
            pallet_key = pallet_no or 'Unknown'
            if pallet_key not in extra_packed_pallet_groups:
                extra_packed_pallet_groups[pallet_key] = {
                    'pallet_no': pallet_key, 'status': 'Extra Packed', 'count': 0, 'serials': []
                }
            extra_packed_pallet_groups[pallet_key]['count'] += 1
            if len(extra_packed_pallet_groups[pallet_key]['serials']) < 50:
                extra_packed_pallet_groups[pallet_key]['serials'].append(serial)
        
        extra_packed_count = len(extra_packed_set)
        extra_packed_pallet_list = sorted(extra_packed_pallet_groups.values(), key=lambda x: str(x['pallet_no']))
        print(f"[PDI Production] Extra Packed (not in any PDI): {extra_packed_count} serials, {len(extra_packed_pallet_list)} pallets")
        
        debug_info = {
            'matched_company': matched_company,
            'total_pdi_with_dispatch': len(pdi_dispatch_data),
            'total_dispatched_serials': total_dispatched_in_result,
            'company_name': company_name,
            'using_live_api': True,
            'using_cache': False,
            'cache_age_seconds': 0,
            'last_refresh_time': last_refresh_time,
            'server_current_time': server_current_time,
            'live_dispatch_count': len(dispatched_serials_set),
            'live_packed_count': len(packed_lookup),
            'mrp_barcodes_total': len(dispatched_serials_set),
            'packed_barcodes_total': len(packed_lookup),
            'local_serials_total': len(all_local_serials),
            'dispatch_matches': dispatch_matches,
            'packed_matches': packed_matches,
            'sample_mrp_barcodes': sample_mrp_barcodes,
            'sample_packed_barcodes': sample_packed_barcodes,
            'sample_local_serials': sample_local_serials,
            'dispatch_api_error': dispatch_api_error,
            'packing_party_names': packing_party_names,
            'party_fetch_counts': party_fetch_counts
        }

        # 8. Build combined PDI-wise results
        all_pdis = sorted(set(
            list(ftr_pdi_counts.keys()) +
            list(production_pdi_counts.keys()) +
            list(planned_pdi.keys())
        ))

        pdi_wise = []
        grand_produced = 0
        grand_planned = 0
        grand_ftr = 0
        grand_dispatched = 0
        grand_packed = 0
        grand_disp_pending = 0

        for pdi in all_pdis:
            ftr_info = ftr_pdi_counts.get(pdi, {})
            prod_info = production_pdi_counts.get(pdi, {})
            plan_info = planned_pdi.get(pdi, {})

            ftr_count = ftr_info.get('ftr_count', 0)
            produced = prod_info.get('total_production', 0)
            planned = plan_info.get('planned_modules', 0)

            # "Ban gaye" = produced from production records; if zero, fall back to FTR count
            ban_gaye = produced if produced > 0 else ftr_count
            # Pending = planned - produced (if planned exists)
            pending = max(0, planned - ban_gaye) if planned > 0 else 0
            # Progress %
            progress = round((ban_gaye / planned) * 100, 1) if planned > 0 else (100 if ban_gaye > 0 else 0)

            grand_produced += ban_gaye
            grand_planned += planned
            grand_ftr += ftr_count

            # Get dispatch data from 3-category classification
            dispatch_data = pdi_dispatch_data.get(pdi, {})
            dispatched = dispatch_data.get('dispatched', 0)
            packed = dispatch_data.get('packed', 0)
            not_packed = dispatch_data.get('not_packed', 0)
            dispatched_serials = dispatch_data.get('dispatched_serials', [])
            packed_serials = dispatch_data.get('packed_serials', [])
            not_packed_serials = dispatch_data.get('not_packed_serials', [])
            
            # Get pallet groups 
            pallet_groups_dict = dispatch_data.get('pallet_groups', {})
            pallet_list = sorted(pallet_groups_dict.values(), key=lambda x: str(x['pallet_no']))
            # Limit serials per pallet to 50 for response size
            for p in pallet_list:
                if len(p['serials']) > 50:
                    p['serials'] = p['serials'][:50]
            
            # Build dispatch parties summary
            dispatch_parties = {}
            for ds in dispatched_serials:
                dp = ds.get('dispatch_party', '')
                dispatch_parties[dp] = dispatch_parties.get(dp, 0) + 1

            grand_dispatched += dispatched
            grand_packed += packed
            grand_disp_pending += not_packed  # Not packed = dispatch pending

            pdi_wise.append({
                'pdi_number': pdi,
                'produced': ban_gaye,
                'ftr_tested': ftr_count,
                'planned': planned,
                'pending': pending,
                'progress': progress,
                'production_days': prod_info.get('record_count', 0),
                'start_date': prod_info.get('start_date'),
                'last_date': prod_info.get('last_date'),
                'assigned_date': ftr_info.get('assigned_date'),
                'batch_status': plan_info.get('batch_status', 'N/A'),
                'dispatched': dispatched,
                'packed': packed,
                'not_packed': not_packed,
                'dispatch_pending': not_packed,  # Not packed = dispatch pending
                'dispatch_parties': [{'party': k, 'count': v} for k, v in sorted(dispatch_parties.items(), key=lambda x: x[1], reverse=True)],
                'dispatched_serials': dispatched_serials,
                'packed_serials': packed_serials,
                'not_packed_serials': not_packed_serials,
                'pallet_groups': pallet_list
            })

        grand_pending = max(0, grand_planned - grand_produced) if grand_planned > 0 else 0
        grand_progress = round((grand_produced / grand_planned) * 100, 1) if grand_planned > 0 else (100 if grand_produced > 0 else 0)

        resp = jsonify({
            "success": True,
            "company": company['company_name'],
            "order_number": order_number,
            "total_order_qty": total_order_qty,
            "total_ftr": total_ftr,
            "total_ftr_ok": total_ftr_ok,
            "total_rejected": total_rejected,
            "total_available": total_available,
            "dispatch_matches": debug_info.get('total_dispatched_serials', 0),
            "mrp_error": mrp_error,
            "debug_info": debug_info,
            "summary": {
                "total_produced": grand_produced,
                "total_planned": grand_planned,
                "total_pending": grand_pending,
                "total_ftr_assigned": grand_ftr,
                "progress": grand_progress,
                "total_dispatched": grand_dispatched,
                "total_packed": grand_packed,
                "total_not_packed": grand_disp_pending,
                "total_dispatch_pending": grand_disp_pending,
                "extra_dispatched": extra_dispatched_count,
                "extra_packed": extra_packed_count
            },
            "pdi_wise": pdi_wise,
            "extra_dispatched": {
                "count": extra_dispatched_count,
                "serials": extra_dispatched_serials,
                "pallet_groups": extra_pallet_list
            },
            "extra_packed": {
                "count": extra_packed_count,
                "serials": extra_packed_serials,
                "pallet_groups": extra_packed_pallet_list
            }
        })
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp

    except Exception as e:
        print(f"[PDI Production] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/export-not-packed-serials/<int:company_id>', methods=['GET'])
def export_not_packed_serials(company_id):
    """
    Export Not Packed serials to Excel for a company.
    Returns Excel file with serial numbers grouped by PDI.
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get company info
        cursor.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()
        if not company:
            return jsonify({"success": False, "error": "Company not found"}), 404
        
        company_name = company.get('company_name', '')
        lower_name = company_name.lower()
        
        # Get all PDI serials for this company grouped by PDI
        cursor.execute("""
            SELECT pdi_number, serial_number FROM pdi_serial_numbers 
            WHERE company_id = %s AND pdi_number IS NOT NULL AND serial_number IS NOT NULL
        """, (company_id,))
        
        pdi_serials_map = {}
        for row in cursor.fetchall():
            pdi = row['pdi_number']
            serial = row['serial_number']
            if pdi and serial and not serial.strip().startswith('20'):
                if pdi not in pdi_serials_map:
                    pdi_serials_map[pdi] = []
                pdi_serials_map[pdi].append(serial.strip())
        
        # Map company name to packing API party name
        PACKING_PARTY_NAMES = {
            'rays': 'RAYS POWER INFRA PRIVATE LIMITED',
            'l&t': 'LARSEN & TOUBRO LIMITED CONSTRUCTION',
            'larsen': 'LARSEN & TOUBRO LIMITED CONSTRUCTION',
            'lnt': 'LARSEN & TOUBRO LIMITED CONSTRUCTION',
            'sterling': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
            's&w': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
            'ntpc': 'S&W - NTPC'
        }
        
        # Find matching party names
        matching_party_names = []
        for key, party_name in PACKING_PARTY_NAMES.items():
            if key in lower_name:
                if party_name not in matching_party_names:
                    matching_party_names.append(party_name)
        
        # For Sterling/S&W, always include all sub-party names
        if any(k in lower_name for k in ['sterling', 'sterlin', 's&w']):
            for extra_name in ['STERLING AND WILSON RENEWABLE ENERGY LIMITED', 'STERLING AND WILSON', 'S&W', 'S&W - NTPC']:
                if extra_name not in matching_party_names:
                    matching_party_names.append(extra_name)
        
        # Fetch from Packing API
        packed_serials_set = set()
        for party_name in matching_party_names:
            try:
                payload = {"party_name": party_name}
                response = http_requests.post(
                    "https://umanmrp.in/api/get_barcode_tracking.php",
                    json=payload,
                    timeout=120
                )
                if response.status_code == 200:
                    data = response.json()
                    for item in data.get('data', []):
                        barcode = item.get('barcode', '').strip().upper()
                        if barcode:
                            packed_serials_set.add(barcode)
            except Exception as e:
                print(f"[Export] Packing API error ({party_name}): {e}")
        
        # Build not packed list grouped by PDI
        not_packed_by_pdi = {}
        for pdi, serials in pdi_serials_map.items():
            not_packed_list = []
            for serial in serials:
                if serial.strip().upper() not in packed_serials_set:
                    not_packed_list.append(serial)
            if not_packed_list:
                not_packed_by_pdi[pdi] = not_packed_list
        
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Not Packed Serials"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["S.No", "PDI Number", "Serial Number", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        row_num = 2
        serial_no = 1
        for pdi in sorted(not_packed_by_pdi.keys()):
            for serial in not_packed_by_pdi[pdi]:
                ws.cell(row=row_num, column=1, value=serial_no).border = thin_border
                ws.cell(row=row_num, column=2, value=pdi).border = thin_border
                ws.cell(row=row_num, column=3, value=serial).border = thin_border
                ws.cell(row=row_num, column=4, value="Not Packed").border = thin_border
                row_num += 1
                serial_no += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        
        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Company").font = Font(bold=True)
        ws2.cell(row=1, column=2, value=company_name)
        ws2.cell(row=2, column=1, value="Total PDIs").font = Font(bold=True)
        ws2.cell(row=2, column=2, value=len(not_packed_by_pdi))
        ws2.cell(row=3, column=1, value="Total Not Packed").font = Font(bold=True)
        ws2.cell(row=3, column=2, value=serial_no - 1)
        ws2.cell(row=4, column=1, value="Export Date").font = Font(bold=True)
        ws2.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
        
        # PDI-wise summary
        ws2.cell(row=6, column=1, value="PDI Number").font = Font(bold=True)
        ws2.cell(row=6, column=2, value="Not Packed Count").font = Font(bold=True)
        row = 7
        for pdi in sorted(not_packed_by_pdi.keys()):
            ws2.cell(row=row, column=1, value=pdi)
            ws2.cell(row=row, column=2, value=len(not_packed_by_pdi[pdi]))
            row += 1
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"not_packed_serials_{company_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"[Export] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ============ MRP DISPATCH CACHE ENDPOINTS ============

@ftr_bp.route('/sync-mrp-dispatch', methods=['POST'])
def sync_mrp_dispatch():
    """
    Sync dispatch data from MRP API to local cache table.
    Fetches last 1 year of data for specified company.
    Loops through pages 1-1000 to get ALL data.
    """
    try:
        data = request.get_json() or {}
        company_name = data.get('company', '')
        
        # Party ID mapping
        PARTY_IDS = {
            'Rays Power': '931db2c5-b016-4914-b378-69e9f22562a7',
            'L&T': 'a005562f-568a-46e9-bf2e-700affb171e8', 
            'Sterling': '141b81a0-2bab-4790-b825-3c8734d41484'
        }
        
        # Find party_id for company
        party_id = None
        matched_company = None
        for name, pid in PARTY_IDS.items():
            if name.lower() in company_name.lower() or company_name.lower() in name.lower():
                party_id = pid
                matched_company = name
                break
        
        if not party_id:
            return jsonify({
                "success": False, 
                "error": f"No party ID found for company: {company_name}",
                "available_companies": list(PARTY_IDS.keys())
            }), 400
        
        # Date range - last 1 year
        from datetime import datetime, timedelta
        to_date = datetime.now().strftime('%Y-%m-%d')
        from_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        print(f"[MRP Sync] Company: {company_name}, Party ID: {party_id}")
        print(f"[MRP Sync] Date range: {from_date} to {to_date}")
        
        # Fetch from MRP API - Loop pages 1 to 1000
        all_barcodes = []
        limit = 100
        empty_pages = 0
        max_empty = 3
        
        for page in range(1, 1001):
            payload = {
                "party_id": party_id,
                "from_date": from_date,
                "to_date": to_date,
                "page": page,
                "limit": limit
            }
            
            response = http_requests.post(
                "https://umanmrp.in/api/party-dispatch-history.php",
                json=payload,
                timeout=60
            )
            
            if response.status_code != 200:
                print(f"[MRP Sync] API error on page {page}: {response.status_code}")
                break
                
            result = response.json()
            dispatch_summary = result.get('dispatch_summary', [])
            
            if not dispatch_summary:
                empty_pages += 1
                if empty_pages >= max_empty:
                    print(f"[MRP Sync] Page {page}: Empty - stopping (3 consecutive empty)")
                    break
                continue
            else:
                empty_pages = 0
            
            page_count = 0
            for item in dispatch_summary:
                pallet_nos = item.get('pallet_nos', {})
                status = item.get('status', 'Packed')
                dispatch_party = item.get('dispatch_party', '')
                vehicle_no = item.get('vehicle_no', '')
                dispatch_date = item.get('dispatch_date') or item.get('date', '')
                invoice_no = item.get('invoice_no', '')
                
                # Parse pallet_nos - each key is pallet number, value is space-separated serials
                if isinstance(pallet_nos, dict):
                    for pallet_no, serials_str in pallet_nos.items():
                        if serials_str and isinstance(serials_str, str):
                            serials = serials_str.strip().split()
                            for serial in serials:
                                serial = serial.strip()
                                if serial:
                                    all_barcodes.append({
                                        'serial_number': serial.upper(),
                                        'pallet_no': pallet_no,
                                        'status': status,
                                        'dispatch_party': dispatch_party,
                                        'vehicle_no': vehicle_no,
                                        'dispatch_date': dispatch_date,
                                        'invoice_no': invoice_no,
                                        'company': matched_company,
                                        'party_id': party_id
                                    })
                                    page_count += 1
            
            print(f"[MRP Sync] Page {page}: {len(dispatch_summary)} dispatches, {page_count} serials (Total: {len(all_barcodes)})")
        
        print(f"[MRP Sync] Total barcodes fetched: {len(all_barcodes)}")
        
        # Save to local database
        conn = get_db_connection()
        inserted = 0
        updated = 0
        
        try:
            with conn.cursor() as cursor:
                for barcode in all_barcodes:
                    # Use INSERT ... ON DUPLICATE KEY UPDATE
                    cursor.execute("""
                        INSERT INTO mrp_dispatch_cache 
                        (serial_number, pallet_no, status, dispatch_party, vehicle_no, dispatch_date, invoice_no, company, party_id, synced_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON DUPLICATE KEY UPDATE
                        pallet_no = VALUES(pallet_no),
                        status = VALUES(status),
                        dispatch_party = VALUES(dispatch_party),
                        vehicle_no = VALUES(vehicle_no),
                        dispatch_date = VALUES(dispatch_date),
                        invoice_no = VALUES(invoice_no),
                        company = VALUES(company),
                        party_id = VALUES(party_id),
                        synced_at = NOW()
                    """, (
                        barcode['serial_number'],
                        barcode['pallet_no'],
                        barcode['status'],
                        barcode['dispatch_party'],
                        barcode['vehicle_no'],
                        barcode['dispatch_date'] if barcode['dispatch_date'] else None,
                        barcode['invoice_no'],
                        barcode['company'],
                        barcode['party_id']
                    ))
                    
                    if cursor.rowcount == 1:
                        inserted += 1
                    elif cursor.rowcount == 2:
                        updated += 1
                
                conn.commit()
                
                # Get total count
                cursor.execute("SELECT COUNT(*) as total FROM mrp_dispatch_cache WHERE company = %s", (matched_company,))
                total_in_db = cursor.fetchone()['total']
                
        finally:
            conn.close()
        
        return jsonify({
            "success": True,
            "company": matched_company,
            "party_id": party_id,
            "date_range": {"from": from_date, "to": to_date},
            "fetched_from_api": len(all_barcodes),
            "inserted": inserted,
            "updated": updated,
            "total_in_cache": total_in_db
        })
        
    except Exception as e:
        print(f"[MRP Sync] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/mrp-cache-stats', methods=['GET'])
def mrp_cache_stats():
    """Get statistics about MRP dispatch cache"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Get stats by company
            cursor.execute("""
                SELECT 
                    company,
                    COUNT(*) as total_serials,
                    COUNT(DISTINCT pallet_no) as total_pallets,
                    SUM(CASE WHEN status = 'Dispatched' THEN 1 ELSE 0 END) as dispatched,
                    SUM(CASE WHEN status = 'Packed' THEN 1 ELSE 0 END) as packed,
                    MIN(synced_at) as first_sync,
                    MAX(synced_at) as last_sync
                FROM mrp_dispatch_cache
                GROUP BY company
            """)
            stats = cursor.fetchall()
            
            # Get total
            cursor.execute("SELECT COUNT(*) as total FROM mrp_dispatch_cache")
            total = cursor.fetchone()['total']
            
        conn.close()
        
        return jsonify({
            "success": True,
            "total_cached": total,
            "by_company": stats
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/mrp-cache-search', methods=['GET'])
def mrp_cache_search():
    """Search serials in MRP cache"""
    try:
        serial = request.args.get('serial', '').strip().upper()
        company = request.args.get('company', '')
        pallet = request.args.get('pallet', '')
        limit = int(request.args.get('limit', 100))
        
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = "SELECT * FROM mrp_dispatch_cache WHERE 1=1"
            params = []
            
            if serial:
                query += " AND serial_number LIKE %s"
                params.append(f"%{serial}%")
            if company:
                query += " AND company LIKE %s"
                params.append(f"%{company}%")
            if pallet:
                query += " AND pallet_no LIKE %s"
                params.append(f"%{pallet}%")
            
            query += " ORDER BY synced_at DESC LIMIT %s"
            params.append(limit)
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            
        conn.close()
        
        # Convert dates to strings
        for r in results:
            if r.get('dispatch_date'):
                r['dispatch_date'] = str(r['dispatch_date'])
            if r.get('synced_at'):
                r['synced_at'] = str(r['synced_at'])
        
        return jsonify({
            "success": True,
            "count": len(results),
            "results": results
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ============================================================
# Sales Parties (from logistics.umanerp.com getSalesParty API)
# ============================================================
SALES_PARTY_API = 'https://logistics.umanerp.com/api/party/getSalesParty'
SALES_PERSON_ID = os.getenv('SALES_PERSON_ID', 'c8166f4a-0897-4239-a26d-ee42e9cee22a')

# Cache: {'data': [...parties...], 'timestamp': float}
SALES_PARTY_CACHE = {'data': None, 'timestamp': 0}
SALES_PARTY_CACHE_TTL = 600  # 10 minutes


@ftr_bp.route('/sales-parties', methods=['GET'])
def get_sales_parties():
    """
    Returns the full list of sales parties (companies) from the
    logistics.umanerp.com getSalesParty API. Cached for 10 minutes.

    Response shape:
        { "success": true, "count": N, "parties": [ {id, companyName, city, state, gst}, ... ] }
    """
    try:
        force = request.args.get('force_refresh', '').lower() == 'true'
        now = time.time()

        if (not force
                and SALES_PARTY_CACHE['data'] is not None
                and (now - SALES_PARTY_CACHE['timestamp']) < SALES_PARTY_CACHE_TTL):
            return jsonify({
                "success": True,
                "cached": True,
                "count": len(SALES_PARTY_CACHE['data']),
                "parties": SALES_PARTY_CACHE['data']
            })

        person_id = request.args.get('person_id') or SALES_PERSON_ID
        resp = http_requests.post(
            SALES_PARTY_API,
            json={"personId": person_id},
            timeout=60
        )

        if resp.status_code != 200:
            return jsonify({
                "success": False,
                "error": f"Sales party API returned HTTP {resp.status_code}",
                "detail": resp.text[:300]
            }), 502

        payload = resp.json()
        raw_parties = payload.get('data') or []

        parties = []
        for p in raw_parties:
            party_id = p.get('PartyNameId')
            name = (p.get('PartyName') or '').strip()
            if not party_id or not name:
                continue
            parties.append({
                "id": party_id,
                "companyName": name,
                "city": p.get('City') or '',
                "state": p.get('State') or '',
                "gst": p.get('GSTNo') or '',
                "status": p.get('Status') or ''
            })

        # Sort alphabetically
        parties.sort(key=lambda x: x['companyName'].lower())

        SALES_PARTY_CACHE['data'] = parties
        SALES_PARTY_CACHE['timestamp'] = now

        return jsonify({
            "success": True,
            "cached": False,
            "count": len(parties),
            "parties": parties
        })

    except http_requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "Sales party API timeout"}), 504
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================================
# MRP PROXY ENDPOINTS (avoid browser CORS)
# ============================================================

@ftr_bp.route('/mrp-party-pdis', methods=['POST', 'GET'])
def mrp_party_pdis():
    """Proxy for https://umanmrp.in/get/get_all_pdi.php

    Body / query: { party_name_id: '<uuid>' }
    Returns the upstream JSON as-is on success.
    """
    try:
        if request.method == 'POST':
            payload = request.get_json(silent=True) or {}
            party_name_id = (payload.get('party_name_id') or '').strip()
        else:
            party_name_id = (request.args.get('party_name_id') or '').strip()

        if not party_name_id:
            return jsonify({"status": "error", "message": "party_name_id is required"}), 400

        resp = http_requests.post(
            'https://umanmrp.in/get/get_all_pdi.php',
            json={"party_name_id": party_name_id},
            timeout=60
        )
        try:
            data = resp.json()
        except Exception:
            return jsonify({
                "status": "error",
                "message": f"Upstream returned non-JSON (HTTP {resp.status_code})",
                "detail": resp.text[:300]
            }), 502
        return jsonify(data), resp.status_code
    except http_requests.exceptions.Timeout:
        return jsonify({"status": "error", "message": "Upstream timeout"}), 504
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@ftr_bp.route('/mrp-pdi-barcodes', methods=['POST', 'GET'])
def mrp_pdi_barcodes():
    """Proxy for https://mrp.umanerp.com/get/get_pdi_barcodes.php

    Body / query: { pdi_id: '<id>' }
    """
    try:
        if request.method == 'POST':
            payload = request.get_json(silent=True) or {}
            pdi_id = str(payload.get('pdi_id') or '').strip()
        else:
            pdi_id = str(request.args.get('pdi_id') or '').strip()

        if not pdi_id:
            return jsonify({"status": "error", "message": "pdi_id is required"}), 400

        resp = http_requests.post(
            'https://mrp.umanerp.com/get/get_pdi_barcodes.php',
            json={"pdi_id": pdi_id},
            timeout=120
        )
        try:
            data = resp.json()
        except Exception:
            return jsonify({
                "status": "error",
                "message": f"Upstream returned non-JSON (HTTP {resp.status_code})",
                "detail": resp.text[:300]
            }), 502
        return jsonify(data), resp.status_code
    except http_requests.exceptions.Timeout:
        return jsonify({"status": "error", "message": "Upstream timeout"}), 504
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@ftr_bp.route('/pdi-status/<pdi_id>', methods=['GET'])
def pdi_status(pdi_id):
    """Rays-style per-PDI status using BULK party-dispatch-history API
    (same concept as /dispatch-tracking and /dispatch-by-party).

    1. Fetch all barcodes of this PDI from MRP get_pdi_barcodes.php
    2. Fetch the party's full dispatch history in bulk (one API, paginated)
       from umanmrp.in/api/party-dispatch-history.php
    3. Intersect: which PDI barcodes are in dispatch history -> dispatched,
       rest -> pending. Also group dispatched by vehicle/pallet.

    Query params:
        - party_id : REQUIRED, party UUID (same as used by /dispatch-by-party)
        - days     : dispatch window, default 730
        - force    : '1' to bypass cache

    Cached per (pdi_id, party_id) for 5 minutes.
    """
    pdi_id = str(pdi_id or '').strip()
    if not pdi_id:
        return jsonify({"success": False, "error": "pdi_id is required"}), 400

    party_id = (request.args.get('party_id') or '').strip()
    if not party_id or len(party_id) < 10:
        return jsonify({"success": False, "error": "party_id query param is required"}), 400

    try:
        days = int(request.args.get('days', '730'))
    except Exception:
        days = 730
    force = request.args.get('force', '').lower() in ('1', 'true', 'yes')

    cache = pdi_status.__dict__.setdefault('_cache', {})
    cache_key = f"{pdi_id}|{party_id}|{days}"
    now = time.time()
    if not force and cache_key in cache:
        entry = cache[cache_key]
        if (now - entry['timestamp']) < 300:
            return jsonify({**entry['data'], "cached": True})

    # ===== 1. Get PDI barcodes =====
    try:
        barc_resp = http_requests.post(
            'https://mrp.umanerp.com/get/get_pdi_barcodes.php',
            json={"pdi_id": pdi_id},
            timeout=120
        )
        barc_data = barc_resp.json() if barc_resp.status_code == 200 else {}
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to fetch PDI barcodes: {e}"}), 502

    if barc_data.get('status') != 'success':
        return jsonify({
            "success": False,
            "error": barc_data.get('message', 'Upstream PDI barcode fetch failed')
        }), 502

    pdi_details = barc_data.get('pdi_details') or {}
    raw_barcodes = barc_data.get('barcodes') or []
    pdi_barcode_set = {str(b).strip().upper() for b in raw_barcodes if str(b).strip()}
    total_pdi = len(pdi_barcode_set)

    # ===== 2. Fetch party dispatch history (bulk, paginated) =====
    from datetime import timedelta
    to_date = datetime.now().strftime('%Y-%m-%d')
    from_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

    mrp_lookup = {}
    page = 1
    max_pages = 50
    while page <= max_pages:
        try:
            resp = http_requests.post(
                'https://umanmrp.in/api/party-dispatch-history.php',
                json={
                    'party_id': party_id,
                    'from_date': from_date,
                    'to_date': to_date,
                    'page': page,
                    'limit': 10000
                },
                timeout=120,
                headers={'Cache-Control': 'no-cache', 'Pragma': 'no-cache'}
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            dispatch_summary = data.get('dispatch_summary', [])
            if not dispatch_summary:
                break
            for d in dispatch_summary:
                dispatch_date = d.get('dispatch_date') or d.get('date', '')
                vehicle_no = d.get('vehicle_no', '') or 'Unknown'
                invoice_no = d.get('invoice_no', '')
                factory_name = d.get('factory_name', '')
                dispatch_party = d.get('dispatch_party', '') or vehicle_no
                pallet_nos = d.get('pallet_nos', {})
                if isinstance(pallet_nos, dict):
                    for pallet_no, barcodes_str in pallet_nos.items():
                        if not isinstance(barcodes_str, str):
                            continue
                        for serial in barcodes_str.strip().split():
                            s = serial.strip().upper()
                            if not s:
                                continue
                            mrp_lookup[s] = {
                                'pallet_no': pallet_no,
                                'dispatch_party': dispatch_party,
                                'vehicle_no': vehicle_no,
                                'dispatch_date': dispatch_date,
                                'invoice_no': invoice_no,
                                'factory_name': factory_name
                            }
            page += 1
        except Exception as e:
            print(f"[PDI Status] page {page} error: {e}")
            break

    # ===== 3. Intersect =====
    dispatched_set = pdi_barcode_set & set(mrp_lookup.keys())
    not_dispatched_set = pdi_barcode_set - dispatched_set

    # ===== 3b. Detect packed (not dispatched) via per-barcode tracking API =====
    # Cache per barcode for 30 min to avoid re-hitting upstream
    from concurrent.futures import ThreadPoolExecutor, as_completed
    PACK_CAP = int(request.args.get('pack_cap', 3000))
    pack_cache = pdi_status.__dict__.setdefault('_pack_cache', {})
    PACK_TTL = 1800

    packed_set = set()
    pending_set = set()
    packed_info = {}
    pack_unknown = 0

    not_disp_list = list(not_dispatched_set)
    to_check = []
    for s in not_disp_list[:PACK_CAP]:
        entry = pack_cache.get(s)
        if entry and (now - entry['t']) < PACK_TTL:
            if entry['status'] == 'packed':
                packed_set.add(s)
                packed_info[s] = entry.get('info') or {}
            else:
                pending_set.add(s)
        else:
            to_check.append(s)

    def _check_pack(serial):
        try:
            r = http_requests.post(
                'https://umanmrp.in/api/get_barcode_tracking.php',
                data={'barcode': serial}, timeout=8
            )
            if r.status_code != 200:
                return ('unknown', serial, None)
            d = r.json()
            if not d.get('success') or not d.get('data'):
                return ('pending', serial, None)
            pack = (d['data'] or {}).get('packing') or {}
            if pack.get('packing_date'):
                return ('packed', serial, {
                    'packing_date': pack.get('packing_date'),
                    'box_no': pack.get('box_no', ''),
                    'pallet_no': pack.get('pallet_no', '')
                })
            return ('pending', serial, None)
        except Exception:
            return ('unknown', serial, None)

    if to_check:
        with ThreadPoolExecutor(max_workers=30) as ex:
            for f in as_completed([ex.submit(_check_pack, s) for s in to_check]):
                status, serial, info = f.result()
                if status == 'packed':
                    packed_set.add(serial)
                    packed_info[serial] = info or {}
                    pack_cache[serial] = {'t': now, 'status': 'packed', 'info': info}
                elif status == 'pending':
                    pending_set.add(serial)
                    pack_cache[serial] = {'t': now, 'status': 'pending'}
                else:
                    pack_unknown += 1

    # Anything beyond cap counted as unknown_pack
    skipped_pack_check = max(0, len(not_disp_list) - PACK_CAP)
    if skipped_pack_check:
        # treat skipped as pending (best guess) — UI shows the cap
        pending_set.update(not_disp_list[PACK_CAP:])

    # Group dispatched by vehicle and pallet
    vehicle_map = {}
    pallet_map = {}
    for serial in dispatched_set:
        info = mrp_lookup[serial]
        vehicle = info.get('vehicle_no') or 'Unknown'
        pallet_no = info.get('pallet_no', '')
        if vehicle not in vehicle_map:
            vehicle_map[vehicle] = {
                'vehicle_no': vehicle,
                'dispatch_date': info.get('dispatch_date', ''),
                'invoice_no': info.get('invoice_no', ''),
                'factory_name': info.get('factory_name', ''),
                'module_count': 0,
                'pallets': set(),
                'serials': []
            }
        vehicle_map[vehicle]['module_count'] += 1
        if pallet_no:
            vehicle_map[vehicle]['pallets'].add(pallet_no)
        if len(vehicle_map[vehicle]['serials']) < 50:
            vehicle_map[vehicle]['serials'].append(serial)

        if pallet_no:
            if pallet_no not in pallet_map:
                pallet_map[pallet_no] = {
                    'pallet_no': pallet_no,
                    'vehicle_no': vehicle,
                    'dispatch_date': info.get('dispatch_date', ''),
                    'module_count': 0,
                    'serials': []
                }
            pallet_map[pallet_no]['module_count'] += 1
            if len(pallet_map[pallet_no]['serials']) < 20:
                pallet_map[pallet_no]['serials'].append(serial)

    dispatch_groups = []
    for v, vd in vehicle_map.items():
        dispatch_groups.append({
            'vehicle_no': vd['vehicle_no'],
            'dispatch_date': vd['dispatch_date'],
            'invoice_no': vd['invoice_no'],
            'factory_name': vd['factory_name'],
            'module_count': vd['module_count'],
            'pallet_count': len(vd['pallets']),
            'pallets': sorted(list(vd['pallets'])),
            'serials': vd['serials']
        })
    dispatch_groups.sort(key=lambda x: x['module_count'], reverse=True)
    pallet_groups = sorted(pallet_map.values(), key=lambda x: str(x['pallet_no']))

    dispatched = len(dispatched_set)
    packed = len(packed_set)
    pending = len(pending_set)
    dispatched_pct = round((dispatched / total_pdi * 100), 1) if total_pdi else 0
    packed_pct = round((packed / total_pdi * 100), 1) if total_pdi else 0
    pending_pct = round((pending / total_pdi * 100), 1) if total_pdi else 0

    try:
        wattage_num = float(pdi_details.get('wattage') or 0)
    except Exception:
        wattage_num = 0
    total_kw = round((wattage_num * total_pdi) / 1000.0, 2)

    payload = {
        "success": True,
        "cached": False,
        "pdi": {
            "id": pdi_id,
            "name": pdi_details.get('pdi_name', ''),
            "wattage": pdi_details.get('wattage', ''),
            "quantity": int(pdi_details.get('quantity') or 0),
            "total_kw": total_kw
        },
        "party_id": party_id,
        "summary": {
            "total_barcodes": total_pdi,
            "dispatched": dispatched,
            "packed": packed,
            "pending": pending,
            "dispatched_percent": dispatched_pct,
            "packed_percent": packed_pct,
            "pending_percent": pending_pct,
            "party_dispatch_universe": len(mrp_lookup),
            "pack_check_capped_at": PACK_CAP,
            "pack_skipped_due_to_cap": skipped_pack_check,
            "pack_unknown": pack_unknown
        },
        "dispatch_groups": dispatch_groups,
        "pallet_groups": pallet_groups,
        "packed_sample": [{
            "serial": s,
            "packing_date": (packed_info.get(s) or {}).get('packing_date', ''),
            "box_no": (packed_info.get(s) or {}).get('box_no', ''),
            "pallet_no": (packed_info.get(s) or {}).get('pallet_no', '')
        } for s in list(packed_set)[:200]],
        "pending_sample": list(pending_set)[:200],
        "all_dispatched": sorted(dispatched_set),
        "all_packed": sorted(packed_set),
        "all_pending": sorted(pending_set)
    }

    cache[cache_key] = {"timestamp": now, "data": payload}
    return jsonify(payload)


# ---------------------------------------------------------------------------
# Actual PDI Barcodes — independent storage (no FK to PDI card)
# Used only for generating the actual-compare report. One row per pdi_id;
# new upload replaces previous.
# ---------------------------------------------------------------------------

def _ensure_actual_pdi_barcodes_table(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS actual_pdi_barcodes (
            pdi_id          VARCHAR(64) PRIMARY KEY,
            party_name      VARCHAR(255) DEFAULT NULL,
            filename        VARCHAR(255) DEFAULT NULL,
            barcode_count   INT DEFAULT 0,
            barcodes_json   LONGTEXT,
            uploaded_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at      DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)


@ftr_bp.route('/actual-pdi-barcodes/save', methods=['POST'])
def save_actual_pdi_barcodes():
    """Save / replace actual PDI barcodes for a pdi_id (independent storage)."""
    try:
        body = request.get_json(silent=True) or {}
        pdi_id = str(body.get('pdi_id') or '').strip()
        party_name = str(body.get('party_name') or '').strip() or None
        filename = str(body.get('filename') or '').strip() or None
        barcodes = body.get('barcodes') or []
        if not pdi_id:
            return jsonify({"success": False, "error": "pdi_id required"}), 400
        if not isinstance(barcodes, list):
            return jsonify({"success": False, "error": "barcodes must be a list"}), 400

        cleaned = sorted({str(b).strip().upper() for b in barcodes if str(b).strip()})
        payload = json.dumps(cleaned, ensure_ascii=False)

        conn = get_db_connection()
        cursor = conn.cursor()
        _ensure_actual_pdi_barcodes_table(cursor)
        cursor.execute("""
            INSERT INTO actual_pdi_barcodes (pdi_id, party_name, filename, barcode_count, barcodes_json)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                party_name = VALUES(party_name),
                filename = VALUES(filename),
                barcode_count = VALUES(barcode_count),
                barcodes_json = VALUES(barcodes_json)
        """, (pdi_id, party_name, filename, len(cleaned), payload))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "pdi_id": pdi_id, "count": len(cleaned)})
    except Exception as e:
        print(f"[save_actual_pdi_barcodes] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-barcodes/<pdi_id>', methods=['GET'])
def get_actual_pdi_barcodes(pdi_id):
    """Fetch saved actual PDI barcodes for a pdi_id."""
    try:
        pdi_id = str(pdi_id).strip()
        if not pdi_id:
            return jsonify({"success": False, "error": "pdi_id required"}), 400
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        _ensure_actual_pdi_barcodes_table(cursor)
        cursor.execute("""
            SELECT pdi_id, party_name, filename, barcode_count, barcodes_json,
                   uploaded_at, updated_at
            FROM actual_pdi_barcodes WHERE pdi_id = %s
        """, (pdi_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({"success": True, "exists": False, "barcodes": []})
        try:
            bcs = json.loads(row.get('barcodes_json') or '[]')
        except Exception:
            bcs = []
        return jsonify({
            "success": True,
            "exists": True,
            "pdi_id": row.get('pdi_id'),
            "party_name": row.get('party_name'),
            "filename": row.get('filename'),
            "count": row.get('barcode_count') or len(bcs),
            "barcodes": bcs,
            "uploaded_at": str(row.get('uploaded_at') or ''),
            "updated_at": str(row.get('updated_at') or '')
        })
    except Exception as e:
        print(f"[get_actual_pdi_barcodes] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-barcodes/<pdi_id>', methods=['DELETE'])
def delete_actual_pdi_barcodes(pdi_id):
    """Delete saved actual PDI barcodes for a pdi_id."""
    try:
        pdi_id = str(pdi_id).strip()
        if not pdi_id:
            return jsonify({"success": False, "error": "pdi_id required"}), 400
        conn = get_db_connection()
        cursor = conn.cursor()
        _ensure_actual_pdi_barcodes_table(cursor)
        cursor.execute("DELETE FROM actual_pdi_barcodes WHERE pdi_id = %s", (pdi_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print(f"[delete_actual_pdi_barcodes] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Actual PDI BATCHES — party-level multi-batch storage (PDI 1, PDI 2, ... PDI N)
# Each party can have many actual PDI batches. Each batch is independent of
# the planned PDI cards — used only for compare reports.
# ---------------------------------------------------------------------------

def _ensure_actual_pdi_batches_table(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS actual_pdi_batches (
            id              INT AUTO_INCREMENT PRIMARY KEY,
            party_id        VARCHAR(64) NOT NULL,
            party_name      VARCHAR(255) DEFAULT NULL,
            batch_no        INT NOT NULL,
            batch_name      VARCHAR(255) DEFAULT NULL,
            filename        VARCHAR(255) DEFAULT NULL,
            barcode_count   INT DEFAULT 0,
            barcodes_json   LONGTEXT,
            created_at      DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at      DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_party_batch (party_id, batch_no),
            KEY idx_party (party_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)


@ftr_bp.route('/actual-pdi-batches/<party_id>', methods=['GET'])
def list_actual_pdi_batches(party_id):
    """List all actual PDI batches for a party (without barcodes for speed)."""
    try:
        party_id = str(party_id).strip()
        if not party_id:
            return jsonify({"success": False, "error": "party_id required"}), 400
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        _ensure_actual_pdi_batches_table(cursor)
        cursor.execute("""
            SELECT id, party_id, party_name, batch_no, batch_name, filename,
                   barcode_count, created_at, updated_at
            FROM actual_pdi_batches WHERE party_id = %s
            ORDER BY batch_no ASC
        """, (party_id,))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        batches = []
        for r in rows:
            batches.append({
                "id": r.get('id'),
                "party_id": r.get('party_id'),
                "party_name": r.get('party_name'),
                "batch_no": r.get('batch_no'),
                "batch_name": r.get('batch_name') or f"PDI {r.get('batch_no')}",
                "filename": r.get('filename') or '',
                "count": r.get('barcode_count') or 0,
                "created_at": str(r.get('created_at') or ''),
                "updated_at": str(r.get('updated_at') or '')
            })
        return jsonify({"success": True, "batches": batches})
    except Exception as e:
        print(f"[list_actual_pdi_batches] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-batches/<party_id>', methods=['POST'])
def create_actual_pdi_batch(party_id):
    """Create a new actual PDI batch for a party. Auto-assigns batch_no.

    Body: { party_name, batch_name?, filename?, barcodes:[] }
    """
    try:
        party_id = str(party_id).strip()
        if not party_id:
            return jsonify({"success": False, "error": "party_id required"}), 400
        body = request.get_json(silent=True) or {}
        party_name = str(body.get('party_name') or '').strip() or None
        batch_name = str(body.get('batch_name') or '').strip() or None
        filename = str(body.get('filename') or '').strip() or None
        barcodes = body.get('barcodes') or []
        if not isinstance(barcodes, list):
            return jsonify({"success": False, "error": "barcodes must be a list"}), 400
        cleaned = sorted({str(b).strip().upper() for b in barcodes if str(b).strip()})
        payload = json.dumps(cleaned, ensure_ascii=False)

        conn = get_db_connection()
        cursor = conn.cursor()
        _ensure_actual_pdi_batches_table(cursor)
        cursor.execute("SELECT COALESCE(MAX(batch_no),0)+1 FROM actual_pdi_batches WHERE party_id=%s", (party_id,))
        next_no = cursor.fetchone()[0] or 1
        if not batch_name:
            batch_name = f"PDI {next_no}"
        cursor.execute("""
            INSERT INTO actual_pdi_batches
                (party_id, party_name, batch_no, batch_name, filename, barcode_count, barcodes_json)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (party_id, party_name, next_no, batch_name, filename, len(cleaned), payload))
        new_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "id": new_id, "batch_no": next_no, "batch_name": batch_name, "count": len(cleaned)})
    except Exception as e:
        print(f"[create_actual_pdi_batch] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-batches/<party_id>/<int:batch_id>', methods=['GET'])
def get_actual_pdi_batch(party_id, batch_id):
    """Fetch a single batch with its barcodes."""
    try:
        party_id = str(party_id).strip()
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        _ensure_actual_pdi_batches_table(cursor)
        cursor.execute("""
            SELECT * FROM actual_pdi_batches WHERE party_id=%s AND id=%s
        """, (party_id, batch_id))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({"success": False, "error": "batch not found"}), 404
        try:
            bcs = json.loads(row.get('barcodes_json') or '[]')
        except Exception:
            bcs = []
        return jsonify({
            "success": True,
            "id": row.get('id'),
            "party_id": row.get('party_id'),
            "party_name": row.get('party_name'),
            "batch_no": row.get('batch_no'),
            "batch_name": row.get('batch_name'),
            "filename": row.get('filename') or '',
            "count": row.get('barcode_count') or len(bcs),
            "barcodes": bcs,
            "created_at": str(row.get('created_at') or ''),
            "updated_at": str(row.get('updated_at') or '')
        })
    except Exception as e:
        print(f"[get_actual_pdi_batch] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-batches/<party_id>/<int:batch_id>', methods=['PUT'])
def update_actual_pdi_batch(party_id, batch_id):
    """Update batch barcodes / name / filename. Body: any of {batch_name, filename, barcodes}."""
    try:
        body = request.get_json(silent=True) or {}
        sets, params = [], []
        if 'batch_name' in body:
            sets.append("batch_name=%s"); params.append(str(body.get('batch_name') or '').strip() or None)
        if 'filename' in body:
            sets.append("filename=%s"); params.append(str(body.get('filename') or '').strip() or None)
        if 'barcodes' in body:
            barcodes = body.get('barcodes') or []
            if not isinstance(barcodes, list):
                return jsonify({"success": False, "error": "barcodes must be a list"}), 400
            cleaned = sorted({str(b).strip().upper() for b in barcodes if str(b).strip()})
            sets.append("barcodes_json=%s"); params.append(json.dumps(cleaned, ensure_ascii=False))
            sets.append("barcode_count=%s"); params.append(len(cleaned))
        if not sets:
            return jsonify({"success": False, "error": "nothing to update"}), 400
        params.extend([str(party_id).strip(), batch_id])
        conn = get_db_connection()
        cursor = conn.cursor()
        _ensure_actual_pdi_batches_table(cursor)
        cursor.execute(f"UPDATE actual_pdi_batches SET {', '.join(sets)} WHERE party_id=%s AND id=%s", params)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print(f"[update_actual_pdi_batch] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-batches/<party_id>/<int:batch_id>', methods=['DELETE'])
def delete_actual_pdi_batch(party_id, batch_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _ensure_actual_pdi_batches_table(cursor)
        cursor.execute("DELETE FROM actual_pdi_batches WHERE party_id=%s AND id=%s", (str(party_id).strip(), batch_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print(f"[delete_actual_pdi_batch] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/actual-pdi-batch-compare', methods=['POST'])
def actual_pdi_batch_compare():
    """Compare a single Actual PDI batch's barcodes against ALL planned PDI cards
    of the party + party packing/dispatch data. Returns per-card breakdown.

    Body: { party_id, party_name?, barcodes:[] }
    Output: {
      summary: { total_actual, matched_in_any_card, extras_no_card, packed, dispatched, pending },
      card_breakdown: [ { pdi_id, pdi_name, plan_qty, actual_in_card, packed, dispatched, pending } ],
      packed_sample, pending_sample, extras_sample, dispatch_breakdown, running_order_breakdown
    }
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    try:
        body = request.get_json(silent=True) or {}
        party_id = str(body.get('party_id') or '').strip()
        party_name = str(body.get('party_name') or '').strip()
        actual_raw = body.get('barcodes') or []
        if not party_id:
            return jsonify({"success": False, "error": "party_id required"}), 400
        if not isinstance(actual_raw, list):
            return jsonify({"success": False, "error": "barcodes must be a list"}), 400

        actual_set = {str(b).strip().upper() for b in actual_raw if str(b).strip()}
        if not actual_set:
            return jsonify({"success": False, "error": "barcodes empty"}), 400

        # 1. Get all PDIs of party
        pdi_list_resp = http_requests.post(
            'https://umanmrp.in/get/get_all_pdi.php',
            json={"party_name_id": party_id}, timeout=60
        )
        all_pdis = (pdi_list_resp.json() if pdi_list_resp.status_code == 200 else {}).get('data') or []

        # 2. Fetch barcodes per PDI in parallel
        cache = actual_pdi_batch_compare.__dict__.setdefault('_pdi_bc_cache', {})
        now = time.time()

        def _fetch(pid):
            ent = cache.get(pid)
            if ent and (now - ent['t']) < 600:
                return pid, ent['barcodes'], ent['details']
            try:
                r = http_requests.post(
                    'https://mrp.umanerp.com/get/get_pdi_barcodes.php',
                    json={"pdi_id": str(pid)}, timeout=120
                )
                d = r.json() if r.status_code == 200 else {}
                bcs = [str(x).strip().upper() for x in (d.get('barcodes') or []) if str(x).strip()]
                det = d.get('pdi_details') or {}
                cache[pid] = {'t': now, 'barcodes': bcs, 'details': det}
                return pid, bcs, det
            except Exception:
                return pid, [], {}

        pdi_barcode_map = {}
        pdi_detail_map = {}
        serial_to_pdi = {}
        with ThreadPoolExecutor(max_workers=8) as ex:
            futures = [ex.submit(_fetch, item.get('id')) for item in all_pdis if item.get('id')]
            for f in as_completed(futures):
                pid, bcs, det = f.result()
                bset = set(bcs)
                pdi_barcode_map[pid] = bset
                pdi_detail_map[pid] = det
                for s in bset:
                    if s not in serial_to_pdi:
                        serial_to_pdi[s] = pid

        # 3. Bulk packing (party_name based)
        packed_lookup = {}
        if party_name:
            try:
                pr = http_requests.post(
                    'https://umanmrp.in/api/get_barcode_tracking.php',
                    json={'party_name': party_name}, timeout=120
                )
                if pr.status_code == 200:
                    for it in (pr.json().get('data') or []):
                        bc = str(it.get('barcode','') or '').strip().upper()
                        if bc:
                            packed_lookup[bc] = {
                                'pallet_no': it.get('pallet_no',''),
                                'running_order': it.get('running_order','') or it.get('ro_no','') or '',
                                'packed_party': it.get('packed_party','') or it.get('party_name','') or party_name,
                                'packing_date': it.get('packing_date',''),
                                'box_no': it.get('box_no','')
                            }
            except Exception as e:
                print(f"[batch_compare] pack fetch error: {e}")

        # 4. Bulk dispatch (party_name based)
        dispatch_lookup = {}
        if party_name:
            try:
                dr = http_requests.post(
                    'https://umanmrp.in/api/party-dispatch-history.php',
                    json={'party_name': party_name}, timeout=120
                )
                if dr.status_code == 200:
                    for d in (dr.json().get('data') or []):
                        bc = str(d.get('barcode','') or '').strip().upper()
                        if bc:
                            dispatch_lookup[bc] = {
                                'vehicle_no': d.get('vehicle_no',''),
                                'dispatch_date': d.get('dispatch_date',''),
                                'dispatch_party': d.get('dispatch_party','') or party_name,
                                'invoice_no': d.get('invoice_no','')
                            }
            except Exception as e:
                print(f"[batch_compare] dispatch fetch error: {e}")

        # 5. Bucket actual barcodes
        matched_any = set()
        extras = set()
        per_card = {}  # pid -> set of actual barcodes
        for s in actual_set:
            pid = serial_to_pdi.get(s)
            if pid:
                matched_any.add(s)
                per_card.setdefault(pid, set()).add(s)
            else:
                extras.add(s)

        packed_set = {s for s in actual_set if s in packed_lookup}
        dispatched_set = {s for s in actual_set if s in dispatch_lookup}
        pending_set = actual_set - packed_set - dispatched_set

        # 6. Per-card breakdown
        card_breakdown = []
        for item in all_pdis:
            pid = item.get('id')
            if not pid:
                continue
            actual_in_card = per_card.get(pid, set())
            packed_in_card = {s for s in actual_in_card if s in packed_lookup}
            disp_in_card = {s for s in actual_in_card if s in dispatch_lookup}
            pending_in_card = actual_in_card - packed_in_card - disp_in_card
            det = pdi_detail_map.get(pid) or {}
            card_breakdown.append({
                "pdi_id": pid,
                "pdi_name": item.get('pdi_name') or det.get('pdi_name') or f"PDI {pid}",
                "wattage": item.get('wattage') or det.get('wattage') or '',
                "plan_qty": int(item.get('quantity') or det.get('quantity') or 0),
                "card_total_barcodes": len(pdi_barcode_map.get(pid, set())),
                "actual_in_card": len(actual_in_card),
                "packed": len(packed_in_card),
                "dispatched": len(disp_in_card),
                "pending": len(pending_in_card)
            })
        card_breakdown.sort(key=lambda x: -x['actual_in_card'])

        # 7. Running order breakdown (from packed_lookup)
        ro_break = {}
        for s in actual_set:
            info = packed_lookup.get(s)
            if info:
                ro = info.get('running_order') or 'Unknown'
                ro_break.setdefault(ro, 0)
                ro_break[ro] += 1
        ro_breakdown = [{"running_order": k, "count": v} for k, v in sorted(ro_break.items(), key=lambda x: -x[1])]

        # 8. Dispatch vehicle breakdown
        veh_break = {}
        for s in actual_set:
            info = dispatch_lookup.get(s)
            if info:
                key = (info.get('vehicle_no') or 'Unknown', info.get('dispatch_party') or party_name)
                veh_break.setdefault(key, 0)
                veh_break[key] += 1
        dispatch_breakdown = [
            {"vehicle_no": k[0], "dispatch_party": k[1], "count": v}
            for k, v in sorted(veh_break.items(), key=lambda x: -x[1])
        ]

        return jsonify({
            "success": True,
            "summary": {
                "total_actual": len(actual_set),
                "matched_in_any_card": len(matched_any),
                "extras_no_card": len(extras),
                "packed": len(packed_set),
                "dispatched": len(dispatched_set),
                "pending": len(pending_set)
            },
            "card_breakdown": card_breakdown,
            "running_order_breakdown": ro_breakdown,
            "dispatch_breakdown": dispatch_breakdown,
            "packed_sample": [
                {"serial": s, **(packed_lookup.get(s) or {})} for s in list(packed_set)[:300]
            ],
            "pending_sample": list(pending_set)[:300],
            "extras_sample": list(extras)[:300],
            "all_packed": sorted(packed_set),
            "all_pending": sorted(pending_set),
            "all_extras": sorted(extras),
            "all_dispatched": sorted(dispatched_set)
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"[actual_pdi_batch_compare] error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/pdi-actual-compare', methods=['POST'])
def pdi_actual_compare():
    """Compare actual PDI barcodes (uploaded by user) against:
      - the planned PDI barcodes (this pdi_id)
      - all other PDIs of the same party (to identify extras)
      - party dispatch history (dispatched vs not)

    Request JSON:
        {
          "pdi_id": "76",
          "party_id": "<uuid>",
          "actual_barcodes": ["GS04...", ...]
        }

    Returns per-PDI breakdown + matched/missing/extras + dispatched/packed status.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    try:
        body = request.get_json(silent=True) or {}
        pdi_id = str(body.get('pdi_id') or '').strip()
        party_id = str(body.get('party_id') or '').strip()
        party_name = str(body.get('party_name') or '').strip()
        actual_raw = body.get('actual_barcodes') or []
        if not pdi_id or not party_id:
            return jsonify({"success": False, "error": "pdi_id and party_id required"}), 400
        if not isinstance(actual_raw, list):
            return jsonify({"success": False, "error": "actual_barcodes must be a list"}), 400

        actual_set = {str(b).strip().upper() for b in actual_raw if str(b).strip()}

        # 1. All PDIs of this party
        pdi_list_resp = http_requests.post(
            'https://umanmrp.in/get/get_all_pdi.php',
            json={"party_name_id": party_id},
            timeout=60
        )
        pdi_list_json = pdi_list_resp.json() if pdi_list_resp.status_code == 200 else {}
        all_pdis = pdi_list_json.get('data') or []

        # 2. Fetch barcodes of EACH pdi in parallel; cache (reuse pdi_status pack_cache-like cache)
        pdi_bc_cache = pdi_actual_compare.__dict__.setdefault('_pdi_bc_cache', {})
        now = time.time()

        def _fetch_pdi_barcodes(pid):
            ent = pdi_bc_cache.get(pid)
            if ent and (now - ent['t']) < 600:
                return pid, ent['barcodes'], ent['details']
            try:
                r = http_requests.post(
                    'https://mrp.umanerp.com/get/get_pdi_barcodes.php',
                    json={"pdi_id": str(pid)}, timeout=120
                )
                d = r.json() if r.status_code == 200 else {}
                bcs = [str(x).strip().upper() for x in (d.get('barcodes') or []) if str(x).strip()]
                det = d.get('pdi_details') or {}
                pdi_bc_cache[pid] = {'t': now, 'barcodes': bcs, 'details': det}
                return pid, bcs, det
            except Exception:
                return pid, [], {}

        pdi_barcode_map = {}   # pid -> set
        pdi_detail_map = {}    # pid -> {pdi_name, wattage, quantity}
        serial_to_pdi = {}     # barcode -> pid

        with ThreadPoolExecutor(max_workers=8) as ex:
            futures = [ex.submit(_fetch_pdi_barcodes, item.get('id')) for item in all_pdis if item.get('id')]
            for f in as_completed(futures):
                pid, bcs, det = f.result()
                s = set(bcs)
                pdi_barcode_map[str(pid)] = s
                pdi_detail_map[str(pid)] = {
                    'pdi_name': det.get('pdi_name') or '',
                    'wattage': det.get('wattage') or '',
                    'quantity': int(det.get('quantity') or 0)
                }
                for b in bcs:
                    # Keep first owner seen (unique in practice)
                    if b not in serial_to_pdi:
                        serial_to_pdi[b] = str(pid)

        planned_set = pdi_barcode_map.get(pdi_id, set())
        planned_total = len(planned_set)
        actual_total = len(actual_set)

        # 3. Buckets
        matched = planned_set & actual_set                # in plan and actually delivered
        missing = planned_set - actual_set                # planned but not delivered
        extras = actual_set - planned_set                 # delivered but not in plan
        unknown_extras = {b for b in extras if b not in serial_to_pdi}

        # Group extras by owning PDI
        extras_by_pdi = {}
        for b in extras:
            owner = serial_to_pdi.get(b, '__unknown__')
            extras_by_pdi.setdefault(owner, []).append(b)

        extras_breakdown = []
        for owner_id, serials in extras_by_pdi.items():
            if owner_id == '__unknown__':
                extras_breakdown.append({
                    'pdi_id': None,
                    'pdi_name': 'Unknown / Not in any party PDI',
                    'count': len(serials),
                    'serials_sample': serials[:50]
                })
            else:
                det = pdi_detail_map.get(owner_id, {})
                extras_breakdown.append({
                    'pdi_id': owner_id,
                    'pdi_name': det.get('pdi_name') or f'PDI {owner_id}',
                    'count': len(serials),
                    'serials_sample': serials[:50]
                })
        extras_breakdown.sort(key=lambda x: x['count'], reverse=True)

        # Missing breakdown: for each missing barcode, is it sitting under any OTHER pdi?
        # (Should mostly be under the same planned PDI, but we show running-order/other-pdi hits.)
        missing_other_owners = {}
        for b in missing:
            owner = serial_to_pdi.get(b)
            if owner and owner != pdi_id:
                missing_other_owners.setdefault(owner, []).append(b)

        # 4. Party dispatch history (bulk)
        from datetime import timedelta
        to_date = datetime.now().strftime('%Y-%m-%d')
        from_date = (datetime.now() - timedelta(days=730)).strftime('%Y-%m-%d')

        mrp_lookup = {}
        page = 1
        max_pages = 50
        while page <= max_pages:
            try:
                resp = http_requests.post(
                    'https://umanmrp.in/api/party-dispatch-history.php',
                    json={'party_id': party_id, 'from_date': from_date,
                          'to_date': to_date, 'page': page, 'limit': 10000},
                    timeout=120
                )
                if resp.status_code != 200:
                    break
                data = resp.json()
                ds = data.get('dispatch_summary', [])
                if not ds:
                    break
                for d in ds:
                    pallet_nos = d.get('pallet_nos', {}) or {}
                    if isinstance(pallet_nos, dict):
                        for pallet_no, barcodes_str in pallet_nos.items():
                            if not isinstance(barcodes_str, str):
                                continue
                            for s in barcodes_str.strip().split():
                                s = s.strip().upper()
                                if s:
                                    mrp_lookup[s] = {
                                        'pallet_no': pallet_no,
                                        'vehicle_no': d.get('vehicle_no') or 'Unknown',
                                        'dispatch_date': d.get('dispatch_date') or d.get('date', ''),
                                        'invoice_no': d.get('invoice_no', ''),
                                        'dispatch_party': d.get('dispatch_party', '') or d.get('party_name', '') or party_name
                                    }
                page += 1
            except Exception:
                break

        actual_dispatched = actual_set & set(mrp_lookup.keys())
        actual_not_dispatched = actual_set - actual_dispatched

        # 5. Packed check — BULK approach: call get_barcode_tracking.php with party_name ONCE
        #    This returns ALL packed barcodes for the party — no cap, no per-barcode loops.
        packed_info = {}   # serial -> {pallet_no, running_order, packed_party}
        packed_set = set()
        pending_set = set()

        bulk_packed_lookup = {}   # serial -> info, from bulk party call

        if party_name:
            try:
                bulk_resp = http_requests.post(
                    'https://umanmrp.in/api/get_barcode_tracking.php',
                    json={'party_name': party_name},
                    timeout=120
                )
                if bulk_resp.status_code == 200:
                    bulk_data = bulk_resp.json()
                    items = bulk_data.get('data') or []
                    for item in items:
                        bc = str(item.get('barcode', '') or '').strip().upper()
                        if bc:
                            bulk_packed_lookup[bc] = {
                                'pallet_no': item.get('pallet_no', ''),
                                'running_order': item.get('running_order', '') or item.get('ro_no', '') or '',
                                'packed_party': item.get('packed_party', '') or item.get('party_name', '') or party_name,
                                'packing_date': item.get('packing_date', ''),
                                'box_no': item.get('box_no', '')
                            }
            except Exception as e:
                print(f"[pdi_actual_compare] Bulk pack fetch error: {e}")

        # Intersect actual_not_dispatched with bulk_packed_lookup
        for s in actual_not_dispatched:
            if s in bulk_packed_lookup:
                packed_set.add(s)
                packed_info[s] = bulk_packed_lookup[s]
            else:
                pending_set.add(s)

        try:
            planned_wattage = float(pdi_detail_map.get(pdi_id, {}).get('wattage') or 0)
        except Exception:
            planned_wattage = 0
        planned_kw = round(planned_wattage * planned_total / 1000.0, 2)
        actual_kw = round(planned_wattage * actual_total / 1000.0, 2)

        # Running Order breakdown: group packed barcodes by running_order
        ro_packed_map = {}   # ro_label -> list of packed serials
        ro_unknown_packed = []
        for s in packed_set:
            info = packed_info.get(s) or {}
            ro = (info.get('running_order') or '').strip()
            if ro:
                ro_packed_map.setdefault(ro, []).append(s)
            else:
                ro_unknown_packed.append(s)
        if ro_unknown_packed:
            ro_packed_map['Unknown RO'] = ro_unknown_packed

        running_order_breakdown = []
        for ro_label, ro_serials in sorted(ro_packed_map.items()):
            running_order_breakdown.append({
                'running_order': ro_label,
                'packed_count': len(ro_serials),
                'serials_sample': ro_serials[:50]
            })
        running_order_breakdown.sort(key=lambda x: x['packed_count'], reverse=True)

        # Dispatch breakdown: group dispatched barcodes by vehicle/pallet
        # dispatch_party is already available in mrp_lookup from party-dispatch-history.php
        dispatch_vehicle_map = {}
        for s in actual_dispatched:
            info = mrp_lookup.get(s) or {}
            key = info.get('vehicle_no') or 'Unknown Vehicle'
            dispatch_vehicle_map.setdefault(key, {
                'vehicle_no': key,
                'dispatch_date': info.get('dispatch_date', ''),
                'invoice_no': info.get('invoice_no', ''),
                'dispatch_party': info.get('dispatch_party', '') or party_name,
                'barcodes': []
            })
            dispatch_vehicle_map[key]['barcodes'].append(s)

        dispatch_breakdown = []
        for veh, vdata in dispatch_vehicle_map.items():
            dispatch_breakdown.append({
                'vehicle_no': vdata['vehicle_no'],
                'dispatch_date': vdata['dispatch_date'],
                'invoice_no': vdata['invoice_no'],
                'dispatch_party': vdata['dispatch_party'],
                'module_count': len(vdata['barcodes']),
                'serials_sample': vdata['barcodes'][:50]
            })
        dispatch_breakdown.sort(key=lambda x: x['module_count'], reverse=True)

        return jsonify({
            "success": True,
            "pdi": {
                "id": pdi_id,
                "name": pdi_detail_map.get(pdi_id, {}).get('pdi_name', ''),
                "wattage": pdi_detail_map.get(pdi_id, {}).get('wattage', ''),
                "planned_quantity": pdi_detail_map.get(pdi_id, {}).get('quantity', 0),
                "planned_kw": planned_kw,
                "actual_kw": actual_kw
            },
            "summary": {
                "planned": planned_total,
                "actual_uploaded": actual_total,
                "matched": len(matched),
                "missing": len(missing),            # planned but not delivered
                "extras": len(extras),              # delivered but not in plan
                "unknown_extras": len(unknown_extras),
                "actual_dispatched": len(actual_dispatched),
                "actual_packed": len(packed_set),
                "actual_pending": len(pending_set),
                "variance_percent": round(((actual_total - planned_total) / planned_total * 100), 2) if planned_total else 0
            },
            "extras_breakdown": extras_breakdown,  # which PDI does each extra belong to
            "missing_in_other_pdis": [{
                "pdi_id": k,
                "pdi_name": pdi_detail_map.get(k, {}).get('pdi_name', ''),
                "count": len(v),
                "serials_sample": v[:50]
            } for k, v in missing_other_owners.items()],
            "running_order_breakdown": running_order_breakdown,   # per-RO packed count
            "dispatch_breakdown": dispatch_breakdown,              # per-vehicle dispatched count
            "matched_sample": sorted(list(matched))[:200],
            "missing_sample": sorted(list(missing))[:200],
            "extras_sample": sorted(list(extras))[:200],
            "packed_sample": [{"serial": s, **(packed_info.get(s) or {})} for s in list(packed_set)[:200]],
            "pending_sample": sorted(list(pending_set))[:200],
            "dispatched_sample": [{"serial": s, **mrp_lookup.get(s, {})} for s in list(actual_dispatched)[:200]]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/parties-with-pdis', methods=['GET'])
def parties_with_pdis():
    """Return only parties that have at least one PDI in MRP.

    Internally fetches /ftr/sales-parties list, then for each party_name_id
    calls https://umanmrp.in/get/get_all_pdi.php in parallel and keeps only
    parties whose `data` array is non-empty.

    Cached for 10 minutes. Pass ?force_refresh=true to bypass.

    Response: { success, count, parties: [{id, companyName, pdiCount}] }
    """
    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        force = request.args.get('force_refresh', '').lower() == 'true'
        now = time.time()

        cache = parties_with_pdis.__dict__.setdefault(
            '_cache', {'data': None, 'timestamp': 0}
        )
        if (not force
                and cache['data'] is not None
                and (now - cache['timestamp']) < 600):
            return jsonify({
                "success": True,
                "cached": True,
                "count": len(cache['data']),
                "parties": cache['data']
            })

        # Get full party list (use existing cache if possible)
        if SALES_PARTY_CACHE['data'] is not None and (now - SALES_PARTY_CACHE['timestamp']) < SALES_PARTY_CACHE_TTL:
            all_parties = SALES_PARTY_CACHE['data']
        else:
            person_id = request.args.get('person_id') or SALES_PERSON_ID
            sp_resp = http_requests.post(SALES_PARTY_API, json={"personId": person_id}, timeout=60)
            payload = sp_resp.json() if sp_resp.status_code == 200 else {}
            raw = payload.get('data') or []
            all_parties = []
            for p in raw:
                pid = p.get('PartyNameId')
                name = (p.get('PartyName') or '').strip()
                if pid and name:
                    all_parties.append({"id": pid, "companyName": name})
            SALES_PARTY_CACHE['data'] = all_parties
            SALES_PARTY_CACHE['timestamp'] = now

        def check_party(party):
            try:
                r = http_requests.post(
                    'https://umanmrp.in/get/get_all_pdi.php',
                    json={"party_name_id": party['id']},
                    timeout=30
                )
                d = r.json() if r.status_code == 200 else {}
                pdis = d.get('data') if d.get('status') == 'success' else None
                pdi_count = len(pdis) if isinstance(pdis, list) else 0
                if pdi_count > 0:
                    return {
                        "id": party['id'],
                        "companyName": party['companyName'],
                        "pdiCount": pdi_count
                    }
            except Exception:
                return None
            return None

        results = []
        with ThreadPoolExecutor(max_workers=10) as ex:
            futures = [ex.submit(check_party, p) for p in all_parties]
            for f in as_completed(futures):
                r = f.result()
                if r:
                    results.append(r)

        results.sort(key=lambda x: x['companyName'].lower())

        cache['data'] = results
        cache['timestamp'] = now

        return jsonify({
            "success": True,
            "cached": False,
            "count": len(results),
            "parties": results
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/dispatch-by-party/<party_id>', methods=['GET'])
def get_dispatch_by_party(party_id):
    """
    Fetch dispatch tracking (pallet/vehicle/invoice info) directly from the
    MRP party-dispatch-history.php API for ANY party_id (UUID).

    Unlike /dispatch-tracking/<company_id>, this does NOT require the company
    to exist in the local DB. Useful for new sales parties pulled from
    logistics.umanerp.com getSalesParty.

    Query params:
        - days   : how many days back to fetch (default 730)
        - name   : optional display name to echo back

    Response shape (same as /dispatch-tracking/<company_id>):
        { success, company_name, summary, pallet_groups, dispatch_groups }
    """
    try:
        if not party_id or len(party_id) < 10:
            return jsonify({"success": False, "error": "Invalid party_id"}), 400

        days = int(request.args.get('days', '730'))
        company_name = request.args.get('name', '') or party_id

        from datetime import timedelta
        to_date = datetime.now().strftime('%Y-%m-%d')
        from_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        print(f"[Dispatch By Party] party_id={party_id}, range={from_date}..{to_date}")

        mrp_lookup = {}
        page = 1
        max_pages = 50

        while page <= max_pages:
            try:
                payload = {
                    'party_id': party_id,
                    'from_date': from_date,
                    'to_date': to_date,
                    'page': page,
                    'limit': 10000
                }
                response = http_requests.post(
                    'https://umanmrp.in/api/party-dispatch-history.php',
                    json=payload,
                    timeout=120,
                    headers={'Cache-Control': 'no-cache', 'Pragma': 'no-cache'}
                )

                if response.status_code != 200:
                    print(f"[Dispatch By Party] page {page} HTTP {response.status_code}")
                    break

                data = response.json()
                dispatch_summary = data.get('dispatch_summary', [])
                if not dispatch_summary:
                    break

                for dispatch in dispatch_summary:
                    dispatch_date = dispatch.get('dispatch_date') or dispatch.get('date', '')
                    vehicle_no = dispatch.get('vehicle_no', '') or 'Unknown'
                    invoice_no = dispatch.get('invoice_no', '')
                    factory_name = dispatch.get('factory_name', '')
                    dispatch_party = dispatch.get('dispatch_party', '') or vehicle_no
                    pallet_nos = dispatch.get('pallet_nos', {})

                    if isinstance(pallet_nos, dict):
                        for pallet_no, barcodes_str in pallet_nos.items():
                            if not isinstance(barcodes_str, str):
                                continue
                            for serial in barcodes_str.strip().split():
                                serial = serial.strip().upper()
                                if not serial:
                                    continue
                                mrp_lookup[serial] = {
                                    'pallet_no': pallet_no,
                                    'dispatch_party': dispatch_party,
                                    'vehicle_no': vehicle_no,
                                    'dispatch_date': dispatch_date,
                                    'invoice_no': invoice_no,
                                    'factory_name': factory_name
                                }

                print(f"[Dispatch By Party] page {page}: {len(dispatch_summary)} entries (running total {len(mrp_lookup)})")
                page += 1
            except Exception as e:
                print(f"[Dispatch By Party] page {page} error: {e}")
                break

        total = len(mrp_lookup)

        if total == 0:
            return jsonify({
                "success": True,
                "company_name": company_name,
                "party_id": party_id,
                "summary": {
                    "total_assigned": 0, "packed": 0, "dispatched": 0, "pending": 0,
                    "packed_percent": 0, "dispatched_percent": 0, "pending_percent": 0
                },
                "pallet_groups": [],
                "dispatch_groups": [],
                "message": "No dispatch data found in MRP system for this party"
            })

        vehicle_map = {}
        pallet_map = {}

        for barcode, info in mrp_lookup.items():
            vehicle = info.get('vehicle_no') or info.get('dispatch_party') or 'Unknown'
            pallet_no = info.get('pallet_no', '')
            dispatch_date = info.get('dispatch_date', '')
            invoice_no = info.get('invoice_no', '')
            factory_name = info.get('factory_name', '')

            if vehicle not in vehicle_map:
                vehicle_map[vehicle] = {
                    'dispatch_party': vehicle,
                    'vehicle_no': vehicle,
                    'dispatch_date': dispatch_date,
                    'invoice_no': invoice_no,
                    'factory_name': factory_name,
                    'module_count': 0,
                    'pallets': set(),
                    'serials': []
                }
            vehicle_map[vehicle]['module_count'] += 1
            if pallet_no:
                vehicle_map[vehicle]['pallets'].add(pallet_no)
            if len(vehicle_map[vehicle]['serials']) < 50:
                vehicle_map[vehicle]['serials'].append(barcode)

            if pallet_no:
                if pallet_no not in pallet_map:
                    pallet_map[pallet_no] = {
                        'pallet_no': pallet_no,
                        'module_count': 0,
                        'vehicle_no': vehicle,
                        'dispatch_date': dispatch_date,
                        'serials': []
                    }
                pallet_map[pallet_no]['module_count'] += 1
                if len(pallet_map[pallet_no]['serials']) < 20:
                    pallet_map[pallet_no]['serials'].append(barcode)

        dispatch_groups = []
        for v_name, v_data in vehicle_map.items():
            dispatch_groups.append({
                'dispatch_party': v_name,
                'vehicle_no': v_data['vehicle_no'],
                'dispatch_date': v_data['dispatch_date'],
                'invoice_no': v_data['invoice_no'],
                'factory_name': v_data['factory_name'],
                'module_count': v_data['module_count'],
                'pallet_count': len(v_data['pallets']),
                'pallets': sorted(list(v_data['pallets'])),
                'serials': v_data['serials']
            })
        dispatch_groups.sort(key=lambda x: x['module_count'], reverse=True)

        pallet_groups = sorted(pallet_map.values(), key=lambda x: str(x['pallet_no']))

        return jsonify({
            "success": True,
            "company_name": company_name,
            "party_id": party_id,
            "summary": {
                "total_assigned": total,
                "packed": 0,
                "dispatched": total,
                "pending": 0,
                "packed_percent": 0,
                "dispatched_percent": 100,
                "pending_percent": 0
            },
            "pallet_groups": pallet_groups,
            "dispatch_groups": dispatch_groups
        })

    except http_requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "MRP API timeout"}), 504
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


def _extract_packing_count_fast(party_name):
    """
    Returns packing count for a party by streaming only the beginning of the
    packing API response and extracting the top-level `count` field.
    """
    response = http_requests.post(
        'https://umanmrp.in/api/get_barcode_tracking.php',
        json={'party_name': party_name},
        timeout=60,
        stream=True
    )
    response.raise_for_status()

    buffer = ''
    try:
        for chunk in response.iter_content(chunk_size=4096, decode_unicode=True):
            if not chunk:
                continue
            buffer += chunk
            match = re.search(r'"count"\s*:\s*(\d+)', buffer)
            if match:
                return int(match.group(1))
            if len(buffer) > 120000:
                break
        return 0
    finally:
        response.close()


@ftr_bp.route('/packing-count-by-party/<party_id>', methods=['GET'])
def get_packing_count_by_party(party_id):
    """
    Dynamic packing-count lookup for any sales party.
    This endpoint is read-only and does not change existing dispatch logic.
    """
    try:
        if not party_id or len(party_id) < 10:
            return jsonify({"success": False, "error": "Invalid party_id"}), 400

        # Use cached list first; fetch fresh list if cache is empty.
        parties = SALES_PARTY_CACHE.get('data')
        if not parties:
            person_id = request.args.get('person_id') or SALES_PERSON_ID
            resp = http_requests.post(
                SALES_PARTY_API,
                json={"personId": person_id},
                timeout=60
            )
            if resp.status_code != 200:
                return jsonify({
                    "success": False,
                    "error": f"Sales party API returned HTTP {resp.status_code}",
                    "detail": resp.text[:300]
                }), 502

            payload = resp.json()
            raw_parties = payload.get('data') or []
            parties = []
            for p in raw_parties:
                pid = p.get('PartyNameId')
                name = (p.get('PartyName') or '').strip()
                if not pid or not name:
                    continue
                parties.append({
                    "id": pid,
                    "companyName": name,
                    "city": p.get('City') or '',
                    "state": p.get('State') or '',
                    "gst": p.get('GSTNo') or '',
                    "status": p.get('Status') or ''
                })
            parties.sort(key=lambda x: x['companyName'].lower())
            SALES_PARTY_CACHE['data'] = parties
            SALES_PARTY_CACHE['timestamp'] = time.time()

        selected = None
        for party in parties:
            if str(party.get('id')) == str(party_id):
                selected = party
                break

        if not selected:
            return jsonify({"success": False, "error": "Party not found for party_id"}), 404

        party_name = selected.get('companyName', '').strip()
        count = _extract_packing_count_fast(party_name)

        return jsonify({
            "success": True,
            "party_id": party_id,
            "party_name": party_name,
            "packing_count": count,
            "has_packing_data": count > 0
        })

    except http_requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "Packing API timeout"}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def _count_serial_tokens(value):
    if not value:
        return 0
    tokens = re.split(r'[\s,;|]+', str(value).strip())
    normalized = set()
    for token in tokens:
        token = token.strip().upper()
        if token:
            normalized.add(token)
    return len(normalized)


def _ensure_party_workspace_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS party_reallocation_workspace (
            id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
            party_id VARCHAR(64) NOT NULL,
            party_name VARCHAR(255) NOT NULL,
            pdi_serials LONGTEXT NULL,
            running_order_serials LONGTEXT NULL,
            barcode_serials LONGTEXT NULL,
            rejection_serials LONGTEXT NULL,
            smt_module_serials LONGTEXT NULL,
            pdi_number VARCHAR(120) NULL,
            running_order_number VARCHAR(120) NULL,
            rfid_data_json LONGTEXT NULL,
            rfid_row_count INT NOT NULL DEFAULT 0,
            rfid_uploaded_at DATETIME NULL,
            pdi_count INT NOT NULL DEFAULT 0,
            running_order_count INT NOT NULL DEFAULT 0,
            barcode_count INT NOT NULL DEFAULT 0,
            rejection_count INT NOT NULL DEFAULT 0,
            smt_module_count INT NOT NULL DEFAULT 0,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uq_party_reallocation_workspace_party_id (party_id),
            KEY idx_party_reallocation_workspace_updated_at (updated_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )

    # Backward-compatible column adds for already-created tables.
    cursor.execute("SHOW COLUMNS FROM party_reallocation_workspace")
    existing = {row['Field'] for row in (cursor.fetchall() or [])}
    column_adds = {
        'pdi_number': "ALTER TABLE party_reallocation_workspace ADD COLUMN pdi_number VARCHAR(120) NULL",
        'running_order_number': "ALTER TABLE party_reallocation_workspace ADD COLUMN running_order_number VARCHAR(120) NULL",
        'rfid_data_json': "ALTER TABLE party_reallocation_workspace ADD COLUMN rfid_data_json LONGTEXT NULL",
        'rfid_row_count': "ALTER TABLE party_reallocation_workspace ADD COLUMN rfid_row_count INT NOT NULL DEFAULT 0",
        'rfid_uploaded_at': "ALTER TABLE party_reallocation_workspace ADD COLUMN rfid_uploaded_at DATETIME NULL"
    }
    for col, ddl in column_adds.items():
        if col not in existing:
            cursor.execute(ddl)


def _normalize_rfid_col(name):
    return re.sub(r'[^a-z0-9]', '', str(name or '').strip().lower())


def _parse_rfid_excel(file_storage):
    required_cols = [
        'Date', 'ID', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Rs', 'Rsh',
        'Eff', 'T_Object', 'T_Target', 'Irr_Target', 'Class', 'Sweep_Time',
        'Irr_MonCell', 'Isc_MonCell', 'T_MonCell', 'T_Ambient', 'Binning'
    ]
    canonical_map = {_normalize_rfid_col(c): c for c in required_cols}

    df = pd.read_excel(file_storage)
    df.columns = [str(c).strip() for c in df.columns]
    input_map = {_normalize_rfid_col(c): c for c in df.columns}

    missing = [col for key, col in canonical_map.items() if key not in input_map]
    if missing:
        return {
            'ok': False,
            'missing': missing,
            'rows': []
        }

    normalized_df = pd.DataFrame()
    for key, canonical in canonical_map.items():
        normalized_df[canonical] = df[input_map[key]]

    # Keep upload sane for DB payload size while still useful for compare/testing.
    normalized_df = normalized_df.head(5000)

    records = []
    for row in normalized_df.to_dict(orient='records'):
        clean = {}
        for k, v in row.items():
            if pd.isna(v):
                clean[k] = None
            elif isinstance(v, (pd.Timestamp, datetime)):
                clean[k] = v.isoformat()
            else:
                clean[k] = v
        records.append(clean)

    return {
        'ok': True,
        'missing': [],
        'rows': records
    }


def _normalize_pdi_key(value):
    return (value or '').strip()


def _ensure_party_workspace_pdi_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS party_reallocation_workspace_pdi (
            id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
            party_id VARCHAR(64) NOT NULL,
            party_name VARCHAR(255) NOT NULL,
            pdi_key VARCHAR(120) NOT NULL,
            pdi_number VARCHAR(120) NULL,
            running_order_number VARCHAR(120) NULL,
            pdi_serials LONGTEXT NULL,
            running_order_serials LONGTEXT NULL,
            barcode_serials LONGTEXT NULL,
            rejection_serials LONGTEXT NULL,
            smt_module_serials LONGTEXT NULL,
            rfid_data_json LONGTEXT NULL,
            rfid_row_count INT NOT NULL DEFAULT 0,
            rfid_uploaded_at DATETIME NULL,
            pdi_count INT NOT NULL DEFAULT 0,
            running_order_count INT NOT NULL DEFAULT 0,
            barcode_count INT NOT NULL DEFAULT 0,
            rejection_count INT NOT NULL DEFAULT 0,
            smt_module_count INT NOT NULL DEFAULT 0,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uq_party_reallocation_workspace_pdi (party_id, pdi_key),
            KEY idx_party_reallocation_workspace_pdi_updated_at (updated_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )


def _workspace_pdi_row_to_payload(row):
    if not row:
        return None
    return {
        "partyId": row.get('party_id') or '',
        "partyName": row.get('party_name') or '',
        "pdiKey": row.get('pdi_key') or '',
        "pdiNumber": row.get('pdi_number') or '',
        "runningOrderNumber": row.get('running_order_number') or '',
        "pdiSerials": row.get('pdi_serials') or '',
        "runningOrderSerials": row.get('running_order_serials') or '',
        "barcodeSerials": row.get('barcode_serials') or '',
        "rejectionSerials": row.get('rejection_serials') or '',
        "smtModuleSerials": row.get('smt_module_serials') or '',
        "rfidRowCount": int(row.get('rfid_row_count') or 0),
        "rfidUploadedAt": row.get('rfid_uploaded_at').isoformat() if row.get('rfid_uploaded_at') else None,
        "counts": {
            "pdi": int(row.get('pdi_count') or 0),
            "runningOrder": int(row.get('running_order_count') or 0),
            "barcode": int(row.get('barcode_count') or 0),
            "rejection": int(row.get('rejection_count') or 0),
            "smtModule": int(row.get('smt_module_count') or 0)
        },
        "updatedAt": row.get('updated_at').isoformat() if row.get('updated_at') else None
    }


@ftr_bp.route('/party-workspace/<party_id>/pdi-cards', methods=['GET'])
def list_party_workspace_pdi_cards(party_id):
    try:
        conn = get_db_connection()
        not_found = False
        with conn.cursor() as cursor:
            _ensure_party_workspace_pdi_table(cursor)
            cursor.execute(
                """
                SELECT
                    party_id,
                    party_name,
                    pdi_key,
                    pdi_number,
                    running_order_number,
                    rfid_row_count,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count,
                    updated_at
                FROM party_reallocation_workspace_pdi
                WHERE party_id = %s
                ORDER BY updated_at DESC
                """,
                (party_id,)
            )
            rows = cursor.fetchall() or []

        conn.commit()
        conn.close()

        cards = []
        for row in rows:
            cards.append({
                "partyId": row.get('party_id') or '',
                "partyName": row.get('party_name') or '',
                "pdiKey": row.get('pdi_key') or '',
                "pdiNumber": row.get('pdi_number') or '',
                "runningOrderNumber": row.get('running_order_number') or '',
                "rfidRowCount": int(row.get('rfid_row_count') or 0),
                "counts": {
                    "pdi": int(row.get('pdi_count') or 0),
                    "runningOrder": int(row.get('running_order_count') or 0),
                    "barcode": int(row.get('barcode_count') or 0),
                    "rejection": int(row.get('rejection_count') or 0),
                    "smtModule": int(row.get('smt_module_count') or 0)
                },
                "updatedAt": row.get('updated_at').isoformat() if row.get('updated_at') else None
            })

        return jsonify({
            "success": True,
            "party_id": party_id,
            "cards": cards
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>/pdi-cards/<pdi_key>', methods=['GET'])
def get_party_workspace_pdi_card(party_id, pdi_key):
    try:
        normalized_key = _normalize_pdi_key(pdi_key)
        if not normalized_key:
            return jsonify({"success": False, "error": "pdi_key is required"}), 400

        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_pdi_table(cursor)
            cursor.execute(
                """
                SELECT
                    party_id,
                    party_name,
                    pdi_key,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    rfid_row_count,
                    rfid_uploaded_at,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count,
                    updated_at
                FROM party_reallocation_workspace_pdi
                WHERE party_id = %s AND pdi_key = %s
                LIMIT 1
                """,
                (party_id, normalized_key)
            )
            row = cursor.fetchone()

        conn.commit()
        conn.close()

        if not row:
            return jsonify({
                "success": True,
                "party_id": party_id,
                "pdi_key": normalized_key,
                "workspace": {
                    "partyId": party_id,
                    "partyName": "",
                    "pdiKey": normalized_key,
                    "pdiNumber": normalized_key,
                    "runningOrderNumber": "",
                    "pdiSerials": "",
                    "runningOrderSerials": "",
                    "barcodeSerials": "",
                    "rejectionSerials": "",
                    "smtModuleSerials": "",
                    "rfidRowCount": 0,
                    "rfidUploadedAt": None,
                    "counts": {
                        "pdi": 0,
                        "runningOrder": 0,
                        "barcode": 0,
                        "rejection": 0,
                        "smtModule": 0
                    },
                    "updatedAt": None
                }
            })

        return jsonify({
            "success": True,
            "party_id": party_id,
            "pdi_key": normalized_key,
            "workspace": _workspace_pdi_row_to_payload(row)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>/pdi-cards/<pdi_key>', methods=['POST'])
def save_party_workspace_pdi_card(party_id, pdi_key):
    try:
        normalized_key = _normalize_pdi_key(pdi_key)
        if not normalized_key:
            return jsonify({"success": False, "error": "pdi_key is required"}), 400

        data = request.get_json(silent=True) or {}
        party_name = (data.get('partyName') or '').strip()
        pdi_number = (data.get('pdiNumber') or normalized_key).strip()
        running_order_number = (data.get('runningOrderNumber') or '').strip()
        pdi_serials = (data.get('pdiSerials') or '').strip()
        running_order_serials = (data.get('runningOrderSerials') or '').strip()
        barcode_serials = (data.get('barcodeSerials') or '').strip()
        rejection_serials = (data.get('rejectionSerials') or '').strip()
        smt_module_serials = (data.get('smtModuleSerials') or '').strip()

        counts = {
            'pdi': _count_serial_tokens(pdi_serials),
            'running_order': _count_serial_tokens(running_order_serials),
            'barcode': _count_serial_tokens(barcode_serials),
            'rejection': _count_serial_tokens(rejection_serials),
            'smt_module': _count_serial_tokens(smt_module_serials)
        }

        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_pdi_table(cursor)
            cursor.execute(
                """
                INSERT INTO party_reallocation_workspace_pdi (
                    party_id,
                    party_name,
                    pdi_key,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    party_name = VALUES(party_name),
                    pdi_number = VALUES(pdi_number),
                    running_order_number = VALUES(running_order_number),
                    pdi_serials = VALUES(pdi_serials),
                    running_order_serials = VALUES(running_order_serials),
                    barcode_serials = VALUES(barcode_serials),
                    rejection_serials = VALUES(rejection_serials),
                    smt_module_serials = VALUES(smt_module_serials),
                    pdi_count = VALUES(pdi_count),
                    running_order_count = VALUES(running_order_count),
                    barcode_count = VALUES(barcode_count),
                    rejection_count = VALUES(rejection_count),
                    smt_module_count = VALUES(smt_module_count)
                """,
                (
                    party_id,
                    party_name,
                    normalized_key,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    counts['pdi'],
                    counts['running_order'],
                    counts['barcode'],
                    counts['rejection'],
                    counts['smt_module']
                )
            )

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "party_id": party_id,
            "pdi_key": normalized_key,
            "counts": {
                "pdi": counts['pdi'],
                "runningOrder": counts['running_order'],
                "barcode": counts['barcode'],
                "rejection": counts['rejection'],
                "smtModule": counts['smt_module']
            },
            "pdiNumber": pdi_number,
            "runningOrderNumber": running_order_number
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>/pdi-cards/<pdi_key>/upload-rfid-excel', methods=['POST'])
def upload_party_workspace_pdi_rfid_excel(party_id, pdi_key):
    try:
        normalized_key = _normalize_pdi_key(pdi_key)
        if not normalized_key:
            return jsonify({"success": False, "error": "pdi_key is required"}), 400

        if 'file' not in request.files:
            return jsonify({"success": False, "error": "Excel file is required"}), 400

        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"success": False, "error": "No file selected"}), 400

        parsed = _parse_rfid_excel(file)
        if not parsed['ok']:
            return jsonify({
                "success": False,
                "error": "RFID Excel missing required columns",
                "missingColumns": parsed['missing']
            }), 400

        rows = parsed['rows']
        ids = []
        for row in rows:
            val = row.get('ID')
            if val is None:
                continue
            text = str(val).strip().upper()
            if text:
                ids.append(text)
        unique_ids = sorted(set(ids))

        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_pdi_table(cursor)
            cursor.execute(
                """
                UPDATE party_reallocation_workspace_pdi
                SET
                    rfid_data_json = %s,
                    rfid_row_count = %s,
                    rfid_uploaded_at = NOW(),
                    barcode_serials = %s,
                    barcode_count = %s
                WHERE party_id = %s AND pdi_key = %s
                """,
                (
                    json.dumps(rows, ensure_ascii=True),
                    len(rows),
                    "\n".join(unique_ids),
                    len(unique_ids),
                    party_id,
                    normalized_key
                )
            )

            if cursor.rowcount == 0:
                not_found = True

        if not_found:
            conn.commit()
            conn.close()
            return jsonify({
                "success": False,
                "error": "PDI card not found. Create the PDI card first."
            }), 404

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "party_id": party_id,
            "pdi_key": normalized_key,
            "rfidRows": len(rows),
            "barcodeCount": len(unique_ids),
            "message": "RFID Excel uploaded successfully"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>', methods=['GET'])
def get_party_workspace(party_id):
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_table(cursor)
            cursor.execute(
                """
                SELECT
                    party_id,
                    party_name,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    rfid_row_count,
                    rfid_uploaded_at,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count,
                    updated_at
                FROM party_reallocation_workspace
                WHERE party_id = %s
                LIMIT 1
                """,
                (party_id,)
            )
            row = cursor.fetchone()

        conn.commit()
        conn.close()

        if not row:
            return jsonify({
                "success": True,
                "party_id": party_id,
                "workspace": {
                    "partyId": party_id,
                    "partyName": "",
                    "pdiNumber": "",
                    "runningOrderNumber": "",
                    "pdiSerials": "",
                    "runningOrderSerials": "",
                    "barcodeSerials": "",
                    "rejectionSerials": "",
                    "smtModuleSerials": "",
                    "rfidRowCount": 0,
                    "rfidUploadedAt": None,
                    "counts": {
                        "pdi": 0,
                        "runningOrder": 0,
                        "barcode": 0,
                        "rejection": 0,
                        "smtModule": 0
                    },
                    "updatedAt": None
                }
            })

        return jsonify({
            "success": True,
            "party_id": party_id,
            "workspace": {
                "partyId": row.get('party_id'),
                "partyName": row.get('party_name') or '',
                "pdiNumber": row.get('pdi_number') or '',
                "runningOrderNumber": row.get('running_order_number') or '',
                "pdiSerials": row.get('pdi_serials') or '',
                "runningOrderSerials": row.get('running_order_serials') or '',
                "barcodeSerials": row.get('barcode_serials') or '',
                "rejectionSerials": row.get('rejection_serials') or '',
                "smtModuleSerials": row.get('smt_module_serials') or '',
                "rfidRowCount": int(row.get('rfid_row_count') or 0),
                "rfidUploadedAt": row.get('rfid_uploaded_at').isoformat() if row.get('rfid_uploaded_at') else None,
                "counts": {
                    "pdi": int(row.get('pdi_count') or 0),
                    "runningOrder": int(row.get('running_order_count') or 0),
                    "barcode": int(row.get('barcode_count') or 0),
                    "rejection": int(row.get('rejection_count') or 0),
                    "smtModule": int(row.get('smt_module_count') or 0)
                },
                "updatedAt": row.get('updated_at').isoformat() if row.get('updated_at') else None
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>', methods=['POST'])
def save_party_workspace(party_id):
    try:
        data = request.get_json(silent=True) or {}

        party_name = (data.get('partyName') or '').strip()
        pdi_number = (data.get('pdiNumber') or '').strip()
        running_order_number = (data.get('runningOrderNumber') or '').strip()
        pdi_serials = (data.get('pdiSerials') or '').strip()
        running_order_serials = (data.get('runningOrderSerials') or '').strip()
        barcode_serials = (data.get('barcodeSerials') or '').strip()
        rejection_serials = (data.get('rejectionSerials') or '').strip()
        smt_module_serials = (data.get('smtModuleSerials') or '').strip()

        counts = {
            'pdi': _count_serial_tokens(pdi_serials),
            'running_order': _count_serial_tokens(running_order_serials),
            'barcode': _count_serial_tokens(barcode_serials),
            'rejection': _count_serial_tokens(rejection_serials),
            'smt_module': _count_serial_tokens(smt_module_serials)
        }

        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_table(cursor)
            cursor.execute(
                """
                INSERT INTO party_reallocation_workspace (
                    party_id,
                    party_name,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    party_name = VALUES(party_name),
                    pdi_number = VALUES(pdi_number),
                    running_order_number = VALUES(running_order_number),
                    pdi_serials = VALUES(pdi_serials),
                    running_order_serials = VALUES(running_order_serials),
                    barcode_serials = VALUES(barcode_serials),
                    rejection_serials = VALUES(rejection_serials),
                    smt_module_serials = VALUES(smt_module_serials),
                    pdi_count = VALUES(pdi_count),
                    running_order_count = VALUES(running_order_count),
                    barcode_count = VALUES(barcode_count),
                    rejection_count = VALUES(rejection_count),
                    smt_module_count = VALUES(smt_module_count)
                """,
                (
                    party_id,
                    party_name,
                    pdi_number,
                    running_order_number,
                    pdi_serials,
                    running_order_serials,
                    barcode_serials,
                    rejection_serials,
                    smt_module_serials,
                    counts['pdi'],
                    counts['running_order'],
                    counts['barcode'],
                    counts['rejection'],
                    counts['smt_module']
                )
            )

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "party_id": party_id,
            "counts": {
                "pdi": counts['pdi'],
                "runningOrder": counts['running_order'],
                "barcode": counts['barcode'],
                "rejection": counts['rejection'],
                "smtModule": counts['smt_module']
            },
            "pdiNumber": pdi_number,
            "runningOrderNumber": running_order_number
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace/<party_id>/upload-rfid-excel', methods=['POST'])
def upload_party_workspace_rfid_excel(party_id):
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "Excel file is required"}), 400

        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"success": False, "error": "No file selected"}), 400

        parsed = _parse_rfid_excel(file)
        if not parsed['ok']:
            return jsonify({
                "success": False,
                "error": "RFID Excel missing required columns",
                "missingColumns": parsed['missing']
            }), 400

        rows = parsed['rows']
        ids = []
        for row in rows:
            val = row.get('ID')
            if val is None:
                continue
            text = str(val).strip().upper()
            if text:
                ids.append(text)
        unique_ids = sorted(set(ids))

        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_table(cursor)
            cursor.execute(
                """
                UPDATE party_reallocation_workspace
                SET
                    rfid_data_json = %s,
                    rfid_row_count = %s,
                    rfid_uploaded_at = NOW(),
                    barcode_serials = %s,
                    barcode_count = %s
                WHERE party_id = %s
                """,
                (
                    json.dumps(rows, ensure_ascii=True),
                    len(rows),
                    "\n".join(unique_ids),
                    len(unique_ids),
                    party_id
                )
            )

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "party_id": party_id,
            "rfidRows": len(rows),
            "barcodeCount": len(unique_ids),
            "message": "RFID Excel uploaded successfully"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@ftr_bp.route('/party-workspace-summaries', methods=['GET'])
def get_party_workspace_summaries():
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            _ensure_party_workspace_table(cursor)
            cursor.execute(
                """
                SELECT
                    party_id,
                    party_name,
                    pdi_number,
                    running_order_number,
                    rfid_row_count,
                    rfid_uploaded_at,
                    pdi_count,
                    running_order_count,
                    barcode_count,
                    rejection_count,
                    smt_module_count,
                    updated_at
                FROM party_reallocation_workspace
                """
            )
            rows = cursor.fetchall() or []

        conn.commit()
        conn.close()

        summary_map = {}
        for row in rows:
            pid = row.get('party_id')
            if not pid:
                continue
            summary_map[pid] = {
                "partyName": row.get('party_name') or '',
                "pdiNumber": row.get('pdi_number') or '',
                "runningOrderNumber": row.get('running_order_number') or '',
                "rfidRowCount": int(row.get('rfid_row_count') or 0),
                "rfidUploadedAt": row.get('rfid_uploaded_at').isoformat() if row.get('rfid_uploaded_at') else None,
                "counts": {
                    "pdi": int(row.get('pdi_count') or 0),
                    "runningOrder": int(row.get('running_order_count') or 0),
                    "barcode": int(row.get('barcode_count') or 0),
                    "rejection": int(row.get('rejection_count') or 0),
                    "smtModule": int(row.get('smt_module_count') or 0)
                },
                "updatedAt": row.get('updated_at').isoformat() if row.get('updated_at') else None
            }

        return jsonify({
            "success": True,
            "summaries": summary_map
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


