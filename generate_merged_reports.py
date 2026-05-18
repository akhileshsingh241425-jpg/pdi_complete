"""FTR via server (stable version with retry + error handling)."""

import os
import glob
import time
import openpyxl
import requests
from io import BytesIO
from PyPDF2 import PdfReader, PdfWriter

SERVER_URL = "https://pdi.gspl.cloud"
API_ENDPOINT = f"{SERVER_URL}/api/ftr/generate-report"

INPUT_BASE = r'E:\BSNL\Generated_Output'

SAMPLE_MODE = True
SAMPLE_LIMIT = 3

DEFAULT_MODULE_POWER = 550
REQUEST_TIMEOUT = 120


def _num(value, default=0.0):
    try:
        if value is None or value == '':
            return float(default)
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _build_payload(row, col_map):
    def g(key, default=None):
        idx = col_map.get(key)
        if idx is None:
            return default
        val = row[idx]
        return val if val not in (None, '') else default

    date_val = g('Date', '')
    time_val = g('Time', '')

    return {
        "producer":     g('Producer', 'Gautam Solar Pvt.Ltd.'),
        "moduleType":   g('ModuleType', 'G2X550-HAD'),
        "serialNumber": g('SerialNumber', ''),
        "testDate":     str(date_val).split(' ')[0] if date_val else '',
        "testTime":     str(time_val) if time_val else '',
        "irradiance":   _num(g('Irradiance'), 1000.0),
        "moduleTemp":   _num(g('ModuleTemp'), 25.0),
        "ambientTemp":  _num(g('AmbientTemp'), 25.0),
        "moduleArea":   _num(g('ModuleArea'), 2.5832),
        "modulePower":  _num(g('ModulePower'), DEFAULT_MODULE_POWER),
        "results": {
            "pmax":       _num(g('Pmax')),
            "vpm":        _num(g('Vpm')),
            "ipm":        _num(g('Ipm')),
            "voc":        _num(g('Voc')),
            "isc":        _num(g('Isc')),
            "fillFactor": _num(g('FillFactor')),
            "rs":         _num(g('Rs')),
            "rsh":        _num(g('Rsh')),
            "efficiency": _num(g('Efficiency')),
        },
    }


def _fetch_pdf_from_server(session, payload):
    for attempt in range(3):
        try:
            resp = session.post(
                API_ENDPOINT,
                json=payload,
                timeout=REQUEST_TIMEOUT,
                headers={"Accept": "application/pdf"}
            )

            if resp.status_code != 200:
                raise RuntimeError(f"API Error {resp.status_code}: {resp.text[:200]}")

            content_type = resp.headers.get("Content-Type", "")
            if "application/pdf" not in content_type:
                raise RuntimeError(f"Expected PDF, got {content_type} | {resp.text[:200]}")

            return resp.content

        except Exception as e:
            print(f"  Retry {attempt+1}/3 failed: {e}")
            time.sleep(2)

    raise RuntimeError("Failed after 3 retries")


def process_excel(session, excel_path, subfolder_path, subfolder_name):
    excel_name = os.path.basename(excel_path)
    base_name = os.path.splitext(excel_name)[0]

    prefix = "SAMPLE_" if SAMPLE_MODE else ""
    output_pdf_path = os.path.join(subfolder_path, f"{prefix}Merged_FTR_{base_name}.pdf")

    print(f"\n=== {excel_name}  (folder: {subfolder_name}) ===")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers) if h}

    if 'SerialNumber' not in col_map:
        print(f"  [SKIP] SerialNumber column missing. Found: {list(col_map.keys())}")
        return None

    merged_writer = PdfWriter()
    processed = 0
    failed = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        serial = row[col_map['SerialNumber']]
        if not serial:
            continue

        try:
            payload = _build_payload(row, col_map)
            print(f"  Processing: {serial}")

            pdf_bytes = _fetch_pdf_from_server(session, payload)

            reader = PdfReader(BytesIO(pdf_bytes))
            if len(reader.pages) == 0:
                raise RuntimeError("Empty PDF received")

            merged_writer.add_page(reader.pages[0])

            processed += 1
            print(f"  [{processed}] OK")

            if SAMPLE_MODE and processed >= SAMPLE_LIMIT:
                break

        except Exception as e:
            failed += 1
            print(f"  [ERR] {serial}: {e}")

    if processed:
        with open(output_pdf_path, "wb") as f:
            merged_writer.write(f)

        print(f"\n[DONE] {processed} pages | failed: {failed}")
        print(f"Saved -> {output_pdf_path}")
        return output_pdf_path

    print("[SKIP] No pages generated.")
    return None


def main():
    print(f"Server      : {SERVER_URL}")
    print(f"Input base  : {INPUT_BASE}")
    print(f"Mode        : {'SAMPLE ' + str(SAMPLE_LIMIT) if SAMPLE_MODE else 'FULL'}")

    if not os.path.exists(INPUT_BASE):
        print(f"[FATAL] INPUT_BASE not found: {INPUT_BASE}")
        return

    try:
        r = requests.get(SERVER_URL, timeout=10)
        print(f"Server ping : HTTP {r.status_code}")
    except Exception as e:
        print(f"[FATAL] Can't reach server: {e}")
        return

    subfolders = sorted(
        d for d in os.listdir(INPUT_BASE)
        if os.path.isdir(os.path.join(INPUT_BASE, d)) and d.startswith('QA')
    )

    if not subfolders:
        print("No QA* subfolders found.")
        return

    session = requests.Session()
    start = time.time()

    if SAMPLE_MODE:
        sf = subfolders[0]
        sf_path = os.path.join(INPUT_BASE, sf)

        excel_files = glob.glob(os.path.join(sf_path, "*.xlsx"))
        if not excel_files:
            print(f"No Excel files in: {sf_path}")
            return

        process_excel(session, excel_files[0], sf_path, sf)

    else:
        for sf in subfolders:
            sf_path = os.path.join(INPUT_BASE, sf)

            for excel_file in glob.glob(os.path.join(sf_path, "*.xlsx")):
                process_excel(session, excel_file, sf_path, sf)

    print(f"\nTotal time: {time.time() - start:.1f}s")


if __name__ == "__main__":
    main()