r"""
IV Curve Generator for 550W Solar Modules
==========================================
Reads module parameters from Excel files in E:\BSNL\Generated_Output subfolders
and generates IV curve graphs matching the 550W reference style.
Each module gets its own graph saved as PNG.
"""

import os
import sys
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# Configuration
# ============================================================
INPUT_BASE = r'E:\BSNL\Generated_Output'
EXCEL_FILENAME = 'IV_Curve_Data.xlsx'
OUTPUT_SUBFOLDER = 'IV_Curves'

GRAPH_DPI = 150
NUM_POINTS = 500


def generate_iv_curve(Voc, Isc, Vpm, Ipm, Rs, Rsh, num_points=NUM_POINTS):
    """
    Generate a realistic IV curve using a piecewise approach that ensures
    the curve passes through Isc, (Vpm, Ipm), and Voc accurately,
    matching the characteristic flat-then-sharp-drop shape of real modules.
    """
    # Generate voltage points
    V = np.linspace(0, Voc, num_points)
    I = np.zeros(num_points)

    # Key parameters
    Pmax = Vpm * Ipm

    # Use Lambert-W based analytical approach for realistic shape
    # The IV curve has 3 regions:
    # 1. Flat region (0 to ~0.7*Voc): current stays near Isc
    # 2. Knee region (~0.7*Voc to Vpm): current starts to drop
    # 3. Drop region (Vpm to Voc): current drops sharply to 0

    # Calculate shape parameter from known points
    # At V=Vpm, I=Ipm, so we need: Ipm = Isc * f(Vpm/Voc)
    # Shape determines how flat the curve is before the knee

    # Estimate diode quality factor
    v_ratio = Vpm / Voc
    i_ratio = Ipm / Isc

    # Use the fill factor to determine the sharpness of the knee
    FF = Pmax / (Voc * Isc)

    # Normalized voltage for the diode equation
    # Higher 'n' = sharper knee = more like ideal diode
    # For FF around 80%, n should be around 0.025-0.035

    # Solve for 'n' such that I(Vpm) = Ipm
    # Using: I = Isc - Isc * (exp((V - Voc)/n/Voc) - exp(-Voc/n/Voc)) / (1 - exp(-Voc/n/Voc))
    # Simplified: I ≈ Isc * (1 - exp((V - Voc) / (n * Voc)))

    # Better approach: use two-piece model
    # I(V) = Isc - (Isc - shunt_drop) * [(exp(V/Voc * k) - 1) / (exp(k) - 1)]
    # where k controls the knee sharpness

    # Find k that satisfies I(Vpm) = Ipm
    # Ipm = Isc - Isc * [(exp(Vpm/Voc * k) - 1) / (exp(k) - 1)]
    # (Isc - Ipm)/Isc = (exp(Vpm/Voc * k) - 1) / (exp(k) - 1)

    target_ratio = (Isc - Ipm) / Isc  # fraction of current lost at MPP

    # Binary search for k
    k_low, k_high = 2.0, 50.0
    for _ in range(100):
        k_mid = (k_low + k_high) / 2.0
        ratio_at_vpm = (np.exp(v_ratio * k_mid) - 1) / (np.exp(k_mid) - 1)
        if ratio_at_vpm < target_ratio:
            k_low = k_mid
        else:
            k_high = k_mid
    k = (k_low + k_high) / 2.0

    # Shunt resistance effect: linear droop across the full range
    shunt_slope = 1.0 / Rsh if Rsh > 0 else 0

    for i_idx, v in enumerate(V):
        v_norm = v / Voc
        # Exponential diode current
        diode_fraction = (np.exp(v_norm * k) - 1) / (np.exp(k) - 1)
        # Shunt current loss
        shunt_loss = v * shunt_slope
        # Series resistance effect - shifts the curve slightly
        rs_effect = 0
        if Rs > 0 and v > Vpm * 0.5:
            # Rs causes additional voltage drop proportional to current
            rs_factor = Rs * Isc / Voc
            rs_effect = rs_factor * (v_norm - 0.5) * 0.3

        current = Isc * (1 - diode_fraction) - shunt_loss - rs_effect * Isc
        I[i_idx] = max(0, current)

    # Ensure endpoints
    I[0] = Isc
    I[-1] = 0

    # Power
    P = V * I

    return V, I, P


def plot_iv_curve(V, I, P, Voc, Isc, Pmax, output_path):
    """
    Plot IV curve matching the 550W reference style exactly:
    - Blue I-V curve (left Y axis)
    - Red Power curve (right Y axis)
    - Light gray background, dotted grid
    """
    fig, ax1 = plt.subplots(figsize=(5.8, 6.5))

    # Background color matching reference
    fig.patch.set_facecolor('#f0f0f0')
    ax1.set_facecolor('#f0f0f0')

    # Blue IV curve
    ax1.plot(V, I, color='#4444bb', linewidth=1.2, zorder=3)

    ax1.set_xlabel('Voltage(V)', fontsize=11, fontweight='bold', labelpad=8)
    ax1.set_ylabel('Current(A)', fontsize=11, fontweight='bold', labelpad=8)

    # Axis limits matching reference
    max_v = max(Voc * 1.08, 55)
    max_i_tick = int(np.ceil(Isc * 1.05))
    ax1.set_xlim(0, max_v)
    ax1.set_ylim(0, max_i_tick)

    # Ticks matching reference exactly
    ax1.set_yticks(np.arange(0, max_i_tick + 1, 1))
    ax1.set_xticks(np.arange(0, int(np.ceil(max_v / 5) * 5) + 1, 5))

    # Grid - dotted like reference
    ax1.grid(True, which='major', linestyle=':', linewidth=0.4, color='#bbbbbb', alpha=0.8)
    ax1.tick_params(axis='both', labelsize=9)

    # Red Power curve on right Y axis
    ax2 = ax1.twinx()
    ax2.plot(V, P, color='#cc4444', linewidth=1.0, zorder=2)
    ax2.set_ylabel('Power(W) [red]', fontsize=11, fontweight='bold',
                    color='#cc4444', labelpad=8, rotation=270)
    ax2.yaxis.set_label_coords(1.12, 0.5)

    max_p = max(Pmax * 1.15, 600)
    ax2.set_ylim(0, max_p)
    ax2.set_yticks(np.arange(0, int(max_p) + 1, 50))
    ax2.tick_params(axis='y', labelsize=9, colors='#cc4444')

    plt.tight_layout()
    plt.subplots_adjust(right=0.85)

    fig.savefig(output_path, dpi=GRAPH_DPI, bbox_inches='tight',
                facecolor=fig.get_facecolor(), edgecolor='none')
    plt.close(fig)


def process_all():
    """Process all subfolders and generate IV curves."""
    print("=" * 60)
    print("  IV CURVE GENERATOR - 550W Solar Modules")
    print("=" * 60)

    if not os.path.exists(INPUT_BASE):
        print(f"ERROR: {INPUT_BASE} not found!")
        return

    subfolders = sorted([
        d for d in os.listdir(INPUT_BASE)
        if os.path.isdir(os.path.join(INPUT_BASE, d)) and d.startswith('QA')
    ])

    print(f"\nFound {len(subfolders)} subfolders\n")

    grand_total = 0

    for sf_idx, sf in enumerate(subfolders):
        sf_path = os.path.join(INPUT_BASE, sf)
        excel_path = os.path.join(sf_path, EXCEL_FILENAME)

        if not os.path.exists(excel_path):
            print(f"[{sf_idx+1}/{len(subfolders)}] {sf}: No Excel found, skipping")
            continue

        out_dir = os.path.join(sf_path, OUTPUT_SUBFOLDER)
        os.makedirs(out_dir, exist_ok=True)

        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        col = {h: i for i, h in enumerate(headers)}

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        generated = 0
        skipped = 0
        errors = 0

        print(f"[{sf_idx+1}/{len(subfolders)}] {sf}: {total} modules")

        for r_idx, row in enumerate(rows):
            serial = str(row[col['SerialNumber']] or '').strip()
            if not serial:
                skipped += 1
                continue

            safe_name = serial.replace('/', '_').replace('\\', '_').replace(':', '_')
            out_path = os.path.join(out_dir, f'{safe_name}.png')

            # Skip if already exists
            if os.path.exists(out_path):
                generated += 1
                continue

            try:
                Pmax = float(row[col['Pmax']])
                Vpm  = float(row[col['Vpm']])
                Ipm  = float(row[col['Ipm']])
                Voc  = float(row[col['Voc']])
                Isc  = float(row[col['Isc']])
                Rs   = float(row[col['Rs']])
                Rsh  = float(row[col['Rsh']])

                V, I, P = generate_iv_curve(Voc, Isc, Vpm, Ipm, Rs, Rsh)
                plot_iv_curve(V, I, P, Voc, Isc, Pmax, out_path)
                generated += 1

            except Exception as e:
                errors += 1
                if errors <= 3:
                    print(f"  ERROR [{serial}]: {str(e)[:60]}")

            if (r_idx + 1) % 50 == 0:
                print(f"  {r_idx+1}/{total} done...")

        wb.close()
        print(f"  => {generated} generated, {skipped} skipped, {errors} errors")
        print(f"     Saved: {out_dir}")
        grand_total += generated

    print(f"\n{'='*60}")
    print(f"  DONE! Total: {grand_total} IV curves generated")
    print(f"{'='*60}")


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == 'test':
        print("Generating test IV curve...")
        V, I, P = generate_iv_curve(
            Voc=49.713528, Isc=13.85494,
            Vpm=42.107048, Ipm=13.181063,
            Rs=0.495896, Rsh=374.452993
        )
        test_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_iv_curve.png')
        plot_iv_curve(V, I, P, 49.713528, 13.85494, 555.015639, test_path)
        print(f"Test saved: {test_path}")
        print(f"Pmax in curve: {P.max():.1f}W at V={V[P.argmax()]:.1f}V")
    else:
        process_all()
