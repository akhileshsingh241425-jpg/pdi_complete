"""
PDI Documentation Generator Routes - v5
Individual downloads: IPQC Excel, Witness Excel, Sampling Excel, MOM PDF
Each document downloads separately for easy upload.
"""

from flask import Blueprint, request, jsonify, send_file, current_app
from app.models.database import db
from sqlalchemy import text
from datetime import datetime, timedelta
import os
import io
import random
import traceback
import json
import math

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

pdi_doc_bp = Blueprint('pdi_doc', __name__)

# ============ Excel Styles ============
if EXCEL_AVAILABLE:
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    light_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ============ DB Helpers ============

def get_companies():
    try:
        result = db.session.execute(text(
            "SELECT DISTINCT company_id, company_name FROM companies ORDER BY company_name"
        ))
        return [{'id': row[0], 'name': row[1]} for row in result.fetchall()]
    except Exception:
        try:
            result = db.session.execute(text(
                "SELECT DISTINCT company_id FROM ftr_master_serials ORDER BY company_id"
            ))
            return [{'id': row[0], 'name': row[0]} for row in result.fetchall()]
        except Exception:
            return []


def get_pdis_for_company(company_id):
    try:
        result = db.session.execute(text("""
            SELECT id, pdi_number, total_modules, created_at
            FROM pdi_batches WHERE company_id = :cid ORDER BY created_at DESC
        """), {'cid': company_id})
        return [{
            'id': row[0], 'pdi_number': row[1],
            'total_modules': row[2],
            'created_at': row[3].isoformat() if row[3] else None
        } for row in result.fetchall()]
    except Exception as e:
        print(f"Error getting PDIs: {e}")
        return []


def get_serials_for_pdi(pdi_id):
    try:
        result = db.session.execute(text(
            "SELECT serial_number FROM module_serial_numbers WHERE pdi_batch_id = :pid ORDER BY serial_number"
        ), {'pid': pdi_id})
        return [row[0] for row in result.fetchall()]
    except Exception as e:
        print(f"Error getting serials: {e}")
        return []


def get_ftr_data(serial_numbers):
    if not serial_numbers:
        return {}
    try:
        placeholders = ','.join([f':s{i}' for i in range(len(serial_numbers))])
        params = {f's{i}': s for i, s in enumerate(serial_numbers)}
        result = db.session.execute(text(f"""
            SELECT serial_number, pmax, isc, voc, ipm, vpm, ff, efficiency, binning
            FROM ftr_master_serials WHERE serial_number IN ({placeholders})
        """), params)
        data = {}
        for row in result.fetchall():
            data[row[0]] = {
                'pmax': float(row[1]) if row[1] else 0,
                'isc': float(row[2]) if row[2] else 0,
                'voc': float(row[3]) if row[3] else 0,
                'ipm': float(row[4]) if row[4] else 0,
                'vpm': float(row[5]) if row[5] else 0,
                'ff': float(row[6]) if row[6] else 0,
                'efficiency': float(row[7]) if row[7] else 0,
                'binning': row[8] if row[8] else ''
            }
        return data
    except Exception as e:
        print(f"Error getting FTR data: {e}")
        return {}


def get_calibration_instruments():
    try:
        result = db.session.execute(text("""
            SELECT sr_no, instrument_id, machine_name, make, model_name,
                   item_sr_no, range_capacity, least_count, location,
                   calibration_agency, date_of_calibration, due_date,
                   calibration_frequency, calibration_standards, certificate_no, status
            FROM calibration_instruments ORDER BY sr_no
        """))
        instruments = []
        for row in result.fetchall():
            instruments.append({
                'sr_no': row[0], 'instrument_id': row[1], 'machine_name': row[2],
                'make': row[3], 'model_name': row[4], 'item_sr_no': row[5],
                'range_capacity': row[6], 'least_count': row[7], 'location': row[8],
                'calibration_agency': row[9],
                'date_of_calibration': row[10].strftime('%d/%m/%Y') if row[10] else '',
                'due_date': row[11].strftime('%d/%m/%Y') if row[11] else '',
                'calibration_frequency': row[12], 'calibration_standards': row[13],
                'certificate_no': row[14], 'status': row[15]
            })
        return instruments
    except Exception as e:
        print(f"Error getting calibration instruments: {e}")
        return []


def aql_sample_size(lot_size):
    aql_table = [
        (8, 5), (15, 5), (25, 8), (50, 13), (90, 20),
        (150, 32), (280, 50), (500, 80), (1200, 125),
        (3200, 200), (10000, 315), (35000, 500),
        (150000, 800), (500000, 1250), (float('inf'), 2000)
    ]
    for max_lot, sample in aql_table:
        if lot_size <= max_lot:
            return sample
    return 2000


def safe_filename(name):
    for ch in '/\\:*?"<>|':
        name = name.replace(ch, '_')
    return name


def parse_request_data():
    """Parse common POST body fields"""
    data = request.json
    if not data:
        return None, 'No data provided'

    serial_numbers = data.get('serial_numbers', [])
    if not serial_numbers:
        return None, 'No serial numbers provided'

    total_qty = len(serial_numbers)
    sample_size = min(aql_sample_size(total_qty), total_qty)
    sampled = random.sample(serial_numbers, sample_size) if sample_size < total_qty else serial_numbers

    parsed = {
        'company_id': data.get('company_id', ''),
        'company_name': data.get('company_name', data.get('company_id', '')),
        'pdi_number': data.get('pdi_number', 'PDI-001'),
        'serial_numbers': serial_numbers,
        'production_days': data.get('production_days', 3),
        'report_date': data.get('report_date', datetime.now().strftime('%d/%m/%Y')),
        'module_type': data.get('module_type', 'G2G580'),
        'total_qty': total_qty,
        'sample_size': sample_size,
        'sampled_serials': sampled,
        'ftr_data': get_ftr_data(serial_numbers),
    }
    return parsed, None


# ============ ROUTES ============

@pdi_doc_bp.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'success': True, 'message': 'PDI Docs API v5 running',
        'excel_available': EXCEL_AVAILABLE, 'pdf_available': PDF_AVAILABLE
    }), 200


@pdi_doc_bp.route('/companies', methods=['GET'])
def list_companies():
    return jsonify({'success': True, 'companies': get_companies()}), 200


@pdi_doc_bp.route('/pdis/<company_id>', methods=['GET'])
def list_pdis(company_id):
    return jsonify({'success': True, 'pdis': get_pdis_for_company(company_id)}), 200


@pdi_doc_bp.route('/serials/<int:pdi_id>', methods=['GET'])
def list_serials(pdi_id):
    serials = get_serials_for_pdi(pdi_id)
    return jsonify({'success': True, 'serials': serials, 'count': len(serials)}), 200


@pdi_doc_bp.route('/template-info', methods=['GET'])
def template_info():
    try:
        from app.models.ipqc_data import IPQCTemplate
        template = IPQCTemplate.get_template()
        total_cp = sum(len(s.get('checkpoints', [])) for s in template)
        return jsonify({'success': True, 'total_stages': len(template), 'total_checkpoints': total_cp}), 200
    except Exception:
        return jsonify({'success': True, 'total_stages': 8, 'total_checkpoints': 30}), 200


# -------- Individual Download Endpoints --------

@pdi_doc_bp.route('/download/ipqc', methods=['POST'])
def download_ipqc():
    """Download IPQC Report as Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    try:
        d, err = parse_request_data()
        if err:
            return jsonify({'success': False, 'error': err}), 400

        buf = build_ipqc_excel(d)
        fname = f"IPQC_Report_{safe_filename(d['pdi_number'])}.xlsx"
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        print(f"IPQC download error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/download/witness', methods=['POST'])
def download_witness():
    """Download Witness Report as Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    try:
        d, err = parse_request_data()
        if err:
            return jsonify({'success': False, 'error': err}), 400

        buf = build_witness_excel(d)
        fname = f"Witness_Report_{safe_filename(d['pdi_number'])}.xlsx"
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        print(f"Witness download error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/download/sampling', methods=['POST'])
def download_sampling():
    """Download Sampling Plan as Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    try:
        d, err = parse_request_data()
        if err:
            return jsonify({'success': False, 'error': err}), 400

        buf = build_sampling_excel(d)
        fname = f"Sampling_Plan_{safe_filename(d['pdi_number'])}.xlsx"
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        print(f"Sampling download error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/download/calibration', methods=['POST'])
def download_calibration():
    """Download Calibration Instrument List as Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    try:
        d, err = parse_request_data()
        if err:
            return jsonify({'success': False, 'error': err}), 400

        instruments = get_calibration_instruments()
        buf = build_calibration_excel(d, instruments)
        fname = f"Calibration_List_{safe_filename(d['pdi_number'])}.xlsx"
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        print(f"Calibration download error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/download/mom', methods=['POST'])
def download_mom():
    """Download MOM as PDF"""
    if not PDF_AVAILABLE:
        return jsonify({'success': False, 'error': 'reportlab not installed'}), 500
    try:
        d, err = parse_request_data()
        if err:
            return jsonify({'success': False, 'error': err}), 400

        buf = build_mom_pdf(d)
        fname = f"MOM_{safe_filename(d['pdi_number'])}.pdf"
        return send_file(buf, mimetype='application/pdf',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        print(f"MOM PDF download error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============ EXCEL BUILDERS ============
# Pattern: direct cell assignment, borders only on data cells, no borders on merged titles.
# Matches the working witness_report_routes.py exactly.


def build_ipqc_excel(d):
    """Build IPQC Report workbook"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPQC Report"

    company = d['company_name']
    pdi = d['pdi_number']
    serials = d['serial_numbers']
    sampled = d['sampled_serials']
    report_date = d['report_date']
    module_type = d['module_type']
    prod_days = d['production_days']

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"IPQC INSPECTION REPORT - {company}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Info rows 2-4
    info_rows = [
        ('PDI Number', pdi, 'Date', report_date),
        ('Total Modules', str(len(serials)), 'Production Days', str(prod_days)),
        ('Module Type', module_type, 'Sample Size', str(len(sampled))),
    ]
    for r_idx, (l1, v1, l2, v2) in enumerate(info_rows, 2):
        for col in range(1, 9):
            cell = ws.cell(row=r_idx, column=col, value='')
            cell.border = thin_border
            cell.alignment = center_align
        ws.cell(row=r_idx, column=1, value=l1).font = Font(bold=True)
        ws.cell(row=r_idx, column=1).fill = light_fill
        ws.cell(row=r_idx, column=2, value=v1)
        ws.cell(row=r_idx, column=5, value=l2).font = Font(bold=True)
        ws.cell(row=r_idx, column=5).fill = light_fill
        ws.cell(row=r_idx, column=6, value=v2)

    # Header row (row 6)
    hdr_row = 6
    headers = ['Sr.No', 'Stage', 'Checkpoint', 'Acceptance Criteria', 'Sample Size', 'Frequency', 'Result', 'Remarks']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # IPQC stages
    row = hdr_row + 1
    try:
        from app.services.form_generator import IPQCFormGenerator
        generator = IPQCFormGenerator()
        form_data = generator.generate_form(
            date=report_date, shift='A', customer_id=company,
            po_number=pdi, module_count=len(serials)
        )
        stages = form_data.get('stages', [])
    except Exception as e:
        print(f"IPQCFormGenerator error: {e}")
        stages = []

    if stages:
        for stage in stages:
            stage_name = stage.get('stage', '')
            sr_no = stage.get('sr_no', '')
            for i, cp in enumerate(stage.get('checkpoints', [])):
                vals = [
                    sr_no if i == 0 else '',
                    stage_name if i == 0 else '',
                    cp.get('checkpoint', ''),
                    cp.get('acceptance_criteria', ''),
                    cp.get('sample_size', ''),
                    cp.get('frequency', ''),
                    cp.get('monitoring_result', 'OK'),
                    cp.get('remarks', ''),
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align if col not in (3, 4) else left_wrap
                if vals[6] == 'OK':
                    ws.cell(row=row, column=7).font = Font(color="008000", bold=True)
                row += 1
    else:
        basic = [
            ('1', 'Incoming', 'Raw Material Check', 'As per BOM', str(len(sampled)), 'Each Lot', 'OK', ''),
            ('2', 'Cell Sorting', 'Cell Efficiency', '>=22%', str(len(sampled)), 'Each Lot', 'OK', ''),
            ('3', 'Stringing', 'Solder Quality', 'No cold joints', str(len(sampled)), 'Hourly', 'OK', ''),
            ('4', 'Layup', 'Alignment', '+/-1mm', str(len(sampled)), 'Each Module', 'OK', ''),
            ('5', 'Lamination', 'Temp Profile', '145C +/-5C', str(len(sampled)), 'Each Batch', 'OK', ''),
            ('6', 'Trimming', 'Edge Quality', 'No rough edges', str(len(sampled)), 'Each Module', 'OK', ''),
            ('7', 'Framing', 'Frame Fit', 'No gaps', str(len(sampled)), 'Each Module', 'OK', ''),
            ('8', 'JB & Curing', 'Junction Box', 'Proper adhesion', str(len(sampled)), 'Each Module', 'OK', ''),
            ('9', 'Final Test', 'Flasher Test', 'Power >= Nameplate', str(len(sampled)), '100%', 'OK', ''),
            ('10', 'Packing', 'Packing', 'No damage', str(len(sampled)), 'Each Pallet', 'OK', ''),
        ]
        for vals in basic:
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center_align
            ws.cell(row=row, column=7).font = Font(color="008000", bold=True)
            row += 1

    widths = [8, 20, 35, 30, 12, 12, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Serial Numbers
    ws2 = wb.create_sheet("Serial Numbers")
    for col, h in enumerate(['Sr.No', 'Serial Number', 'Sampled'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, s in enumerate(serials, 1):
        ws2.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws2.cell(row=idx+1, column=2, value=s).border = thin_border
        is_sampled = 'YES' if s in sampled else ''
        cell = ws2.cell(row=idx+1, column=3, value=is_sampled)
        cell.border = thin_border
        if is_sampled:
            cell.font = Font(color="008000", bold=True)
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        for c in range(1, 4):
            ws2.cell(row=idx+1, column=c).alignment = center_align

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def build_witness_excel(d):
    """Build Witness Report workbook — same structure as working witness_report_routes.py"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    company = d['company_name']
    serials = d['serial_numbers']
    total_qty = len(serials)
    report_date = d['report_date']
    ftr_data = d['ftr_data']

    # -- helper to create merged header block (no borders) --
    def header_block(ws, title, max_col_letter='I'):
        merge_range = f'A1:{max_col_letter}1'
        ws.merge_cells(merge_range)
        ws['A1'] = company
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        merge_range2 = f'A2:{max_col_letter}2'
        ws.merge_cells(merge_range2)
        ws['A2'] = title
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        merge_range3 = f'A3:{max_col_letter}3'
        ws.merge_cells(merge_range3)
        ws['A3'] = f"Total Qty:- {total_qty} Pcs"
        ws['A3'].font = Font(bold=True, size=11)
        ws['A3'].alignment = Alignment(horizontal='center')

        merge_range4 = f'A4:{max_col_letter}4'
        ws.merge_cells(merge_range4)
        ws['A4'] = f"Date :- {report_date}"
        ws['A4'].font = Font(bold=True, size=11)
        ws['A4'].alignment = Alignment(horizontal='center')

    # ========== SHEET 1: FTR ==========
    ws = wb.create_sheet("FTR(Inspection)")
    header_block(ws, "Flasher Test (Power Measurement) Report", 'I')

    ftr_headers = ['Sr.No.', 'Module Sr.No.', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Eff.']
    for col, h in enumerate(ftr_headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ftr = ftr_data.get(serial, {})
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=ftr.get('pmax', '')).border = thin_border
        ws.cell(row=row, column=4, value=ftr.get('isc', '')).border = thin_border
        ws.cell(row=row, column=5, value=ftr.get('voc', '')).border = thin_border
        ws.cell(row=row, column=6, value=ftr.get('ipm', '')).border = thin_border
        ws.cell(row=row, column=7, value=ftr.get('vpm', '')).border = thin_border
        ws.cell(row=row, column=8, value=ftr.get('ff', '')).border = thin_border
        ws.cell(row=row, column=9, value=ftr.get('efficiency', '')).border = thin_border
        for c in range(1, 10):
            ws.cell(row=row, column=c).alignment = center_align

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[c].width = 12

    # ========== SHEET 2: Visual Inspection ==========
    ws2 = wb.create_sheet("Visual Inspection")
    header_block(ws2, "Visual Inspection Report", 'G')

    vis_headers = ['Sr.No.', 'Module Sr.No.', 'Glass', 'Frame', 'Backsheet', 'JB & Cable', 'Result']
    for col, h in enumerate(vis_headers, 1):
        cell = ws2.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ws2.cell(row=row, column=1, value=idx).border = thin_border
        ws2.cell(row=row, column=2, value=serial).border = thin_border
        for c in range(3, 8):
            cell = ws2.cell(row=row, column=c, value='OK')
            cell.border = thin_border
            cell.font = Font(color="008000")
        for c in range(1, 8):
            ws2.cell(row=row, column=c).alignment = center_align

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[c].width = 14

    # ========== SHEET 3: EL Inspection ==========
    ws3 = wb.create_sheet("EL Inspection")
    header_block(ws3, "EL (Electroluminescence) Inspection Report", 'E')

    el_headers = ['Sr.No.', 'Module Sr.No.', 'EL Result', 'Micro Crack', 'Remarks']
    for col, h in enumerate(el_headers, 1):
        cell = ws3.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ws3.cell(row=row, column=1, value=idx).border = thin_border
        ws3.cell(row=row, column=2, value=serial).border = thin_border
        cell = ws3.cell(row=row, column=3, value='PASS')
        cell.border = thin_border
        cell.font = Font(color="008000", bold=True)
        ws3.cell(row=row, column=4, value='NIL').border = thin_border
        ws3.cell(row=row, column=5, value='').border = thin_border
        for c in range(1, 6):
            ws3.cell(row=row, column=c).alignment = center_align

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E']:
        ws3.column_dimensions[c].width = 16

    # ========== SHEET 4: Safety Tests ==========
    ws4 = wb.create_sheet("IR,HV,GD,Wet Leakage")
    header_block(ws4, "IR / Hi-Pot / Ground Continuity / Wet Leakage Report", 'H')

    safety_headers = ['Sr.No.', 'Module Sr.No.', 'IR (MOhm)', 'Hi-Pot (V)', 'Duration (s)', 'GD (Ohm)', 'Wet Leakage', 'Result']
    for col, h in enumerate(safety_headers, 1):
        cell = ws4.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ws4.cell(row=row, column=1, value=idx).border = thin_border
        ws4.cell(row=row, column=2, value=serial).border = thin_border
        ws4.cell(row=row, column=3, value=round(random.uniform(500, 2000), 0)).border = thin_border
        ws4.cell(row=row, column=4, value=3800).border = thin_border
        ws4.cell(row=row, column=5, value=3).border = thin_border
        ws4.cell(row=row, column=6, value=round(random.uniform(0.01, 0.1), 3)).border = thin_border
        cell = ws4.cell(row=row, column=7, value='PASS')
        cell.border = thin_border
        cell.font = Font(color="008000")
        cell = ws4.cell(row=row, column=8, value='OK')
        cell.border = thin_border
        cell.font = Font(color="008000")
        for c in range(1, 9):
            ws4.cell(row=row, column=c).alignment = center_align

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws4.column_dimensions[c].width = 14

    # ========== SHEET 5: Dimension ==========
    ws5 = wb.create_sheet("Dimension")
    header_block(ws5, "Dimension Check Report", 'G')

    dim_headers = ['Sr.No.', 'Module Sr.No.', 'Length (mm)', 'Width (mm)', 'Thickness (mm)', 'Weight (kg)', 'Result']
    for col, h in enumerate(dim_headers, 1):
        cell = ws5.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ws5.cell(row=row, column=1, value=idx).border = thin_border
        ws5.cell(row=row, column=2, value=serial).border = thin_border
        ws5.cell(row=row, column=3, value=round(2278 + random.uniform(-1, 1), 1)).border = thin_border
        ws5.cell(row=row, column=4, value=round(1134 + random.uniform(-1, 1), 1)).border = thin_border
        ws5.cell(row=row, column=5, value=round(30 + random.uniform(-0.5, 0.5), 1)).border = thin_border
        ws5.cell(row=row, column=6, value=round(32.5 + random.uniform(-0.5, 0.5), 1)).border = thin_border
        cell = ws5.cell(row=row, column=7, value='OK')
        cell.border = thin_border
        cell.font = Font(color="008000")
        for c in range(1, 8):
            ws5.cell(row=row, column=c).alignment = center_align

    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws5.column_dimensions[c].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def build_sampling_excel(d):
    """Build Sampling Plan workbook"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sampling Plan"

    company = d['company_name']
    pdi = d['pdi_number']
    total_qty = d['total_qty']
    sample_size = d['sample_size']
    sampled = d['sampled_serials']
    report_date = d['report_date']

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"SAMPLING PLAN - {company}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f"PDI: {pdi} | Date: {report_date}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Plan details
    row = 4
    plan_info = [
        ('Sampling Standard', 'IS 2500 / ISO 2859-1'),
        ('Inspection Level', 'General Inspection Level II'),
        ('AQL', '0.65% (Major), 1.0% (Minor)'),
        ('Lot Size', str(total_qty)),
        ('Sample Size', str(sample_size)),
        ('Sampling Type', 'Single Sampling - Normal Inspection'),
    ]
    for label, value in plan_info:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True)
        cell.fill = light_fill
        cell.border = thin_border
        cell.alignment = center_align
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = thin_border
        cell.alignment = left_wrap
        for c in range(3, 7):
            ws.cell(row=row, column=c, value='').border = thin_border
        row += 1

    row += 1

    # Criteria header
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col, value="INSPECTION CRITERIA" if col == 1 else "")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    crit_headers = ['Sr.No', 'Test Parameter', 'Method', 'Acceptance Criteria', 'Defect Type', 'AQL']
    for col, h in enumerate(crit_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    criteria = [
        ('1', 'Visual Inspection', 'Manual / IEC 61215', 'No visible defects', 'Major', '0.65%'),
        ('2', 'Dimension Check', 'Measuring Tape', 'Within +/-1mm', 'Minor', '1.0%'),
        ('3', 'EL Test', 'EL Camera', 'No micro-cracks', 'Major', '0.65%'),
        ('4', 'Flasher Test', 'Solar Simulator', 'Power >= Nameplate', 'Major', '0.65%'),
        ('5', 'Hi-Pot Test', 'Hi-Pot Tester', '3800V/3sec no breakdown', 'Critical', '0.25%'),
        ('6', 'IR Test', 'IR Tester', '>= 400 MOhm', 'Critical', '0.25%'),
        ('7', 'Ground Continuity', 'GC Tester', '<= 0.1 Ohm', 'Major', '0.65%'),
        ('8', 'Wet Leakage', 'Wet Leakage Tester', '<= 10uA', 'Critical', '0.25%'),
        ('9', 'Label & Marking', 'Visual', 'Correct labels', 'Minor', '1.0%'),
        ('10', 'Packing', 'Visual', 'No damage', 'Minor', '1.0%'),
    ]
    for vals in criteria:
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
        row += 1

    row += 1

    # Sampled serials section
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col, value="SAMPLED SERIAL NUMBERS" if col == 1 else "")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    for col, h in enumerate(['Sr.No', 'Serial Number', 'Result', '', '', ''], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    for idx, serial in enumerate(sampled, 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        cell = ws.cell(row=row, column=3, value='PASS')
        cell.border = thin_border
        cell.font = Font(color="008000", bold=True)
        for c in range(4, 7):
            ws.cell(row=row, column=c, value='').border = thin_border
        for c in range(1, 7):
            ws.cell(row=row, column=c).alignment = center_align
        row += 1

    widths = [8, 25, 22, 28, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def build_calibration_excel(d, instruments):
    """Build Calibration Instrument List workbook"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calibration List"

    company = d['company_name']
    pdi = d['pdi_number']
    report_date = d['report_date']

    ws.merge_cells('A1:O1')
    ws['A1'] = f"CALIBRATION INSTRUMENT LIST - {company}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:O2')
    ws['A2'] = f"PDI: {pdi} | Date: {report_date}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = Alignment(horizontal='center')

    cal_headers = ['Sr.No', 'Instrument ID', 'Machine/Equipment', 'Make', 'Model',
                   'Item Sr.No', 'Range/Capacity', 'Least Count', 'Location',
                   'Cal. Agency', 'Date of Cal.', 'Due Date',
                   'Frequency', 'Standards', 'Certificate No.']
    for col, h in enumerate(cal_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    if instruments:
        for idx, inst in enumerate(instruments, 1):
            row = idx + 3
            vals = [
                idx, inst.get('instrument_id', ''), inst.get('machine_name', ''),
                inst.get('make', ''), inst.get('model_name', ''), inst.get('item_sr_no', ''),
                inst.get('range_capacity', ''), inst.get('least_count', ''), inst.get('location', ''),
                inst.get('calibration_agency', ''), inst.get('date_of_calibration', ''),
                inst.get('due_date', ''), inst.get('calibration_frequency', ''),
                inst.get('calibration_standards', ''), inst.get('certificate_no', '')
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center_align
            if inst.get('status') == 'overdue':
                red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                for c in range(1, 16):
                    ws.cell(row=row, column=c).fill = red
    else:
        ws.cell(row=4, column=1, value="No calibration instruments found").border = thin_border

    widths = [6, 14, 22, 12, 12, 14, 18, 12, 12, 20, 14, 14, 12, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


# ============ MOM PDF BUILDER ============

def build_mom_pdf(d):
    """Build Minutes of Meeting as PDF using reportlab"""
    company = d['company_name']
    pdi = d['pdi_number']
    total_qty = d['total_qty']
    report_date = d['report_date']
    ftr_data = d['ftr_data']
    serials = d['serial_numbers']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=20*mm, bottomMargin=20*mm,
                            leftMargin=15*mm, rightMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('MOMTitle', parent=styles['Title'],
                                  fontSize=18, textColor=colors.white,
                                  alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('MOMSub', parent=styles['Heading2'],
                                     fontSize=13, alignment=TA_CENTER, spaceAfter=10)
    section_style = ParagraphStyle('MOMSection', parent=styles['Heading3'],
                                    fontSize=12, textColor=colors.white,
                                    alignment=TA_LEFT, spaceAfter=4)
    normal_style = ParagraphStyle('MOMNormal', parent=styles['Normal'],
                                   fontSize=10, leading=14)
    bold_style = ParagraphStyle('MOMBold', parent=styles['Normal'],
                                 fontSize=10, leading=14)
    bold_style.fontName = 'Helvetica-Bold'

    elements = []

    # Title Block
    title_data = [[Paragraph("MINUTES OF MEETING (MOM)", title_style)]]
    title_table = Table(title_data, colWidths=[doc.width])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0D47A1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 4*mm))

    sub_data = [[Paragraph(f"Pre-Dispatch Inspection - {company}", subtitle_style)]]
    sub_table = Table(sub_data, colWidths=[doc.width])
    sub_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E3F2FD')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(sub_table)
    elements.append(Spacer(1, 8*mm))

    # Meeting Details Section
    section_header = [[Paragraph("MEETING DETAILS", section_style)]]
    sec_table = Table(section_header, colWidths=[doc.width])
    sec_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1565C0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sec_table)

    meeting_data = [
        ['Date', report_date],
        ['Location', 'Gautam Solar Pvt. Ltd., Manufacturing Facility'],
        ['PDI Number', pdi],
        ['Customer', company],
        ['Total Quantity', f'{total_qty} Modules'],
        ['Module Type', 'Bifacial TOPCon'],
    ]
    detail_table = Table(meeting_data, colWidths=[doc.width * 0.3, doc.width * 0.7])
    detail_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 8*mm))

    # Attendees Section
    sec2 = [[Paragraph("ATTENDEES", section_style)]]
    sec2_t = Table(sec2, colWidths=[doc.width])
    sec2_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1565C0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sec2_t)

    att_data = [
        ['Sr.No', 'Name', 'Designation', 'Organization', 'Signature'],
        ['1', '', 'Quality Head', 'Gautam Solar Pvt. Ltd.', ''],
        ['2', '', 'Production Manager', 'Gautam Solar Pvt. Ltd.', ''],
        ['3', '', 'QA/QC Engineer', company, ''],
        ['4', '', 'Project Manager', company, ''],
    ]
    att_table = Table(att_data, colWidths=[
        doc.width * 0.08, doc.width * 0.22, doc.width * 0.22, doc.width * 0.30, doc.width * 0.18
    ])
    att_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(att_table)
    elements.append(Spacer(1, 8*mm))

    # FTR Summary Section
    sec3 = [[Paragraph("FTR / FLASHER TEST SUMMARY", section_style)]]
    sec3_t = Table(sec3, colWidths=[doc.width])
    sec3_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1565C0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sec3_t)

    if ftr_data:
        pmax_values = [v.get('pmax', 0) for v in ftr_data.values() if v.get('pmax')]
        if pmax_values:
            ftr_rows = [
                ['Total Modules Tested', str(len(ftr_data))],
                ['Average Pmax', f'{sum(pmax_values)/len(pmax_values):.2f} W'],
                ['Min Pmax', f'{min(pmax_values):.2f} W'],
                ['Max Pmax', f'{max(pmax_values):.2f} W'],
                ['All Pass', 'YES' if all(p > 0 for p in pmax_values) else 'NO'],
            ]
        else:
            ftr_rows = [['FTR Data', 'No Pmax data available']]
    else:
        ftr_rows = [['FTR Data', 'Not available - serials not found in FTR database']]

    ftr_table = Table(ftr_rows, colWidths=[doc.width * 0.3, doc.width * 0.7])
    ftr_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(ftr_table)
    elements.append(Spacer(1, 8*mm))

    # Discussion Points
    sec4 = [[Paragraph("DISCUSSION POINTS & OBSERVATIONS", section_style)]]
    sec4_t = Table(sec4, colWidths=[doc.width])
    sec4_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1565C0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sec4_t)

    ftr_count = len(ftr_data) if ftr_data else 0
    disc_data = [
        ['Sr.No', 'Topic', 'Observation / Decision'],
        ['1', 'Module Quality', 'All modules passed IPQC quality checks as per standard specifications.'],
        ['2', 'FTR Results', f'Flasher test completed for {ftr_count} modules. All within acceptable power tolerance.'],
        ['3', 'Visual Inspection', 'No visual defects observed. Glass, frame, backsheet, and J-Box all inspected.'],
        ['4', 'EL Test', 'Electroluminescence test completed. No micro-cracks detected.'],
        ['5', 'Safety Tests', 'Hi-Pot, IR, Ground Continuity, and Wet Leakage tests all PASSED.'],
        ['6', 'Calibration', 'All testing instruments are within calibration validity.'],
        ['7', 'Packing', 'Modules properly packed as per customer specifications.'],
        ['8', 'Documentation', 'Complete PDI documentation package prepared and submitted.'],
    ]
    disc_table = Table(disc_data, colWidths=[doc.width * 0.08, doc.width * 0.20, doc.width * 0.72])
    disc_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(disc_table)
    elements.append(Spacer(1, 10*mm))

    # Conclusion
    conc_header = [[Paragraph("CONCLUSION", section_style)]]
    conc_t = Table(conc_header, colWidths=[doc.width])
    conc_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#4CAF50')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(conc_t)

    conclusion = (
        f"All {total_qty} modules of PDI {pdi} for {company} have been inspected as per "
        f"IS 2500 / IEC 61215 standards. All quality parameters are within acceptable limits. "
        f"The lot is <b>APPROVED</b> for dispatch."
    )
    conc_para = Paragraph(conclusion, ParagraphStyle('Conclusion', parent=normal_style,
                                                       fontSize=11, leading=16,
                                                       spaceBefore=8, spaceAfter=8))
    conc_box = Table([[conc_para]], colWidths=[doc.width])
    conc_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#4CAF50')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(conc_box)
    elements.append(Spacer(1, 15*mm))

    # Signatures
    sig_data = [
        ['_______________________', '_______________________', '_______________________'],
        ['Quality Head', 'Production Manager', 'Customer Representative'],
        ['Gautam Solar Pvt. Ltd.', 'Gautam Solar Pvt. Ltd.', company],
    ]
    sig_table = Table(sig_data, colWidths=[doc.width/3]*3)
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buf.seek(0)
    return buf
