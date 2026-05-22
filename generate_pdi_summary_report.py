import requests, json, sys
from datetime import datetime

API_BASE = sys.argv[1] if len(sys.argv) > 1 else 'http://localhost:5003'

companies = {
    3: 'Larsen & Toubro',
    4: 'Rays Power',
    5: 'Sterlin and Wilson',
}

all_data = {}

for cid, cname in companies.items():
    print(f'Fetching {cname}...')
    try:
        r = requests.get(f'{API_BASE}/api/ftr/pdi-production-status/{cid}?force_refresh=true', timeout=300)
        if r.status_code == 200:
            d = r.json()
            all_data[cname] = {
                'pdis': d.get('pdi_wise', []),
                'summary': d.get('summary', {}),
                'debug': d.get('debug_info', {})
            }
            print(f'  -> {len(all_data[cname]["pdis"])} PDIs')
        else:
            print(f'  -> Error {r.status_code}')
    except Exception as e:
        print(f'  -> Failed: {e}')

# Print table
print(f'\n{"="*120}')
print(f'PDI COMPLETE SUMMARY REPORT - {datetime.now().strftime("%Y-%m-%d %H:%M")}')
print(f'{"="*120}')

grand_total = {'produced': 0, 'ftr': 0, 'dispatched': 0, 'packed': 0, 'not_packed': 0}

for cname in ['Larsen & Toubro', 'Rays Power', 'Sterlin and Wilson']:
    if cname not in all_data:
        continue
    data = all_data[cname]
    pdis = data['pdis']
    summary = data['summary']
    
    print(f'\n{"-"*120}')
    print(f'  {cname}')
    print(f'{"-"*120}')
    print(f'  {"PDI":<12} {"Produced":>10} {"FTR":>10} {"Dispatched":>12} {"Packed":>10} {"Not Packed":>12} {"Disp%":>8}')
    print(f'  {"-"*100}')
    
    co_total = {'produced': 0, 'ftr': 0, 'dispatched': 0, 'packed': 0, 'not_packed': 0}
    
    for p in pdis:
        disp = p.get('dispatched', 0)
        packed = p.get('packed', 0)
        not_packed = p.get('not_packed', 0)
        ftr = p.get('ftr_tested', 0)
        produced = p.get('produced', 0)
        total = disp + packed + not_packed
        disp_pct = round((disp / total * 100), 1) if total > 0 else 0
        
        print(f'  {p["pdi_number"]:<12} {produced:>10} {ftr:>10} {disp:>12} {packed:>10} {not_packed:>12} {disp_pct:>7}%')
        
        co_total['produced'] += produced
        co_total['ftr'] += ftr
        co_total['dispatched'] += disp
        co_total['packed'] += packed
        co_total['not_packed'] += not_packed
    
    total = co_total['dispatched'] + co_total['packed'] + co_total['not_packed']
    disp_pct = round((co_total['dispatched'] / total * 100), 1) if total > 0 else 0
    print(f'  {"-"*100}')
    print(f'  {"TOTAL":<12} {co_total["produced"]:>10} {co_total["ftr"]:>10} {co_total["dispatched"]:>12} {co_total["packed"]:>10} {co_total["not_packed"]:>12} {disp_pct:>7}%')
    
    for k in grand_total:
        grand_total[k] += co_total[k]

print(f'\n{"="*120}')
total = grand_total['dispatched'] + grand_total['packed'] + grand_total['not_packed']
disp_pct = round((grand_total['dispatched'] / total * 100), 1) if total > 0 else 0
print(f'  GRAND TOTAL: Produced={grand_total["produced"]} FTR={grand_total["ftr"]} Dispatched={grand_total["dispatched"]} Packed={grand_total["packed"]} Not Packed={grand_total["not_packed"]} ({disp_pct}% dispatched)')
print(f'{"="*120}')

# Generate Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    
    # Summary Sheet
    ws = wb.active
    ws.title = 'PDI Summary'
    
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    
    row = 1
    ws.cell(row=row, column=1, value=f'PDI Complete Summary Report - {datetime.now().strftime("%Y-%m-%d %H:%M")}').font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    
    row = 3
    headers = ['Company', 'PDI', 'Produced', 'FTR Assigned', 'Dispatched', 'Packed', 'Not Packed', 'Dispatch %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    for cname in ['Larsen & Toubro', 'Rays Power', 'Sterlin and Wilson']:
        if cname not in all_data:
            continue
        pdis = all_data[cname]['pdis']
        for p in pdis:
            disp = p.get('dispatched', 0)
            packed = p.get('packed', 0)
            not_packed = p.get('not_packed', 0)
            total = disp + packed + not_packed
            disp_pct = round((disp / total * 100), 1) if total > 0 else 0
            ws.cell(row=row, column=1, value=cname)
            ws.cell(row=row, column=2, value=p['pdi_number'])
            ws.cell(row=row, column=3, value=p.get('produced', 0))
            ws.cell(row=row, column=4, value=p.get('ftr_tested', 0))
            ws.cell(row=row, column=5, value=disp)
            ws.cell(row=row, column=6, value=packed)
            ws.cell(row=row, column=7, value=not_packed)
            ws.cell(row=row, column=8, value=disp_pct)
            row += 1
    
    # Add totals
    bold_font = Font(bold=True)
    for cname in ['Larsen & Toubro', 'Rays Power', 'Sterlin and Wilson']:
        if cname not in all_data:
            continue
        ws.cell(row=row, column=1, value=f'{cname} TOTAL').font = bold_font
        ws.cell(row=row, column=3, value=sum(p.get('produced',0) for p in all_data[cname]['pdis'])).font = bold_font
        ws.cell(row=row, column=4, value=sum(p.get('ftr_tested',0) for p in all_data[cname]['pdis'])).font = bold_font
        ws.cell(row=row, column=5, value=sum(p.get('dispatched',0) for p in all_data[cname]['pdis'])).font = bold_font
        ws.cell(row=row, column=6, value=sum(p.get('packed',0) for p in all_data[cname]['pdis'])).font = bold_font
        ws.cell(row=row, column=7, value=sum(p.get('not_packed',0) for p in all_data[cname]['pdis'])).font = bold_font
        row += 1
    
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
    
    # Per-company sheets
    for cname in ['Larsen & Toubro', 'Rays Power', 'Sterlin and Wilson']:
        if cname not in all_data:
            continue
        ws2 = wb.create_sheet(cname[:31])
        pdis = all_data[cname]['pdis']
        
        ws2.cell(row=1, column=1, value=f'{cname} - PDI Detailed Report').font = title_font
        ws2.merge_cells('A1:H1')
        
        headers2 = ['PDI', 'Produced', 'FTR Assigned', 'Dispatched', 'Packed', 'Not Packed', 'Dispatch %', 'Assigned Date']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        r = 4
        for p in pdis:
            disp = p.get('dispatched', 0)
            packed = p.get('packed', 0)
            not_packed = p.get('not_packed', 0)
            total = disp + packed + not_packed
            disp_pct = round((disp / total * 100), 1) if total > 0 else 0
            ws2.cell(row=r, column=1, value=p['pdi_number'])
            ws2.cell(row=r, column=2, value=p.get('produced', 0))
            ws2.cell(row=r, column=3, value=p.get('ftr_tested', 0))
            ws2.cell(row=r, column=4, value=disp)
            ws2.cell(row=r, column=5, value=packed)
            ws2.cell(row=r, column=6, value=not_packed)
            ws2.cell(row=r, column=7, value=disp_pct)
            ws2.cell(row=r, column=8, value=p.get('assigned_date', ''))
            r += 1
        
        # Auto-width
        for col in ws2.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
    
    filename = f'PDI_Summary_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(filename)
    print(f'\n✅ Excel report saved: {filename}')
except ImportError:
    print('\n⚠️  openpyxl not installed. Install with: pip install openpyxl')
    print('   Table output printed above instead.')
except Exception as e:
    print(f'\n❌ Excel error: {e}')
