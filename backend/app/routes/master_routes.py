"""
Master Data Routes - Upload & Manage Pre-generated Production Data
"""

from flask import Blueprint, request, jsonify
from app.models.master_data import MasterOrder, MasterModule, DailyProduction
from app.models.database import db
from datetime import datetime
import random

master_bp = Blueprint('master', __name__, url_prefix='/api/master')

@master_bp.route('/upload-excel', methods=['POST'])
def upload_excel_data():
    """
    Upload Excel file with RFID FTR data
    
    Excel Format (RFID format):
    Column A: Date
    Column B: ID (Serial Number)
    Column C-U: Pmax, Isc, Voc, Ipm, Vpm, FF, Rs, Rsh, Eff, T_Object, T_Target, Irr_Target, Class, Sweep_Time, Irr_Monitor, Isc_Monitor, T_Monitor, Cell_Temp, T_Ambient, Binning
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Only Excel files allowed'}), 400
        
        # Get form data
        company_name = request.form.get('company_name')
        order_number = request.form.get('order_number')
        
        if not company_name or not order_number:
            return jsonify({'error': 'company_name and order_number required'}), 400
        
        # Check if order already exists
        existing_order = MasterOrder.query.filter_by(order_number=order_number).first()
        if existing_order:
            return jsonify({'error': 'Order number already exists'}), 400
        
        # Read Excel file
        import pandas as pd
        import io
        
        # Read Excel in chunks for large files
        file_content = file.read()
        df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Check if first row contains header keywords
        first_row_str = ' '.join([str(val).lower() for val in df.iloc[0].values if pd.notna(val)])
        if 'pmax' in first_row_str or 'date' in first_row_str or 'id' in first_row_str:
            # First row is header, use it as column names
            df.columns = df.iloc[0]
            df = df.drop(df.index[0]).reset_index(drop=True)
        
        # Normalize column names
        df.columns = df.columns.str.strip()
        
        # Validate max rows (500,000 limit)
        if len(df) > 500000:
            return jsonify({'error': f'Excel file has {len(df)} rows. Maximum 500,000 rows allowed.'}), 400
        
        # Check if ID column exists
        if 'ID' not in df.columns:
            return jsonify({'error': 'Missing required column: ID (Serial Number)'}), 400
        
        # Remove rows where ID is empty or invalid
        df = df[df['ID'].notna()]
        df = df[df['ID'].astype(str).str.strip() != '']
        
        # Extract serial prefix from first serial
        first_serial = str(df['ID'].iloc[0]).strip()
        # Assuming format like GS04890TG3002500001
        serial_prefix = ''.join([c for c in first_serial if not c.isdigit()])
        
        # Create master order
        order = MasterOrder(
            company_name=company_name,
            order_number=order_number,
            total_modules=len(df),
            serial_prefix=serial_prefix,
            rejection_percentage=0.0  # All FTR data
        )
        db.session.add(order)
        db.session.flush()
        
        # Store progress in session/cache for polling
        import json
        import tempfile
        import os
        temp_dir = tempfile.gettempdir()
        progress_file = os.path.join(temp_dir, f'upload_progress_{order.id}.json')
        
        def update_progress(current, total, status):
            try:
                with open(progress_file, 'w') as f:
                    json.dump({
                        'current': current,
                        'total': total,
                        'status': status,
                        'percent': int((current / total) * 100) if total > 0 else 0
                    }, f)
            except:
                pass
        
        # Process modules in larger batches for performance
        batch_size = 5000
        modules_batch = []
        processed = 0
        total_rows = len(df)
        
        update_progress(0, total_rows, 'Starting processing...')
        
        for idx, row in df.iterrows():
            processed += 1
            if processed % 5000 == 0:
                print(f"Processed {processed}/{total_rows} rows...")
                update_progress(processed, total_rows, f'Processing rows... {processed:,} / {total_rows:,}')
            # Get serial number from ID column
            serial = str(row['ID']).strip() if pd.notna(row['ID']) else ''
            
            if not serial:
                continue
            
            module = MasterModule(
                order_id=order.id,
                serial_number=serial,
                sequence_number=idx + 1,
                is_rejected=False,  # All FTR data
                
                # Store all FTR parameters
                date=str(row.get('Date', '')) if pd.notna(row.get('Date')) else None,
                pmax=float(row['Pmax']) if pd.notna(row.get('Pmax')) else None,
                isc=float(row['Isc']) if pd.notna(row.get('Isc')) else None,
                voc=float(row['Voc']) if pd.notna(row.get('Voc')) else None,
                ipm=float(row['Ipm']) if pd.notna(row.get('Ipm')) else None,
                vpm=float(row['Vpm']) if pd.notna(row.get('Vpm')) else None,
                ff=float(row['FF']) if pd.notna(row.get('FF')) else None,
                rs=float(row['Rs']) if pd.notna(row.get('Rs')) else None,
                rsh=float(row['Rsh']) if pd.notna(row.get('Rsh')) else None,
                eff=float(row['Eff']) if pd.notna(row.get('Eff')) else None,
                t_object=float(row['T_Object']) if pd.notna(row.get('T_Object')) else None,
                t_target=float(row['T_Target']) if pd.notna(row.get('T_Target')) else None,
                irr_target=float(row['Irr_Target']) if pd.notna(row.get('Irr_Target')) else None,
                class_grade=str(row.get('Class', '')) if pd.notna(row.get('Class')) else None,
                sweep_time=float(row['Sweep_Time']) if pd.notna(row.get('Sweep_Time')) else None,
                irr_monitor=float(row['Irr_Monitor']) if pd.notna(row.get('Irr_Monitor')) else None,
                isc_monitor=float(row['Isc_Monitor']) if pd.notna(row.get('Isc_Monitor')) else None,
                t_monitor=float(row['T_Monitor']) if pd.notna(row.get('T_Monitor')) else None,
                cell_temp=float(row['Cell_Temp']) if pd.notna(row.get('Cell_Temp')) else None,
                t_ambient=float(row['T_Ambient']) if pd.notna(row.get('T_Ambient')) else None,
                binning=str(row.get('Binning', '')) if pd.notna(row.get('Binning')) else None
            )
            modules_batch.append(module)
            
            # Bulk insert
            if len(modules_batch) >= batch_size:
                db.session.bulk_save_objects(modules_batch)
                db.session.commit()
                modules_batch = []
        
        # Insert remaining
        if modules_batch:
            db.session.bulk_save_objects(modules_batch)
        
        db.session.commit()
        
        update_progress(total_rows, total_rows, 'Complete!')
        
        # Cleanup progress file
        try:
            import os
            if os.path.exists(progress_file):
                os.remove(progress_file)
        except:
            pass
        
        return jsonify({
            'message': 'FTR Excel data uploaded successfully',
            'order': {
                'id': order.id,
                'company_name': order.company_name,
                'order_number': order.order_number,
                'total_modules': len(df),
                'ftr_count': len(df),
                'rejection_count': 0,
                'serial_prefix': serial_prefix
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@master_bp.route('/upload-progress/<int:order_id>', methods=['GET'])
def get_upload_progress(order_id):
    """Get upload progress for an order"""
    try:
        import json
        import os
        import tempfile
        temp_dir = tempfile.gettempdir()
        progress_file = os.path.join(temp_dir, f'upload_progress_{order_id}.json')
        
        if os.path.exists(progress_file):
            with open(progress_file, 'r') as f:
                progress = json.load(f)
            return jsonify(progress), 200
        else:
            return jsonify({'current': 0, 'total': 0, 'status': 'No upload in progress', 'percent': 0}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@master_bp.route('/orders', methods=['GET'])
def get_master_orders():
    """Get all master orders"""
    try:
        orders = MasterOrder.query.all()
        
        orders_list = []
        for order in orders:
            total = order.total_modules
            produced = MasterModule.query.filter_by(order_id=order.id, is_produced=True).count()
            rejected = MasterModule.query.filter_by(order_id=order.id, is_rejected=True).count()
            delivered = MasterModule.query.filter_by(order_id=order.id, is_delivered=True).count()
            
            orders_list.append({
                'id': order.id,
                'company_name': order.company_name,
                'order_number': order.order_number,
                'total_modules': total,
                'produced_modules': produced,
                'remaining_modules': total - produced,
                'rejection_count': rejected,
                'delivered_modules': delivered,
                'rejection_percentage': order.rejection_percentage,
                'created_at': order.created_at.isoformat()
            })
        
        return jsonify({'orders': orders_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@master_bp.route('/daily-production', methods=['POST'])
def assign_daily_production():
    """
    Assign modules from master data to daily production
    
    Request body:
    {
        "order_id": 1,
        "production_date": "2025-11-28",
        "shift": "A",
        "line_number": 1,
        "modules_count": 500
    }
    """
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['order_id', 'production_date', 'shift', 'line_number', 'modules_count']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        order_id = data['order_id']
        production_date = datetime.strptime(data['production_date'], '%Y-%m-%d').date()
        shift = data['shift']
        line_number = data['line_number']
        modules_count = data['modules_count']
        
        # Get order
        order = MasterOrder.query.get(order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Get next available modules (not yet produced)
        available_modules = MasterModule.query.filter_by(
            order_id=order_id,
            is_produced=False
        ).order_by(MasterModule.sequence_number).limit(modules_count).all()
        
        if len(available_modules) < modules_count:
            return jsonify({
                'error': f'Only {len(available_modules)} modules available, requested {modules_count}'
            }), 400
        
        # Assign production data
        ftr_count = 0
        rejection_count = 0
        
        for module in available_modules:
            module.is_produced = True
            module.production_date = production_date
            module.production_shift = shift
            module.line_number = line_number
            
            if module.is_rejected:
                rejection_count += 1
            else:
                ftr_count += 1
        
        # Create daily production record
        daily_prod = DailyProduction(
            order_id=order_id,
            production_date=production_date,
            shift=shift,
            line_number=line_number,
            modules_produced=modules_count,
            ftr_count=ftr_count,
            rejection_count=rejection_count
        )
        db.session.add(daily_prod)
        db.session.commit()
        
        return jsonify({
            'message': 'Daily production assigned successfully',
            'production': {
                'date': production_date.isoformat(),
                'shift': shift,
                'line_number': line_number,
                'modules_produced': modules_count,
                'ftr_count': ftr_count,
                'rejection_count': rejection_count,
                'rejection_serials': [m.serial_number for m in available_modules if m.is_rejected]
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@master_bp.route('/daily-report/<int:order_id>', methods=['GET'])
def get_daily_report(order_id):
    """
    Get daily production report for specific date
    Query params: date (YYYY-MM-DD), shift (optional)
    """
    try:
        date_str = request.args.get('date')
        shift = request.args.get('shift')
        
        if not date_str:
            return jsonify({'error': 'Date parameter required'}), 400
        
        production_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Build query
        query = DailyProduction.query.filter_by(
            order_id=order_id,
            production_date=production_date
        )
        
        if shift:
            query = query.filter_by(shift=shift)
        
        productions = query.all()
        
        if not productions:
            return jsonify({'error': 'No production data found for this date'}), 404
        
        # Get rejection details
        rejection_modules = MasterModule.query.filter_by(
            order_id=order_id,
            production_date=production_date,
            is_rejected=True,
            is_produced=True
        )
        
        if shift:
            rejection_modules = rejection_modules.filter_by(production_shift=shift)
        
        rejection_modules = rejection_modules.all()
        
        # Compile report
        total_produced = sum(p.modules_produced for p in productions)
        total_ftr = sum(p.ftr_count for p in productions)
        total_rejected = sum(p.rejection_count for p in productions)
        
        report = {
            'order_id': order_id,
            'date': production_date.isoformat(),
            'shift': shift if shift else 'ALL',
            'summary': {
                'total_produced': total_produced,
                'ftr_count': total_ftr,
                'rejection_count': total_rejected,
                'rejection_percentage': round((total_rejected / total_produced * 100), 2) if total_produced > 0 else 0
            },
            'rejections': [
                {
                    'serial_number': m.serial_number,
                    'reason': m.rejection_reason,
                    'shift': m.production_shift,
                    'line': m.line_number
                }
                for m in rejection_modules
            ],
            'line_wise': []
        }
        
        # Line-wise breakdown
        for prod in productions:
            report['line_wise'].append({
                'line_number': prod.line_number,
                'shift': prod.shift,
                'modules_produced': prod.modules_produced,
                'ftr_count': prod.ftr_count,
                'rejection_count': prod.rejection_count
            })
        
        return jsonify({'report': report}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@master_bp.route('/upload-rejections', methods=['POST'])
def upload_rejections():
    """
    Upload simple Excel with rejection serial numbers
    
    Excel Format:
    Column A: Serial Number
    Optional Column B: Rejection Reason
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        order_id = request.form.get('order_id')
        
        if not order_id:
            return jsonify({'error': 'order_id required'}), 400
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Only Excel files allowed'}), 400
        
        # Check if order exists
        order = MasterOrder.query.get(order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Read Excel file
        import pandas as pd
        import io
        
        file_content = file.read()
        df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Get first column as serial numbers
        if len(df.columns) == 0:
            return jsonify({'error': 'Excel file is empty'}), 400
        
        # First column is serial number
        serial_col = df.columns[0]
        reason_col = df.columns[1] if len(df.columns) > 1 else None
        
        rejected_count = 0
        not_found = []
        
        for idx, row in df.iterrows():
            serial = str(row[serial_col]).strip() if pd.notna(row[serial_col]) else ''
            
            if not serial:
                continue
            
            # Find module in this order
            module = MasterModule.query.filter_by(
                order_id=order_id,
                serial_number=serial
            ).first()
            
            if module:
                module.is_rejected = True
                if reason_col and pd.notna(row[reason_col]):
                    module.rejection_reason = str(row[reason_col])
                else:
                    module.rejection_reason = 'Rejected via upload'
                rejected_count += 1
            else:
                not_found.append(serial)
        
        db.session.commit()
        
        response_data = {
            'message': f'Successfully marked {rejected_count} modules as rejected',
            'rejected_count': rejected_count,
            'order_id': order_id
        }
        
        if not_found:
            response_data['warning'] = f'{len(not_found)} serial numbers not found in order'
            response_data['not_found_serials'] = not_found[:10]  # First 10 only
        
        return jsonify(response_data), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@master_bp.route('/upload-delivered-ftr', methods=['POST'])
def upload_delivered_ftr():
    """
    Upload Excel with FTR serial numbers that have already been delivered to customer
    These serials will be skipped in future FTR downloads
    
    Excel Format:
    Column A: Serial Number (required)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        order_id = request.form.get('order_id')
        
        if not order_id:
            return jsonify({'error': 'order_id required'}), 400
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Only Excel files allowed'}), 400
        
        # Check if order exists
        order = MasterOrder.query.get(order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Read Excel file
        import pandas as pd
        import io
        
        file_content = file.read()
        df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Get first column as serial numbers
        if len(df.columns) == 0:
            return jsonify({'error': 'Excel file is empty'}), 400
        
        # First column is serial number
        serial_col = df.columns[0]
        
        delivered_count = 0
        not_found = []
        already_rejected = []
        
        for idx, row in df.iterrows():
            serial = str(row[serial_col]).strip() if pd.notna(row[serial_col]) else ''
            
            if not serial:
                continue
            
            # Find module in this order
            module = MasterModule.query.filter_by(
                order_id=order_id,
                serial_number=serial
            ).first()
            
            if module:
                if module.is_rejected:
                    already_rejected.append(serial)
                else:
                    module.is_delivered = True
                    module.delivered_date = datetime.now().date()
                    delivered_count += 1
            else:
                not_found.append(serial)
        
        db.session.commit()
        
        response_data = {
            'message': f'Successfully marked {delivered_count} modules as delivered',
            'delivered_count': delivered_count,
            'order_id': order_id
        }
        
        if not_found:
            response_data['warning'] = f'{len(not_found)} serial numbers not found in order'
            response_data['not_found_serials'] = not_found[:10]  # First 10 only
        
        if already_rejected:
            response_data['rejected_warning'] = f'{len(already_rejected)} serials were already rejected (skipped)'
            response_data['rejected_serials'] = already_rejected[:10]  # First 10 only
        
        return jsonify(response_data), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@master_bp.route('/download-ftr-by-quantity', methods=['POST'])
def download_ftr_by_quantity():
    """
    Download FTR Excel starting from a serial number with specific quantity
    
    Request body:
    {
        "order_id": 1,
        "start_serial": "GS04890TG3002500001",
        "quantity": 2832
    }
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import tempfile
        from flask import send_file
        
        data = request.json
        order_id = data.get('order_id')
        start_serial = data.get('start_serial')
        quantity = data.get('quantity')
        
        if not order_id or not start_serial or not quantity:
            return jsonify({'error': 'order_id, start_serial and quantity required'}), 400
        
        # Get order
        order = MasterOrder.query.get(order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Get modules starting from start_serial, only non-rejected and non-delivered, limit by quantity
        modules = MasterModule.query.filter(
            MasterModule.order_id == order_id,
            MasterModule.is_rejected == False,
            MasterModule.is_delivered == False,
            MasterModule.serial_number >= start_serial
        ).order_by(MasterModule.serial_number).limit(quantity).all()
        
        if not modules:
            return jsonify({'error': 'No modules found starting from given serial'}), 404
        
        if len(modules) < quantity:
            return jsonify({
                'error': f'Only {len(modules)} available modules (non-rejected & non-delivered) starting from {start_serial}. Requested: {quantity}'
            }), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'FTR Data'
        
        # Header style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers - Only required columns
        headers = ['SN', 'ID', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Rs', 'Eff', 'Binning']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, module in enumerate(modules, 2):
            ws.cell(row=row_num, column=1).value = row_num - 1  # SN
            ws.cell(row=row_num, column=2).value = module.serial_number  # ID
            ws.cell(row=row_num, column=3).value = module.pmax
            ws.cell(row=row_num, column=4).value = module.isc
            ws.cell(row=row_num, column=5).value = module.voc
            ws.cell(row=row_num, column=6).value = module.ipm
            ws.cell(row=row_num, column=7).value = module.vpm
            ws.cell(row=row_num, column=8).value = module.ff
            ws.cell(row=row_num, column=9).value = module.rs
            ws.cell(row=row_num, column=10).value = module.eff
            ws.cell(row=row_num, column=11).value = module.binning
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        filename = f'FTR_Data_{order.order_number}_{quantity}_modules.xlsx'
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@master_bp.route('/download-ftr-by-serials', methods=['POST'])
def download_ftr_by_serials():
    """
    Download FTR Excel for specific serial numbers
    
    Request body:
    {
        "order_id": 1,
        "serial_numbers": ["GS04890TG3002500001", "GS04890TG3002500002", ...]
        OR
        "serial_range": {
            "start": "GS04890TG3002500001",
            "end": "GS04890TG3002500100"
        }
    }
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import tempfile
        from flask import send_file
        
        data = request.json
        order_id = data.get('order_id')
        serial_numbers = data.get('serial_numbers', [])
        serial_range = data.get('serial_range')
        
        if not order_id:
            return jsonify({'error': 'order_id required'}), 400
        
        # Get order
        order = MasterOrder.query.get(order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Build query - ONLY NON-REJECTED MODULES
        if serial_range:
            # Range query
            start_serial = serial_range.get('start')
            end_serial = serial_range.get('end')
            
            modules = MasterModule.query.filter(
                MasterModule.order_id == order_id,
                MasterModule.is_rejected == False,
                MasterModule.is_delivered == False,
                MasterModule.serial_number >= start_serial,
                MasterModule.serial_number <= end_serial
            ).order_by(MasterModule.serial_number).all()
        elif serial_numbers:
            # Specific serials - ONLY NON-REJECTED
            modules = MasterModule.query.filter(
                MasterModule.order_id == order_id,
                MasterModule.is_rejected == False,
                MasterModule.is_delivered == False,
                MasterModule.serial_number.in_(serial_numbers)
            ).order_by(MasterModule.serial_number).all()
        else:
            return jsonify({'error': 'serial_numbers or serial_range required'}), 400
        
        if not modules:
            return jsonify({'error': 'No modules found for given serials'}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'FTR Data'
        
        # Header style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers - Only required columns
        headers = ['SN', 'ID', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Rs', 'Eff', 'Binning']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows - Only required columns
        for row_num, module in enumerate(modules, 2):
            ws.cell(row=row_num, column=1).value = row_num - 1  # SN (serial count starting from 1)
            ws.cell(row=row_num, column=2).value = module.serial_number  # ID
            ws.cell(row=row_num, column=3).value = module.pmax
            ws.cell(row=row_num, column=4).value = module.isc
            ws.cell(row=row_num, column=5).value = module.voc
            ws.cell(row=row_num, column=6).value = module.ipm
            ws.cell(row=row_num, column=7).value = module.vpm
            ws.cell(row=row_num, column=8).value = module.ff
            ws.cell(row=row_num, column=9).value = module.rs
            ws.cell(row=row_num, column=10).value = module.eff
            ws.cell(row=row_num, column=11).value = module.binning
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        filename = f'FTR_Data_{order.order_number}_{len(modules)}_modules.xlsx'
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@master_bp.route('/modules/<int:order_id>', methods=['GET'])
def get_modules(order_id):
    """Get modules for an order with pagination and search"""
    try:
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        search = request.args.get('search', '')
        
        query = MasterModule.query.filter_by(order_id=order_id)
        
        if search:
            query = query.filter(MasterModule.serial_number.like(f'%{search}%'))
        
        total = query.count()
        modules = query.order_by(MasterModule.sequence_number).offset(offset).limit(limit).all()
        
        modules_list = []
        for m in modules:
            modules_list.append({
                'id': m.id,
                'serial_number': m.serial_number,
                'date': m.date,
                'pmax': m.pmax,
                'isc': m.isc,
                'voc': m.voc,
                'ipm': m.ipm,
                'vpm': m.vpm,
                'ff': m.ff,
                'rs': m.rs,
                'rsh': m.rsh,
                'eff': m.eff,
                't_object': m.t_object,
                't_target': m.t_target,
                'irr_target': m.irr_target,
                'class_grade': m.class_grade,
                'sweep_time': m.sweep_time,
                'irr_monitor': m.irr_monitor,
                'isc_monitor': m.isc_monitor,
                't_monitor': m.t_monitor,
                'cell_temp': m.cell_temp,
                't_ambient': m.t_ambient,
                'binning': m.binning
            })
        
        return jsonify({'modules': modules_list, 'total': total}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@master_bp.route('/module/<int:module_id>', methods=['PUT'])
def update_module(module_id):
    """Update a module"""
    try:
        data = request.json
        module = MasterModule.query.get(module_id)
        
        if not module:
            return jsonify({'error': 'Module not found'}), 404
        
        # Update fields
        if 'serial_number' in data:
            module.serial_number = data['serial_number']
        if 'pmax' in data:
            module.pmax = data['pmax']
        if 'isc' in data:
            module.isc = data['isc']
        if 'voc' in data:
            module.voc = data['voc']
        if 'eff' in data:
            module.eff = data['eff']
        if 'class_grade' in data:
            module.class_grade = data['class_grade']
        
        db.session.commit()
        
        return jsonify({'message': 'Module updated successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@master_bp.route('/module/<int:module_id>', methods=['DELETE'])
def delete_module(module_id):
    """Delete a module"""
    try:
        module = MasterModule.query.get(module_id)
        
        if not module:
            return jsonify({'error': 'Module not found'}), 404
        
        db.session.delete(module)
        db.session.commit()
        
        return jsonify({'message': 'Module deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@master_bp.route('/order/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    """Delete an order and all its modules"""
    try:
        order = MasterOrder.query.get(order_id)
        
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        # Delete all modules first (cascade should handle this, but explicit is better)
        MasterModule.query.filter_by(order_id=order_id).delete()
        
        # Delete order
        db.session.delete(order)
        db.session.commit()
        
        return jsonify({'message': 'Order and all modules deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
