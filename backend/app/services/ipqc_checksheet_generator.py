"""
IPQC Check Sheet Generator — Exact replica of "IPQC Check Sheet.xlsx"
Generates a filled IPQC checksheet in the original Gautam Solar format.
140 rows × 15 columns (A-O), single sheet named "IPQC".
"""

import os
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────
THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray (theme 0 tint -0.15)
FONT_TITLE = Font(name='Calibri', size=16, bold=True)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)
FONT_NORMAL = Font(name='Calibri', size=11)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER_V = Alignment(vertical='center', wrap_text=True)


# ──────────────────────────────────────────────
# Column widths  (from reference)
# ──────────────────────────────────────────────
COL_WIDTHS = {
    'A': 5.22, 'B': 14.55, 'C': 22.22, 'D': 11.55, 'E': 15.22,
    'F': 20.22, 'G': 17.56, 'H': 8.78, 'I': 13.0, 'J': 15.89,
    'K': 8.78, 'L': 13.0, 'M': 11.44, 'N': 13.66, 'O': 16.44,
}

# ──────────────────────────────────────────────
# Row heights  (from reference)
# ──────────────────────────────────────────────
ROW_HEIGHTS = {
    3: 21, 7: 28.2, 8: 28.2, 9: 28.2, 10: 28.2, 11: 28.2, 12: 28.2,
    13: 43.5, 14: 61.2, 15: 25.5, 16: 30.6, 18: 28.8, 19: 28.8, 20: 27,
    21: 28.8, 22: 14.55, 23: 23.55, 25: 22.5, 26: 15.45, 27: 18, 28: 13.8,
    29: 27, 30: 26.55, 31: 28.2, 32: 24.45, 33: 24.45, 34: 24.45, 35: 21,
    36: 42, 37: 26.4, 38: 38.4, 39: 31.5, 40: 56.1, 41: 46.2, 42: 39.6,
    43: 64.2, 44: 18, 45: 28.95, 46: 51.45, 47: 15.6, 48: 28.05,
    49: 15.6, 50: 24, 51: 48.6, 52: 48.6, 53: 28.8, 54: 28.8, 55: 28.8,
    56: 28.8, 57: 28.8, 58: 49.8, 59: 49.8, 60: 49.8, 61: 49.8, 62: 34.2,
    63: 47.4, 64: 53.55, 65: 70.05, 66: 27, 67: 62.4, 68: 28.8, 69: 28.8,
    70: 28.8, 71: 28.8, 72: 28.8, 73: 26.4, 74: 24.6, 75: 24.6, 76: 24.6,
    77: 24.6, 78: 24.6, 79: 46.05, 80: 31.5, 81: 26.25, 82: 16.5,
    83: 24, 84: 38.1, 85: 37.5, 86: 37.5, 87: 42.6, 88: 42.6,
    89: 40.8, 90: 40.8, 91: 40.8, 92: 47.4, 93: 24.6, 94: 27, 95: 26.1,
    96: 36.6, 97: 28.8, 98: 28.8, 99: 28.8, 100: 28.8, 101: 28.8,
    102: 37.5, 103: 50.55, 104: 61.5, 105: 23.55, 106: 49.05, 107: 34.2,
    108: 26.4, 109: 26.4, 110: 26.4, 111: 26.4, 112: 26.4,
    113: 43.2, 114: 22.8, 115: 22.8, 116: 22.8, 117: 22.8, 118: 22.8,
    119: 24, 120: 43.05, 121: 25.2, 122: 25.2, 123: 25.2, 124: 25.2, 125: 25.2,
    126: 33.6, 127: 33.6, 128: 33.6, 129: 33.6, 130: 33.6,
    131: 30, 132: 30, 133: 30, 134: 30, 135: 30,
    136: 45, 137: 45, 138: 45, 139: 45, 140: 45,
}


# ──────────────────────────────────────────────
# Helper functions
# ──────────────────────────────────────────────
def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Write to cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _rr(base, tol):
    """Random realistic value within base ± tol."""
    return round(base + random.uniform(-abs(tol), abs(tol)), 2)


def _gen_serials(prefix, start, count=5):
    """Generate sorted serial number strings."""
    pool = list(range(start, min(start + 200, 100000)))
    picks = sorted(random.sample(pool, min(count, len(pool))))
    return [f"{prefix}{str(n).zfill(5)}" for n in picks]


# ──────────────────────────────────────────────
# Main generator
# ──────────────────────────────────────────────
def generate_ipqc_checksheet(
    date=None,
    shift='A',
    po_number='',
    cell_manufacturer='Solar Space',
    cell_efficiency=25.7,
    jb_cable_length=1200,
    golden_module_number='GM-2024-001',
    serial_prefix='GS04875KG302250',
    serial_start=1,
    module_count=1,
    customer_id='GSPL/IPQC/IPC/003',
    checked_by='',
    reviewed_by='',
):
    """
    Generate a filled IPQC Check Sheet in the exact reference format.
    Returns the file path of the generated .xlsx file.
    """
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')

    wb = Workbook()
    ws = wb.active
    ws.title = 'IPQC'

    # ── Column widths ──
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row heights ──
    for r, h in ROW_HEIGHTS.items():
        ws.row_dimensions[r].height = h

    # ══════════════════════════════════════════
    # ROWS 1-6: Header block
    # ══════════════════════════════════════════
    # A1:C3  — Logo / company placeholder (merged)
    ws.merge_cells('A1:C3')
    _cell(ws, 1, 1, 'GAUTAM SOLAR', FONT_BOLD, alignment=ALIGN_CENTER)

    # D1:K2 — Title
    ws.merge_cells('D1:K2')
    _cell(ws, 1, 4, 'Gautam Solar Private Limited', FONT_TITLE, alignment=ALIGN_CENTER)

    # D3:K3 — Check Sheet Title
    ws.merge_cells('D3:K3')
    _cell(ws, 3, 4, 'IPQC Check Sheet', FONT_TITLE, alignment=ALIGN_CENTER)

    # Document info (right side)
    ws.merge_cells('L1:M1'); _cell(ws, 1, 12, 'Document No.', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('N1:O1'); _cell(ws, 1, 14, 'GSPL/IPQC/IPC/003', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('L2:M2'); _cell(ws, 2, 12, 'Issue Date', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('N2:O2'); _cell(ws, 2, 14, '01/12/2024', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('L3:M3'); _cell(ws, 3, 12, 'Rev. No./Rev.Date ', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('N3:O3'); _cell(ws, 3, 14, '01/30-08-2025', FONT_BOLD, alignment=ALIGN_CENTER)

    # Row 4 — Date / Time / Shift / PO
    ws.merge_cells('A4:C4')
    _cell(ws, 4, 1, f'Date :-  {date}', FONT_BOLD, alignment=Alignment(horizontal='left', vertical='center'))
    _cell(ws, 4, 4, f' Time :-  {_shift_time(shift)}', FONT_BOLD)
    _cell(ws, 4, 6, f'Shift  {shift}', FONT_BOLD)
    ws.merge_cells('H4:O4')
    _cell(ws, 4, 8, f'Po.no.:- {po_number}', FONT_BOLD, alignment=ALIGN_LEFT_CENTER)

    # Row 5-6 — Column headers
    ws.merge_cells('A5:A6'); _cell(ws, 5, 1, 'Sr.No.', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    ws.merge_cells('B5:B6'); _cell(ws, 5, 2, 'Stage', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    ws.merge_cells('C5:C6'); _cell(ws, 5, 3, 'Check point', FONT_BOLD, HEADER_FILL, ALIGN_CENTER_V)
    ws.merge_cells('D5:E5')
    _cell(ws, 5, 4, 'Quantum of Check', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    _cell(ws, 6, 4, 'Sample Size', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    _cell(ws, 6, 5, 'Frequency', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    ws.merge_cells('F5:G6'); _cell(ws, 5, 6, 'Acceptance Criteria', FONT_BOLD, HEADER_FILL, ALIGN_CENTER_V)
    ws.merge_cells('H5:N6'); _cell(ws, 5, 8, 'Monitoring Result', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)
    ws.merge_cells('O5:O6'); _cell(ws, 5, 15, 'Remarks,if any', FONT_BOLD, HEADER_FILL, ALIGN_CENTER)

    # ══════════════════════════════════════════
    # ROWS 7-139:  All 33 checkpoints
    # ══════════════════════════════════════════
    _write_all_stages(ws, serial_prefix, serial_start, cell_manufacturer, cell_efficiency,
                      jb_cable_length, golden_module_number, shift, date)

    # Row 140 — Checked By / Reviewed By
    ws.merge_cells('A140:B140'); _cell(ws, 140, 1, 'Checked By', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('C140:E140'); _cell(ws, 140, 3, checked_by, FONT_NORMAL, alignment=ALIGN_CENTER)
    ws.merge_cells('F140:K140'); _cell(ws, 140, 6, '', FONT_NORMAL)
    ws.merge_cells('L140:M140'); _cell(ws, 140, 12, 'Reviewed By', FONT_BOLD, alignment=ALIGN_CENTER)
    ws.merge_cells('N140:O140'); _cell(ws, 140, 14, reviewed_by, FONT_NORMAL, alignment=ALIGN_CENTER)

    # ── Apply borders to all cells ──
    for row in ws.iter_rows(min_row=1, max_row=140, min_col=1, max_col=15):
        for cell in row:
            cell.border = THIN_BORDER

    # ── Save ──
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'generated_pdfs')
    os.makedirs(output_dir, exist_ok=True)
    safe_date = date.replace('-', '') if date else datetime.now().strftime('%Y%m%d')
    filename = f"IPQC_CheckSheet_{safe_date}_Shift{shift}_{datetime.now().strftime('%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ──────────────────────────────────────────────
# Shift time helper
# ──────────────────────────────────────────────
def _shift_time(shift):
    mapping = {'A': '06:00 - 14:00', 'B': '14:00 - 22:00', 'C': '22:00 - 06:00'}
    return mapping.get(shift, '06:00 - 14:00')


# ══════════════════════════════════════════════
# All 33 Stages  (rows 7 – 139)
# ══════════════════════════════════════════════
def _write_all_stages(ws, prefix, start, cell_mfr, cell_eff, cable_len, golden, shift, date):
    """Write every stage exactly matching the reference template."""

    # Convenience aliases
    B = FONT_BOLD
    N = FONT_NORMAL
    AC = ALIGN_CENTER
    AV = ALIGN_CENTER_V
    # Monitoring result alignment
    AM = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def sno(row_start, row_end=None):
        """Generate sorted serial numbers list for multi-row serial areas"""
        nums = _gen_serials(prefix, start, 5)
        return nums

    # ─── Stage 1: Shop Floor (rows 7-8) ───
    ws.merge_cells('A7:A8')
    _cell(ws, 7, 1, 1, B, alignment=AC)
    ws.merge_cells('B7:B8')
    # Only set row 7 — row 8 is part of the merge
    # But for B column, we split B7=Shop Floor (merged with B8)
    # Actually B7:B8 is NOT merged in reference — let me check
    # From reference: B7='Shop Floor', B8 has no value  — but they aren't merged per se
    # Actually in the merged cells list: there's no B7:B8, just individual cells
    # Let me re-check: B9:B10, B11:B13  etc. are merged, but B7 has value and B8 doesn't
    # So B7='Shop Floor' spans conceptually but isn't merged in all cases
    # Let me just write each cell individually following the reference exactly
    
    # Actually wait — let me re-examine the merged cells from the reference output...
    # B7:B8 is listed as merged. Let me do it properly.
    _cell(ws, 7, 2, 'Shop Floor', B, alignment=AC)
    _cell(ws, 7, 3, 'Temperature', B, alignment=AV)
    _cell(ws, 7, 4, 'once', N, alignment=AC)
    _cell(ws, 7, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F7:G7')
    _cell(ws, 7, 6, 'Temp. 25±3°C', B, alignment=AV)
    # Monitoring result
    temp_val = _rr(25, 2)
    ws.merge_cells('H7:N7')
    _cell(ws, 7, 8, f'Time: {_shift_start(shift)}   Temp: {temp_val}°C', N, alignment=AM)
    _cell(ws, 7, 15, 'OK' if 22 <= temp_val <= 28 else 'High', N, alignment=AC)

    # Row 8: Humidity
    _cell(ws, 8, 3, 'Humidity', B, alignment=AV)
    _cell(ws, 8, 4, 'once', N, alignment=AC)
    _cell(ws, 8, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F8:G8')
    _cell(ws, 8, 6, 'RH ≤60%', B, alignment=AV)
    rh_val = random.randint(40, 58)
    ws.merge_cells('H8:N8')
    _cell(ws, 8, 8, f'Time: {_shift_start(shift)}   RH: {rh_val}%', N, alignment=AM)
    _cell(ws, 8, 15, 'OK', N, alignment=AC)

    # ─── Stage 2: Glass Loader (rows 9-10) ───
    ws.merge_cells('A9:A10')
    _cell(ws, 9, 1, 2, B, alignment=AC)
    ws.merge_cells('B9:B10')
    _cell(ws, 9, 2, 'Glass Loader', B, alignment=AC)
    _cell(ws, 9, 3, 'Glass dimension(L*W*T)', B, alignment=AV)
    _cell(ws, 9, 4, 'once', N, alignment=AC)
    _cell(ws, 9, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F9:G9')
    _cell(ws, 9, 6, 'As Per PO', B, alignment=AV)
    gl = _rr(2376, 0.8); gw = _rr(1128, 0.8); gt = _rr(2.0, 0.04)
    ws.merge_cells('H9:N9')
    _cell(ws, 9, 8, f'{gl}mm x {gw}mm x {gt}mm', N, alignment=AM)
    _cell(ws, 9, 15, 'OK', N, alignment=AC)

    _cell(ws, 10, 3, 'Appearance(Visual)', B, alignment=AV)
    _cell(ws, 10, 4, 'once', N, alignment=AC)
    _cell(ws, 10, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F10:G10')
    _cell(ws, 10, 6, 'Glass Broken, Crack, Scratches and Line mark not allowed', B, alignment=AV)
    ws.merge_cells('H10:N10')
    _cell(ws, 10, 8, 'No Scratches/Cracks', N, alignment=AM)
    _cell(ws, 10, 15, 'OK', N, alignment=AC)

    # ─── Stage 3: EVA/EPE Cutting (rows 11-13) ───
    ws.merge_cells('A11:A13')
    _cell(ws, 11, 1, 3, B, alignment=AC)
    ws.merge_cells('B11:B13')
    _cell(ws, 11, 2, 'EVA/EPE Cutting', B, alignment=AC)

    _cell(ws, 11, 3, 'EVA/EPE Type', B, alignment=AV)
    _cell(ws, 11, 4, 'once', N, alignment=AC)
    _cell(ws, 11, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F11:G11')
    _cell(ws, 11, 6, 'As per approved BOM', B, alignment=AV)
    ws.merge_cells('H11:N11')
    _cell(ws, 11, 8, 'EPE', N, alignment=AM)
    _cell(ws, 11, 15, 'OK', N, alignment=AC)

    _cell(ws, 12, 3, 'EVA/EPE dimension(L*W*T)', B, alignment=AV)
    _cell(ws, 12, 4, 'once', N, alignment=AC)
    _cell(ws, 12, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F12:G12')
    _cell(ws, 12, 6, 'As per Specification', B, alignment=AV)
    el = _rr(2378, 0.8); ew = _rr(1125, 0.8); et = _rr(0.70, 0.03)
    ws.merge_cells('H12:N12')
    _cell(ws, 12, 8, f'{el}mm x {ew}mm x {et}mm', N, alignment=AM)
    _cell(ws, 12, 15, 'OK', N, alignment=AC)

    _cell(ws, 13, 3, 'EVA/EPE Status', B, alignment=AV)
    _cell(ws, 13, 4, 'once', N, alignment=AC)
    _cell(ws, 13, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F13:G13')
    _cell(ws, 13, 6, 'Not allowed dust & foreign particle/Cut & non Uniform Embossing /Mfg Date', B, alignment=AV)
    ws.merge_cells('H13:N13')
    _cell(ws, 13, 8, 'No Damage, Uniform Embossing', N, alignment=AM)
    _cell(ws, 13, 15, 'OK', N, alignment=AC)

    # ─── Stage 4:  EVA/EPE Soldering (row 14) ───
    _cell(ws, 14, 1, 4, B, alignment=AC)
    _cell(ws, 14, 2, 'Eva/EPE Soldering at edge(If Applicable)', B, alignment=AC)
    _cell(ws, 14, 3, 'Soldering Temprature and Quality of Soldering', B, alignment=AV)
    _cell(ws, 14, 4, 'Once', N, alignment=AC)
    _cell(ws, 14, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F14:G14')
    _cell(ws, 14, 6, 'As per specification and Should be properly soldered ( 400 ± 20°C)', B, alignment=AV)
    solder_t = _rr(400, 15)
    ws.merge_cells('H14:N14')
    _cell(ws, 14, 8, f'{solder_t}°C — Properly Soldered', N, alignment=AM)
    _cell(ws, 14, 15, 'OK', N, alignment=AC)

    # ─── Stage 5: Cell Loading (rows 15-20) ───
    ws.merge_cells('A15:A20')
    _cell(ws, 15, 1, 5, B, alignment=AC)
    ws.merge_cells('B15:B20')
    _cell(ws, 15, 2, 'Cell Loading', B, alignment=AC)

    _cell(ws, 15, 3, 'Cell Manufacturer & Eff.', B, alignment=AV)
    _cell(ws, 15, 4, 'once', N, alignment=AC); _cell(ws, 15, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F15:G15'); _cell(ws, 15, 6, 'Refer Process Card', B, alignment=AV)
    ws.merge_cells('H15:N15'); _cell(ws, 15, 8, f'{cell_mfr}, Eff: {cell_eff}%', N, alignment=AM)
    _cell(ws, 15, 15, 'OK', N, alignment=AC)

    _cell(ws, 16, 3, 'Cell Size(L*W)', B, alignment=AV)
    _cell(ws, 16, 4, 'once', N, alignment=AC); _cell(ws, 16, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F16:G16'); _cell(ws, 16, 6, 'Refer Process Card', B, alignment=AV)
    cl = _rr(182.53, 0.15); cw = _rr(105.04, 0.15)
    ws.merge_cells('H16:N16'); _cell(ws, 16, 8, f'{cl}mm x {cw}mm', N, alignment=AM)
    _cell(ws, 16, 15, 'OK', N, alignment=AC)

    _cell(ws, 17, 3, 'Cell Condition', B, alignment=AV)
    _cell(ws, 17, 4, 'once', N, alignment=AC); _cell(ws, 17, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F17:G17'); _cell(ws, 17, 6, 'Free From dust,finger spot,color variation', B, alignment=AV)
    ws.merge_cells('H17:N17'); _cell(ws, 17, 8, 'No Dust/Finger Spot/Color Variation', N, alignment=AM)
    _cell(ws, 17, 15, 'OK', N, alignment=AC)

    _cell(ws, 18, 3, 'Cleanliness of Cell Loading Area', B, alignment=AV)
    _cell(ws, 18, 4, 'once', N, alignment=AC); _cell(ws, 18, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F18:G18'); _cell(ws, 18, 6, 'No unwanted or waste material should be at Cell Loading Area', B, alignment=AV)
    ws.merge_cells('H18:N18'); _cell(ws, 18, 8, 'Clean — No waste material', N, alignment=AM)
    _cell(ws, 18, 15, 'OK', N, alignment=AC)

    _cell(ws, 19, 3, 'Verification of Process Parameter', B, alignment=AV)
    _cell(ws, 19, 4, 'once', N, alignment=AC); _cell(ws, 19, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F19:G19'); _cell(ws, 19, 6, 'ATW Stringer Specification', B, alignment=AV)
    ws.merge_cells('H19:N19'); _cell(ws, 19, 8, 'Monitoring of ATW STRINGER — Verified', N, alignment=AM)
    _cell(ws, 19, 15, 'OK', N, alignment=AC)

    _cell(ws, 20, 3, 'Cell Cross cutting', B, alignment=AV)
    _cell(ws, 20, 4, 'once', N, alignment=AC); _cell(ws, 20, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F20:G20'); _cell(ws, 20, 6, 'Both side cutting should be equal.', B, alignment=AV)
    ws.merge_cells('H20:N20')
    crosscut = _rr(0, 0.08)
    _cell(ws, 20, 8, f'Difference: {crosscut}mm — equal', N, alignment=AM)
    _cell(ws, 20, 15, 'OK', N, alignment=AC)

    # ─── Stage 6: Tabber & Stringer (rows 21-31) ───
    ws.merge_cells('A21:A29')
    _cell(ws, 21, 1, 6, B, alignment=AC)
    ws.merge_cells('B21:B29')
    _cell(ws, 21, 2, 'Tabber & stringer', B, alignment=AC)

    # 21 — ATW Temp Validation
    _cell(ws, 21, 3, 'Verification of Process Parameter', B, alignment=AV)
    _cell(ws, 21, 4, 'once', N, alignment=AC); _cell(ws, 21, 5, 'Month', N, alignment=AC)
    ws.merge_cells('F21:G21'); _cell(ws, 21, 6, 'ATW Stringer specification', B, alignment=AV)
    ws.merge_cells('H21:N21'); _cell(ws, 21, 8, 'ATW Temprature Validation — Verified', N, alignment=AM)
    _cell(ws, 21, 15, 'OK', N, alignment=AC)

    # 22-23 — Visual Check after Stringing (TS columns)
    ws.merge_cells('C22:C23')
    _cell(ws, 22, 3, 'Visual Check after Stringing', B, alignment=AV)
    ws.merge_cells('D22:D23'); _cell(ws, 22, 4, 'once', N, alignment=AC)
    ws.merge_cells('E22:E23'); _cell(ws, 22, 5, '1 String/TS shift', N, alignment=AC)
    ws.merge_cells('F22:G23'); _cell(ws, 22, 6, 'TS Visual Criteria', B, alignment=AV)
    _cell(ws, 22, 15, 'OK', N, alignment=AC)
    # TS columns: H=TS01A, I=TS01B, J=TS02A, K=TS02B, L=TS03A, M=TS03B, N=TS04A, O=TS04B
    ts_headers_22 = ['TS01A', 'TS01B', 'TS02A', 'TS02B', 'TS03A', 'TS03B', 'TS04A', 'TS04B']
    for i, h in enumerate(ts_headers_22):
        _cell(ws, 22, 8+i, h, B, alignment=AC)
    for i in range(8):
        _cell(ws, 23, 8+i, 'OK', N, alignment=AC)

    # 24-25 — EL Image of Strings
    ws.merge_cells('C24:C25')
    _cell(ws, 24, 3, 'EL Image of Strings', B, alignment=AV)
    ws.merge_cells('D24:D25'); _cell(ws, 24, 4, 'once', N, alignment=AC)
    ws.merge_cells('E24:E25'); _cell(ws, 24, 5, '1 String/TS/shift', N, alignment=AC)
    ws.merge_cells('F24:G25'); _cell(ws, 24, 6, 'TS EL Criteria', B, alignment=AV)
    _cell(ws, 24, 15, 'OK', N, alignment=AC)
    for i, h in enumerate(ts_headers_22):
        _cell(ws, 24, 8+i, h, B, alignment=AC)
    for i in range(8):
        _cell(ws, 25, 8+i, 'OK', N, alignment=AC)

    # 26-27 — String length
    ws.merge_cells('C26:C27')
    _cell(ws, 26, 3, 'String length ', B, alignment=AV)
    ws.merge_cells('D26:D29'); _cell(ws, 26, 4, 'once', N, alignment=AC)
    ws.merge_cells('E26:E29'); _cell(ws, 26, 5, '1 String/Stringer/ shift', N, alignment=AC)
    ws.merge_cells('F26:G29'); _cell(ws, 26, 6, 'Refer Process Card ', B, alignment=AV)
    for i, h in enumerate(ts_headers_22):
        _cell(ws, 26, 8+i, h, B, alignment=AC)
    for i in range(8):
        sl = _rr(1163, 0.8)
        _cell(ws, 27, 8+i, f'{sl:.1f}', N, alignment=AC)

    # 28-29 — Cell to Cell Gap
    ws.merge_cells('C28:C29')
    _cell(ws, 28, 3, ' Cell to Cell Gap', B, alignment=AV)
    for i, h in enumerate(ts_headers_22):
        _cell(ws, 28, 8+i, h, B, alignment=AC)
    for i in range(8):
        gap = _rr(0.77, 0.05)
        _cell(ws, 29, 8+i, f'{gap:.2f}', N, alignment=AC)

    # 30-31 — Peel Strength (Row 30-31, merged A30:A31)
    ws.merge_cells('A30:A31')
    ws.merge_cells('B30:B31')
    ws.merge_cells('C30:C31')
    _cell(ws, 30, 3, 'Verification of Soldering Peel Strength', B, alignment=AV)
    ws.merge_cells('D30:D31'); _cell(ws, 30, 4, '2 cell each stringer Front & Back.', N, alignment=AC)
    ws.merge_cells('E30:E31'); _cell(ws, 30, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F30:G31'); _cell(ws, 30, 6, 'Peel Strength ≥1N', B, alignment=AV)
    ws.merge_cells('H30:N31')
    peel_vals = [_rr(1.5, 0.3) for _ in range(4)]
    _cell(ws, 30, 8, f'Ribbon to cell soldering peel strength: {", ".join(f"{v:.2f}N" for v in peel_vals)}', N, alignment=AM)
    ws.merge_cells('O30:O31')
    _cell(ws, 30, 15, 'OK', N, alignment=AC)

    # ─── Stage 7: Auto bussing, layup & Tapping (rows 32-42) ───
    ws.merge_cells('A32:A42')
    _cell(ws, 32, 1, 7, B, alignment=AC)
    ws.merge_cells('B32:B42')
    _cell(ws, 32, 2, 'Auto bussing , layup & Tapping ', B, alignment=AC)

    # 32 — String to String Gap
    _cell(ws, 32, 3, 'String to String Gap', B, alignment=AV)
    _cell(ws, 32, 4, 'once', N, alignment=AC); _cell(ws, 32, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F32:G35')  # Reference merges F32:G35
    _cell(ws, 32, 6, 'Refer Process Card & Module Drawing', B, alignment=AV)
    ws.merge_cells('H32:N32')
    ss_gap = _rr(2.5, 0.5)
    _cell(ws, 32, 8, f'{ss_gap:.2f}mm', N, alignment=AM)
    _cell(ws, 32, 15, 'OK', N, alignment=AC)

    # 33-35 — Cell edge to Glass edge
    ws.merge_cells('C33:C35')
    _cell(ws, 33, 3, 'Cell edge to Glass edge distance (Top,bottom & sides)', B, alignment=AV)
    ws.merge_cells('D33:D35'); _cell(ws, 33, 4, 'once', N, alignment=AC)
    ws.merge_cells('E33:E35'); _cell(ws, 33, 5, 'per shift', N, alignment=AC)
    # No acceptance criteria merged in ref, just individual rows
    ws.merge_cells('H33:N33')
    top_dist = _rr(19.72, 0.3)
    _cell(ws, 33, 8, f'TOP: {top_dist}mm', N, alignment=AM)
    ws.merge_cells('H34:N34')
    bot_dist = _rr(18.82, 0.3)
    _cell(ws, 34, 8, f'Bottom: {bot_dist}mm', N, alignment=AM)
    ws.merge_cells('H35:N35')
    side_dist = _rr(13.21, 0.2)
    _cell(ws, 35, 8, f'Sides: {side_dist}mm', N, alignment=AM)
    _cell(ws, 33, 15, 'OK', N, alignment=AC)

    # 36 — Soldering Peel Strength busbar
    _cell(ws, 36, 3, 'Soldering Peel Strength b/w Ribbon to busbar interconnector', B, alignment=AV)
    _cell(ws, 36, 4, 'once', N, alignment=AC); _cell(ws, 36, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F36:G36'); _cell(ws, 36, 6, '≥2N', B, alignment=AV)
    ws.merge_cells('H36:N36')
    peel_bb = _rr(3.2, 0.8)
    _cell(ws, 36, 8, f'Ribbon to busbar peel test: {peel_bb:.2f}N', N, alignment=AM)
    _cell(ws, 36, 15, 'OK', N, alignment=AC)

    # 37-38 — Terminal busbar to edge of Cell
    ws.merge_cells('D37:D38'); _cell(ws, 37, 4, 'once', N, alignment=AC)
    ws.merge_cells('E37:E38'); _cell(ws, 37, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('C37:C38')
    _cell(ws, 37, 3, 'Terminal busbar to edge of Cell', B, alignment=AV)
    ws.merge_cells('F37:G37'); _cell(ws, 37, 6, '132 Cell module drawing', B, alignment=AV)
    ws.merge_cells('F38:G38'); _cell(ws, 38, 6, 'Refer Module Drawing: GSPL/N144/G/001', B, alignment=AV)
    ws.merge_cells('H37:N37')
    tb_edge = _rr(6.0, 0.8)
    _cell(ws, 37, 8, f'{tb_edge:.2f}mm', N, alignment=AM)
    ws.merge_cells('H38:N38')
    _cell(ws, 38, 8, 'As per Drawing', N, alignment=AM)
    _cell(ws, 37, 15, 'OK', N, alignment=AC)

    # 39 — Soldering Quality Ribbon to busbar
    _cell(ws, 39, 3, 'Soldering Quality of Ribbon to busbar', B, alignment=AV)
    _cell(ws, 39, 4, 'Every 4h', N, alignment=AC); _cell(ws, 39, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F39:G39'); _cell(ws, 39, 6, 'No Dry/Poor Soldering', B, alignment=AV)
    ws.merge_cells('H39:I39'); _cell(ws, 39, 8, 'OK', N, alignment=AC)
    ws.merge_cells('J39:L39'); _cell(ws, 39, 10, 'No Dry Solder', N, alignment=AC)
    ws.merge_cells('M39:N39'); _cell(ws, 39, 13, 'OK', N, alignment=AC)
    _cell(ws, 39, 15, 'OK', N, alignment=AC)

    # 40 — Top & Bottom Creepage Distance
    _cell(ws, 40, 3, 'Top & Bottom Creepage Distance/Terminal busbar to Glass Edge.', B, alignment=AV)
    _cell(ws, 40, 4, 'Every 4h', N, alignment=AC); _cell(ws, 40, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F40:G40'); _cell(ws, 40, 6, 'Creepage distance should be as per process card/Drawing', B, alignment=AV)
    ws.merge_cells('H40:I40')
    cr_top = _rr(11.7, 0.2)
    _cell(ws, 40, 8, f'Top: {cr_top}mm', N, alignment=AM)
    ws.merge_cells('J40:L40')
    cr_bot = _rr(11.6, 0.2)
    _cell(ws, 40, 10, f'Bottom: {cr_bot}mm', N, alignment=AM)
    ws.merge_cells('M40:N40'); _cell(ws, 40, 13, 'OK', N, alignment=AC)
    _cell(ws, 40, 15, 'OK', N, alignment=AC)

    # 41 — Verification of Process Parameter
    _cell(ws, 41, 3, 'Verification of Process Parameter', B, alignment=AV)
    _cell(ws, 41, 4, 'once', N, alignment=AC); _cell(ws, 41, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F41:G41'); _cell(ws, 41, 6, 'Specification for Auto Bussing', B, alignment=AV)
    ws.merge_cells('H41:N41'); _cell(ws, 41, 8, 'Verified — As per Specification', N, alignment=AM)
    _cell(ws, 41, 15, 'OK', N, alignment=AC)

    # 42 — Quality of auto taping
    _cell(ws, 42, 3, 'Quality of auto taping', B, alignment=AV)
    _cell(ws, 42, 4, 'Every 4h', N, alignment=AC); _cell(ws, 42, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F42:G42'); _cell(ws, 42, 6, 'Taping should be proper,no Cell Shifting allowed', B, alignment=AV)
    ws.merge_cells('H42:I42'); _cell(ws, 42, 8, 'Proper', N, alignment=AC)
    ws.merge_cells('J42:L42'); _cell(ws, 42, 10, 'No Cell Shift', N, alignment=AC)
    ws.merge_cells('M42:N42'); _cell(ws, 42, 13, 'OK', N, alignment=AC)
    _cell(ws, 42, 15, 'OK', N, alignment=AC)

    # ─── Stage 8: Auto RFID / Logo / Barcode (row 43) ───
    _cell(ws, 43, 1, 8, B, alignment=AC)
    _cell(ws, 43, 2, 'Auto RFID  Logo/Barcode placing  (If Applicable)', B, alignment=AC)
    _cell(ws, 43, 3, 'Position verification of RFID& Logo /Barcode placing', B, alignment=AV)
    _cell(ws, 43, 4, 'Every 4h', N, alignment=AC); _cell(ws, 43, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F43:G43'); _cell(ws, 43, 6, 'Should not be tilt', B, alignment=AV)
    ws.merge_cells('H43:I43'); _cell(ws, 43, 8, 'Center', N, alignment=AC)
    ws.merge_cells('J43:L43'); _cell(ws, 43, 10, 'No Tilt', N, alignment=AC)
    ws.merge_cells('M43:N43'); _cell(ws, 43, 13, 'OK', N, alignment=AC)
    _cell(ws, 43, 15, 'OK', N, alignment=AC)

    # ─── Stage 9: EVA/EPE cutting (rows 44-46) ───
    ws.merge_cells('A44:A46')
    _cell(ws, 44, 1, 9, B, alignment=AC)
    ws.merge_cells('B44:B46')
    _cell(ws, 44, 2, 'EVA/EPE cutting', B, alignment=AC)

    _cell(ws, 44, 3, 'EVA/EPE Type', B, alignment=AV)
    _cell(ws, 44, 4, 'once', N, alignment=AC); _cell(ws, 44, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F44:G44'); _cell(ws, 44, 6, 'EVA', B, alignment=AV)
    ws.merge_cells('H44:N44'); _cell(ws, 44, 8, 'EVA', N, alignment=AM)
    _cell(ws, 44, 15, 'OK', N, alignment=AC)

    _cell(ws, 45, 3, 'EVA/EPE dimension(L*W*T)', B, alignment=AV)
    _cell(ws, 45, 4, 'once', N, alignment=AC); _cell(ws, 45, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F45:G45'); _cell(ws, 45, 6, 'As per Specification', B, alignment=AV)
    ws.merge_cells('H45:N45')
    eva2l = _rr(2378, 0.8); eva2w = _rr(1125, 0.8)
    _cell(ws, 45, 8, f'{eva2l}mm x {eva2w}mm', N, alignment=AM)
    _cell(ws, 45, 15, 'OK', N, alignment=AC)

    _cell(ws, 46, 3, 'EVA/EPE Status', B, alignment=AV)
    _cell(ws, 46, 4, 'once', N, alignment=AC); _cell(ws, 46, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F46:G46'); _cell(ws, 46, 6, 'Not allowed dust & foreign particle/Cut & non Uniform Embossing /Mfg Date', B, alignment=AV)
    ws.merge_cells('H46:N46'); _cell(ws, 46, 8, 'Clean Surface, Uniform', N, alignment=AM)
    _cell(ws, 46, 15, 'OK', N, alignment=AC)

    # ─── Stage 10: Back Glass Loader (rows 47-50) ───
    ws.merge_cells('A47:A48')
    _cell(ws, 47, 1, 10, B, alignment=AC)
    ws.merge_cells('B47:B48')
    _cell(ws, 47, 2, 'Back Glass Loader', B, alignment=AC)
    ws.merge_cells('C47:C48')
    _cell(ws, 47, 3, 'Glass dimension(L*W*T)', B, alignment=AV)
    ws.merge_cells('D47:D48'); _cell(ws, 47, 4, 'once', N, alignment=AC)
    ws.merge_cells('E47:E48'); _cell(ws, 47, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F47:G48'); _cell(ws, 47, 6, 'As per PO', B, alignment=AV)
    ws.merge_cells('H47:N48')
    bg_l = _rr(2376, 0.8); bg_w = _rr(1128, 0.8); bg_t = _rr(2.0, 0.04)
    _cell(ws, 47, 8, f'{bg_l}mm x {bg_w}mm x {bg_t}mm', N, alignment=AM)
    ws.merge_cells('O47:O48'); _cell(ws, 47, 15, 'OK', N, alignment=AC)

    # 49-50: Holes
    ws.merge_cells('A49:A50')
    _cell(ws, 49, 1, '', B, alignment=AC)
    ws.merge_cells('B49:B50')
    ws.merge_cells('C49:C50')
    _cell(ws, 49, 3, 'No. of Holes/ Holes dimension', B, alignment=AV)
    ws.merge_cells('D49:D50'); _cell(ws, 49, 4, 'once', N, alignment=AC)
    ws.merge_cells('E49:E50'); _cell(ws, 49, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F49:G50'); _cell(ws, 49, 6, '3 hole with dimension 12mm±0.5mm', B, alignment=AV)
    ws.merge_cells('H49:N50')
    h1 = _rr(12, 0.3); h2 = _rr(12, 0.3); h3 = _rr(12, 0.3)
    _cell(ws, 49, 8, f'3 holes: {h1:.2f}mm, {h2:.2f}mm, {h3:.2f}mm', N, alignment=AM)
    ws.merge_cells('O49:O50'); _cell(ws, 49, 15, 'OK', N, alignment=AC)

    # ─── Stage 11: Auto Busbar Flatten (rows 51-52) ───
    ws.merge_cells('A51:A52')
    _cell(ws, 51, 1, 11, B, alignment=AC)
    ws.merge_cells('B51:B52')
    _cell(ws, 51, 2, 'Auto Busbar Flatten  (If Applicable)', B, alignment=AC)
    ws.merge_cells('C51:C52')
    _cell(ws, 51, 3, 'Visual Inspection', B, alignment=AV)
    ws.merge_cells('D51:D52'); _cell(ws, 51, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E51:E52'); _cell(ws, 51, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F51:G52'); _cell(ws, 51, 6, 'No  cracks/ breaks in busbar &properly flattened without bending and twisting', B, alignment=AV)
    ws.merge_cells('H51:I52'); _cell(ws, 51, 8, 'S.No', B, alignment=AC)
    ws.merge_cells('J51:J52')
    ws.merge_cells('K51:L52')
    ws.merge_cells('M51:M52')
    ws.merge_cells('N51:N52')
    snos = _gen_serials(prefix, start, 5)
    _cell(ws, 51, 10, snos[0] if len(snos)>0 else '', N, alignment=AC)
    _cell(ws, 51, 11, snos[1] if len(snos)>1 else '', N, alignment=AC)
    _cell(ws, 51, 13, snos[2] if len(snos)>2 else '', N, alignment=AC)
    _cell(ws, 51, 14, snos[3] if len(snos)>3 else '', N, alignment=AC)
    ws.merge_cells('O51:O52'); _cell(ws, 51, 15, 'OK', N, alignment=AC)

    # ─── Stage 12: Pre lamination EL & Visual (rows 53-57) ───
    ws.merge_cells('A53:A57')
    _cell(ws, 53, 1, 12, B, alignment=AC)
    ws.merge_cells('B53:B57')
    _cell(ws, 53, 2, 'Pre lamination EL &Visual inspection', B, alignment=AC)
    ws.merge_cells('C53:C57')
    _cell(ws, 53, 3, 'EL Inspection and Visual inspection', B, alignment=AV)
    ws.merge_cells('D53:D57'); _cell(ws, 53, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E53:E57'); _cell(ws, 53, 5, 'per shift', N, alignment=AC)

    ws.merge_cells('F53:G55')
    _cell(ws, 53, 6, 'Pre EL Inspection Criteria', B, alignment=AV)

    # Serial number rows for Pre EL
    snos_pre = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([53, 54, 55]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_pre[idx] if idx < len(snos_pre) else '', N, alignment=AM)

    ws.merge_cells('F56:G57')
    _cell(ws, 56, 6, 'Pre EL Visual Criteria', B, alignment=AV)
    for idx, r in enumerate([56, 57]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_pre[3+idx] if 3+idx < len(snos_pre) else '', N, alignment=AM)
    _cell(ws, 53, 15, 'OK', N, alignment=AC)

    # ─── Stage 13: String Rework Station (rows 58-59) ───
    ws.merge_cells('A58:A59')
    _cell(ws, 58, 1, 13, B, alignment=AC)
    ws.merge_cells('B58:B59')
    _cell(ws, 58, 2, 'String Rework Station', B, alignment=AC)

    _cell(ws, 58, 3, 'cleaning of rework station/Soldering iron and  sponge', B, alignment=AV)
    _cell(ws, 58, 4, 'once', N, alignment=AC); _cell(ws, 58, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F58:G58'); _cell(ws, 58, 6, 'Rework Station should be Clean/Sponge should be Wet', B, alignment=AV)
    ws.merge_cells('H58:N58'); _cell(ws, 58, 8, 'Clean — Sponge Wet', N, alignment=AM)
    _cell(ws, 58, 15, 'OK', N, alignment=AC)

    _cell(ws, 59, 3, 'Soldering Iron Temp.', B, alignment=AV)
    _cell(ws, 59, 4, 'once', N, alignment=AC); _cell(ws, 59, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F59:G59'); _cell(ws, 59, 6, '400±30°C', B, alignment=AV)
    iron_t1 = _rr(400, 20)
    ws.merge_cells('H59:N59'); _cell(ws, 59, 8, f'Time: {_shift_start(shift)}  {iron_t1}°C', N, alignment=AM)
    _cell(ws, 59, 15, 'OK', N, alignment=AC)

    # ─── Stage 14: Module Rework Station (rows 60-62) ───
    ws.merge_cells('A60:A62')
    _cell(ws, 60, 1, 14, B, alignment=AC)
    ws.merge_cells('B60:B62')
    _cell(ws, 60, 2, 'Module Rework Station', B, alignment=AC)

    _cell(ws, 60, 3, 'Method of Rework', B, alignment=AV)
    _cell(ws, 60, 4, 'once', N, alignment=AC); _cell(ws, 60, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F60:G60'); _cell(ws, 60, 6, 'As per WI (GSPL/P/WI/012)', B, alignment=AV)
    ws.merge_cells('H60:N60'); _cell(ws, 60, 8, 'As per WI', N, alignment=AM)
    _cell(ws, 60, 15, 'OK', N, alignment=AC)

    _cell(ws, 61, 3, 'Cleaning of Rework station/Soldering iron sponge', B, alignment=AV)
    _cell(ws, 61, 4, 'once', N, alignment=AC); _cell(ws, 61, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F61:G61'); _cell(ws, 61, 6, 'Rework Station should be Clean/Sponge should be Wet', B, alignment=AV)
    ws.merge_cells('H61:N61'); _cell(ws, 61, 8, 'Clean — Sponge Wet', N, alignment=AM)
    _cell(ws, 61, 15, 'OK', N, alignment=AC)

    _cell(ws, 62, 3, 'Soldering Iron Temp.', B, alignment=AV)
    _cell(ws, 62, 4, 'once', N, alignment=AC); _cell(ws, 62, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F62:G62'); _cell(ws, 62, 6, '400±30°C', B, alignment=AV)
    iron_t2 = _rr(400, 20)
    ws.merge_cells('H62:N62'); _cell(ws, 62, 8, f'Time: {_shift_start(shift)}  {iron_t2}°C', N, alignment=AM)
    _cell(ws, 62, 15, 'OK', N, alignment=AC)

    # ─── Stage 15: Laminator (rows 63-66) ───
    ws.merge_cells('A63:A64')
    _cell(ws, 63, 1, 15, B, alignment=AC)
    ws.merge_cells('B63:B64')
    _cell(ws, 63, 2, 'Laminator', B, alignment=AC)

    _cell(ws, 63, 3, 'Monitoring of Laminator Process parameter', B, alignment=AV)
    _cell(ws, 63, 4, 'once', N, alignment=AC); _cell(ws, 63, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F63:G63'); _cell(ws, 63, 6, 'Process Parameter of jinchen Laminator', B, alignment=AV)
    ws.merge_cells('H63:N63'); _cell(ws, 63, 8, 'Verified — As per Specification', N, alignment=AM)
    _cell(ws, 63, 15, 'OK', N, alignment=AC)

    _cell(ws, 64, 3, 'Cleaning of Diaphragm/release sheet', B, alignment=AV)
    _cell(ws, 64, 4, 'once', N, alignment=AC); _cell(ws, 64, 5, '24h', N, alignment=AC)
    ws.merge_cells('F64:G64'); _cell(ws, 64, 6, 'Diaphragm/Release sheet should be clean,No EVA residue is allowed', B, alignment=AV)
    ws.merge_cells('H64:N64'); _cell(ws, 64, 8, 'Clean — No EVA Residue', N, alignment=AM)
    _cell(ws, 64, 15, 'OK', N, alignment=AC)

    # 65-66 — Peel test & Gel content (these are separate rows under stage 15 conceptually)
    ws.merge_cells('A65:A66')
    ws.merge_cells('B65:B66')
    ws.merge_cells('D65:D66'); _cell(ws, 65, 4, 'All position', N, alignment=AC)
    ws.merge_cells('E65:E66'); _cell(ws, 65, 5, 'All laminators to be coverd in a month', N, alignment=AC)

    _cell(ws, 65, 3, 'Peel of Test b/w: \nEVA/Backsheet EVA/EPE/POE to Glass ', B, alignment=AV)
    ws.merge_cells('F65:G65'); _cell(ws, 65, 6, 'E/G ≥60N/cm E/B≥60N/cm', B, alignment=AV)
    ws.merge_cells('H65:N65'); _cell(ws, 65, 8, 'Refer Document GSPL/IPQC/QC/001', N, alignment=AM)
    _cell(ws, 65, 15, 'OK', N, alignment=AC)

    _cell(ws, 66, 3, 'Gel Content Test', B, alignment=AV)
    ws.merge_cells('F66:G66'); _cell(ws, 66, 6, '75to 95%', B, alignment=AV)
    ws.merge_cells('H66:N66'); _cell(ws, 66, 8, 'Refer Document GSPL/IPQC/QC/001', N, alignment=AM)
    _cell(ws, 66, 15, 'OK', N, alignment=AC)

    # ─── Stage 16: Auto Tape Removing (row 67) ───
    _cell(ws, 67, 1, 16, B, alignment=AC)
    _cell(ws, 67, 2, 'Auto Tape Removing  (If Applicable)', B, alignment=AC)
    _cell(ws, 67, 3, 'Visual Check after Lamination', B, alignment=AV)
    _cell(ws, 67, 4, '5 pieces', N, alignment=AC); _cell(ws, 67, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F67:G67')
    _cell(ws, 67, 6, 'Check Tape Removing Should be smooth and No visual bubble Should be found. ', B, alignment=AV)
    ws.merge_cells('H67:I67'); _cell(ws, 67, 8, 'Smooth', N, alignment=AC)
    ws.merge_cells('K67:L67'); _cell(ws, 67, 11, 'No Bubble', N, alignment=AC)
    _cell(ws, 67, 15, 'OK', N, alignment=AC)

    # ─── Stage 17: Auto Edge Trimming (rows 68-73) ───
    ws.merge_cells('A68:A73')
    _cell(ws, 68, 1, 17, B, alignment=AC)
    ws.merge_cells('B68:B73')
    _cell(ws, 68, 2, 'Auto Edge Trimming', B, alignment=AC)

    ws.merge_cells('C68:C72')
    _cell(ws, 68, 3, 'Trimming Quality', B, alignment=AV)
    ws.merge_cells('D68:D72'); _cell(ws, 68, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E68:E72'); _cell(ws, 68, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F68:G72'); _cell(ws, 68, 6, 'Excess layer from the glass edge should be removed,Uneven Trimming not allowed', B, alignment=AV)

    snos_trim = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([68, 69, 70, 71, 72]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No.', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_trim[idx] if idx < len(snos_trim) else '', N, alignment=AM)
    _cell(ws, 68, 15, 'OK', N, alignment=AC)

    _cell(ws, 73, 3, 'Trimming Blade life cycle', B, alignment=AV)
    _cell(ws, 73, 4, 'once', N, alignment=AC); _cell(ws, 73, 5, 'per month', N, alignment=AC)
    ws.merge_cells('F73:G73'); _cell(ws, 73, 6, 'Worn out not allowed', B, alignment=AV)
    ws.merge_cells('H73:O73'); _cell(ws, 73, 8, 'Blade OK — Not worn out', N, alignment=AM)

    # ─── Stage 18: 90° Visual inspection (rows 74-78) ───
    ws.merge_cells('A74:A78')
    _cell(ws, 74, 1, 18, B, alignment=AC)
    ws.merge_cells('B74:B78')
    _cell(ws, 74, 2, '90° Visual inspection', B, alignment=AC)
    ws.merge_cells('C74:C78')
    _cell(ws, 74, 3, 'Visual Inspection', B, alignment=AV)
    ws.merge_cells('D74:D78'); _cell(ws, 74, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E74:E78'); _cell(ws, 74, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F74:G78'); _cell(ws, 74, 6, 'Post Lam Visual Inspection Criteria', B, alignment=AV)

    snos_vis = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([74, 75, 76, 77, 78]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_vis[idx] if idx < len(snos_vis) else '', N, alignment=AM)
    _cell(ws, 74, 15, 'OK', N, alignment=AC)

    # ─── Stage 19: Framing (rows 79-82) ───
    ws.merge_cells('A79:A82')
    _cell(ws, 79, 1, 19, B, alignment=AC)
    ws.merge_cells('B79:B82')
    _cell(ws, 79, 2, 'Framing', B, alignment=AC)

    _cell(ws, 79, 3, 'Glue uniformity & continuity in frame groove', B, alignment=AV)
    _cell(ws, 79, 4, '1 set', N, alignment=AC); _cell(ws, 79, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F79:G79'); _cell(ws, 79, 6, 'Should be uniform,Back sealing should be proper', B, alignment=AV)
    ws.merge_cells('H79:N79'); _cell(ws, 79, 8, 'Uniform — Back sealing proper', N, alignment=AM)
    _cell(ws, 79, 15, 'OK', N, alignment=AC)

    # 80-81 Glue weights
    _cell(ws, 80, 3, 'Short Side Glue Weight', B, alignment=AV)
    _cell(ws, 80, 4, 'once', N, alignment=AC); _cell(ws, 80, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F80:G81'); _cell(ws, 80, 6, 'Fill as per Specification', B, alignment=AV)
    ws.merge_cells('H80:N81'); _cell(ws, 80, 8, 'Refer Document GSPL/IPQC/QC/011', N, alignment=AM)
    ws.merge_cells('O80:O81'); _cell(ws, 80, 15, 'OK', N, alignment=AC)
    _cell(ws, 81, 3, 'Long Side Glue Weight', B, alignment=AV)
    _cell(ws, 81, 4, 'once', N, alignment=AC); _cell(ws, 81, 5, 'Per shift', N, alignment=AC)

    # 82 — Anodizing
    _cell(ws, 82, 3, 'Anodizing Thickness', B, alignment=AV)
    _cell(ws, 82, 4, 'once', N, alignment=AC); _cell(ws, 82, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F82:G82'); _cell(ws, 82, 6, '≥15 micron', B, alignment=AV)
    anod = _rr(16.5, 1.2)
    ws.merge_cells('H82:N82'); _cell(ws, 82, 8, f'{anod:.1f} micron', N, alignment=AM)
    _cell(ws, 82, 15, 'OK', N, alignment=AC)

    # ─── Stage 20: Junction Box Assembly (rows 83-85) ───
    ws.merge_cells('A83:A85')
    _cell(ws, 83, 1, 20, B, alignment=AC)
    ws.merge_cells('B83:B85')
    _cell(ws, 83, 2, 'Junction Box Assembly', B, alignment=AC)
    ws.merge_cells('C83:C84')
    _cell(ws, 83, 3, 'Junction Box(Connector Appereance & Cable Length)', B, alignment=AV)
    ws.merge_cells('D83:D84'); _cell(ws, 83, 4, 'once', N, alignment=AC)
    ws.merge_cells('E83:E84'); _cell(ws, 83, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F83:G84'); _cell(ws, 83, 6, 'As per Process Card & module drawing', B, alignment=AV)
    ws.merge_cells('H83:N84'); _cell(ws, 83, 8, f'Cable Length: {cable_len}mm — OK', N, alignment=AM)
    ws.merge_cells('O83:O84'); _cell(ws, 83, 15, 'OK', N, alignment=AC)

    _cell(ws, 85, 3, 'Silicon Glue Weight on the bottom (g)', B, alignment=AV)
    _cell(ws, 85, 4, 'once', N, alignment=AC); _cell(ws, 85, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F85:G85'); _cell(ws, 85, 6, '21±6 gm', B, alignment=AV)
    glue_w = _rr(21, 4)
    ws.merge_cells('H85:N85'); _cell(ws, 85, 8, f'{glue_w:.1f} gm', N, alignment=AM)
    _cell(ws, 85, 15, 'OK', N, alignment=AC)

    # ─── Stage 21: Auto JB Soldering (rows 86-88) ───
    ws.merge_cells('A86:A88')
    _cell(ws, 86, 1, 21, B, alignment=AC)
    ws.merge_cells('B86:B88')
    _cell(ws, 86, 2, 'Auto JB Soldering', B, alignment=AC)

    _cell(ws, 86, 3, 'Max Welding time', B, alignment=AV)
    _cell(ws, 86, 4, 'once', N, alignment=AC); _cell(ws, 86, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F86:G86'); _cell(ws, 86, 6, 'As per Specification', B, alignment=AV)
    ws.merge_cells('H86:N86'); _cell(ws, 86, 8, 'Within Specification', N, alignment=AM)
    _cell(ws, 86, 15, 'OK', N, alignment=AC)

    _cell(ws, 87, 3, 'Soldering current', B, alignment=AV)
    _cell(ws, 87, 4, 'once', N, alignment=AC); _cell(ws, 87, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F87:G87'); _cell(ws, 87, 6, 'As per Specification', B, alignment=AV)
    s_current = _rr(20, 1.5)
    ws.merge_cells('H87:N87'); _cell(ws, 87, 8, f'{s_current:.1f}A', N, alignment=AM)
    _cell(ws, 87, 15, 'OK', N, alignment=AC)

    _cell(ws, 88, 3, 'Soldering Quality', B, alignment=AV)
    _cell(ws, 88, 4, 'once', N, alignment=AC); _cell(ws, 88, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F88:G88'); _cell(ws, 88, 6, 'Welding area should be fully covered & checked by twizzer,no yellowing allowed', B, alignment=AV)
    ws.merge_cells('H88:N88'); _cell(ws, 88, 8, 'Fully Covered — No Yellowing', N, alignment=AM)
    _cell(ws, 88, 15, 'OK', N, alignment=AC)

    # ─── Stage 22: JB Potting (rows 89-91) ───
    ws.merge_cells('A89:A91')
    _cell(ws, 89, 1, 22, B, alignment=AC)
    ws.merge_cells('B89:B91')
    _cell(ws, 89, 2, 'JB Potting', B, alignment=AC)

    _cell(ws, 89, 3, 'A/B Glue Ratio', B, alignment=AV)
    _cell(ws, 89, 4, 'once', N, alignment=AC); _cell(ws, 89, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F89:G89'); _cell(ws, 89, 6, 'As per Specification', B, alignment=AV)
    ws.merge_cells('H89:N89'); _cell(ws, 89, 8, 'Refer Document GSPL/IPQC/QC/011', N, alignment=AM)
    _cell(ws, 89, 15, 'OK', N, alignment=AC)

    _cell(ws, 90, 3, 'Potting material weight', B, alignment=AV)
    _cell(ws, 90, 4, 'once', N, alignment=AC); _cell(ws, 90, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F90:G90'); _cell(ws, 90, 6, '21±6 gm', B, alignment=AV)
    pot_w = _rr(21, 4)
    ws.merge_cells('H90:N90'); _cell(ws, 90, 8, f'{pot_w:.1f} gm', N, alignment=AM)
    _cell(ws, 90, 15, 'OK', N, alignment=AC)

    _cell(ws, 91, 3, 'Nozzle Changing', B, alignment=AV)
    _cell(ws, 91, 4, 'once', N, alignment=AC); _cell(ws, 91, 5, 'every 6h', N, alignment=AC)
    ws.merge_cells('F91:G91'); _cell(ws, 91, 6, 'Should be changed after 6 hours or when found issue of damage or extra amount dispensing.', B, alignment=AV)
    ws.merge_cells('H91:K91'); _cell(ws, 91, 8, f'Time: {_shift_start(shift)}', N, alignment=AM)
    ws.merge_cells('L91:N91'); _cell(ws, 91, 12, f'Time: {_shift_mid(shift)}', N, alignment=AM)
    _cell(ws, 91, 15, 'OK', N, alignment=AC)

    # ─── Stage 23: OLE Potting Inspection (row 92) ───
    _cell(ws, 92, 1, 23, B, alignment=AC)
    _cell(ws, 92, 2, 'OLE Potting Inspection  (If Applicable)', B, alignment=AC)
    _cell(ws, 92, 3, 'Visual Check', B, alignment=AV)
    _cell(ws, 92, 4, 'once', N, alignment=AC); _cell(ws, 92, 5, '5 piece', N, alignment=AC)
    ws.merge_cells('F92:G92'); _cell(ws, 92, 6, 'Potting should be properly filled, and mounting hole should be as per drawing.', B, alignment=AV)
    ws.merge_cells('H92:I92'); _cell(ws, 92, 8, 'Visual', B, alignment=AC)
    ws.merge_cells('K92:L92'); _cell(ws, 92, 11, 'OK', N, alignment=AC)
    _cell(ws, 92, 15, 'OK', N, alignment=AC)

    # ─── Stage 24: Curing (rows 93-95) ───
    ws.merge_cells('A93:A95')
    _cell(ws, 93, 1, 24, B, alignment=AC)
    ws.merge_cells('B93:B95')
    _cell(ws, 93, 2, 'Curing', B, alignment=AC)

    _cell(ws, 93, 3, 'Temperature', B, alignment=AV)
    _cell(ws, 93, 4, 'once', N, alignment=AC); _cell(ws, 93, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F93:G93'); _cell(ws, 93, 6, '25±3℃', B, alignment=AV)
    cure_t = _rr(25, 2)
    ws.merge_cells('H93:N93'); _cell(ws, 93, 8, f'{cure_t}℃', N, alignment=AM)
    _cell(ws, 93, 15, 'OK', N, alignment=AC)

    _cell(ws, 94, 3, 'Humidity', B, alignment=AV)
    _cell(ws, 94, 4, 'once', N, alignment=AC); _cell(ws, 94, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F94:G94'); _cell(ws, 94, 6, '≥50%', B, alignment=AV)
    cure_h = random.randint(52, 62)
    ws.merge_cells('H94:N94'); _cell(ws, 94, 8, f'{cure_h}%', N, alignment=AM)
    _cell(ws, 94, 15, 'OK', N, alignment=AC)

    _cell(ws, 95, 3, 'Curing Time(h)', B, alignment=AV)
    _cell(ws, 95, 4, 'once', N, alignment=AC); _cell(ws, 95, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F95:G95'); _cell(ws, 95, 6, '≥4 hours', B, alignment=AV)
    cure_time = round(random.uniform(4.5, 6.0), 1)
    ws.merge_cells('H95:N95'); _cell(ws, 95, 8, f'{cure_time} hours', N, alignment=AM)
    _cell(ws, 95, 15, 'OK', N, alignment=AC)

    # ─── Stage 25: Buffing (row 96) ───
    _cell(ws, 96, 1, 25, B, alignment=AC)
    _cell(ws, 96, 2, 'Buffing', B, alignment=AC)
    _cell(ws, 96, 3, 'Corner Edge,Buffing belt condition', B, alignment=AV)
    _cell(ws, 96, 4, '5 pieces', N, alignment=AC); _cell(ws, 96, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F96:G96'); _cell(ws, 96, 6, 'Should not be sharp & No worn out', B, alignment=AV)
    ws.merge_cells('H96:I96'); _cell(ws, 96, 8, 'Not Sharp', N, alignment=AC)
    ws.merge_cells('K96:L96'); _cell(ws, 96, 11, 'No Worn Out', N, alignment=AC)
    _cell(ws, 96, 15, 'OK', N, alignment=AC)

    # ─── Stage 26: Cleaning (rows 97-101) ───
    ws.merge_cells('A97:A101')
    _cell(ws, 97, 1, 26, B, alignment=AC)
    ws.merge_cells('B97:B101')
    _cell(ws, 97, 2, 'Cleaning', B, alignment=AC)
    ws.merge_cells('C97:C101')
    _cell(ws, 97, 3, 'Module should be free from Tape,Dust,Dirt,EVA/Backsheet residue,Corner Burrs,Glue residue on (glass,backsheet,JB,Wire etc.)', B, alignment=AV)
    ws.merge_cells('D97:D101'); _cell(ws, 97, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E97:E101'); _cell(ws, 97, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F97:G101'); _cell(ws, 97, 6, 'Post Lam Visual Criteria', B, alignment=AV)

    snos_clean = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([97, 98, 99, 100, 101]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_clean[idx] if idx < len(snos_clean) else '', N, alignment=AM)
    _cell(ws, 97, 15, 'OK', N, alignment=AC)

    # ─── Stage 27: Flash Tester (rows 102-106) ───
    ws.merge_cells('A102:A106')
    _cell(ws, 102, 1, 27, B, alignment=AC)
    ws.merge_cells('B102:B106')
    _cell(ws, 102, 2, 'Flash Tester', B, alignment=AC)

    _cell(ws, 102, 3, 'Ambient Temp.', B, alignment=AV)
    _cell(ws, 102, 4, 'once', N, alignment=AC); _cell(ws, 102, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F102:G102'); _cell(ws, 102, 6, '25±3℃', B, alignment=AV)
    ft_temp = _rr(25, 2)
    ws.merge_cells('H102:N102'); _cell(ws, 102, 8, f'{ft_temp}℃', N, alignment=AM)
    _cell(ws, 102, 15, 'OK', N, alignment=AC)

    _cell(ws, 103, 3, 'Module Temp.', B, alignment=AV)
    _cell(ws, 103, 4, 'once', N, alignment=AC); _cell(ws, 103, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F103:G103'); _cell(ws, 103, 6, '25±3℃', B, alignment=AV)
    mt = _rr(25, 2)
    ws.merge_cells('H103:N103'); _cell(ws, 103, 8, f'{mt}℃', N, alignment=AM)
    _cell(ws, 103, 15, 'OK', N, alignment=AC)

    _cell(ws, 104, 3, 'Sunsimulator Calibration', B, alignment=AV)
    _cell(ws, 104, 4, 'once', N, alignment=AC); _cell(ws, 104, 5, '12h', N, alignment=AC)
    ws.merge_cells('F104:G104'); _cell(ws, 104, 6, 'Sunsimulator must be calibrated at the start of the shift with Golden/Silver module(GSPL/QA/S/11)', B, alignment=AV)
    ws.merge_cells('H104:N104'); _cell(ws, 104, 8, f'Calibrated with Golden Module: {golden}', N, alignment=AM)
    _cell(ws, 104, 15, 'OK', N, alignment=AC)

    _cell(ws, 105, 3, 'Validation ', B, alignment=AV)
    _cell(ws, 105, 4, 'once', N, alignment=AC); _cell(ws, 105, 5, 'every 6h', N, alignment=AC)
    ws.merge_cells('F105:G105'); _cell(ws, 105, 6, 'As per GSPL/QA/S/11', B, alignment=AV)
    ws.merge_cells('H105:J105'); _cell(ws, 105, 8, 'Validated', N, alignment=AC)
    ws.merge_cells('K105:N105'); _cell(ws, 105, 11, 'As per GSPL/QA/S/11', N, alignment=AM)
    _cell(ws, 105, 15, 'OK', N, alignment=AC)

    _cell(ws, 106, 3, 'Silver Reference Module EL Check', B, alignment=AV)
    _cell(ws, 106, 4, 'once', N, alignment=AC); _cell(ws, 106, 5, 'Two weeks', N, alignment=AC)
    ws.merge_cells('F106:G106'); _cell(ws, 106, 6, 'Should be same as orignal EL picture', B, alignment=AV)
    ws.merge_cells('H106:N106'); _cell(ws, 106, 8, 'Same as original EL picture', N, alignment=AM)
    _cell(ws, 106, 15, 'OK', N, alignment=AC)

    # ─── Stage 28: Hipot Test (rows 107-112) ───
    ws.merge_cells('A107:A112')
    _cell(ws, 107, 1, 28, B, alignment=AC)
    ws.merge_cells('B107:B112')
    _cell(ws, 107, 2, 'Hipot Test', B, alignment=AC)
    ws.merge_cells('C107:C112')
    _cell(ws, 107, 3, 'DCW/IR/Ground Continuity', B, alignment=AV)
    ws.merge_cells('D107:D112'); _cell(ws, 107, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E107:E112'); _cell(ws, 107, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F107:G112'); _cell(ws, 107, 6, '≤50µA , >40M\u2126.m2 ,(0-100) m\u2126', B, alignment=AV)

    # Header row for Hipot
    ws.merge_cells('H107:J107')  # Reference merges H107:J107
    _cell(ws, 107, 8, 'S.No', B, alignment=AC)
    ws.merge_cells('K107:L107'); _cell(ws, 107, 11, 'DCW', B, alignment=AC)
    _cell(ws, 107, 13, 'IR', B, alignment=AC)
    _cell(ws, 107, 14, 'Ground Continuity', B, alignment=AC)

    snos_hipot = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([108, 109, 110, 111, 112]):
        ws.merge_cells(f'H{r}:J{r}')
        _cell(ws, r, 8, snos_hipot[idx] if idx < len(snos_hipot) else '', N, alignment=AM)
        ws.merge_cells(f'K{r}:L{r}')
        dcw = _rr(22, 10)
        _cell(ws, r, 11, f'{dcw:.1f}µA', N, alignment=AC)
        ir_v = _rr(80, 25)
        _cell(ws, r, 13, f'{ir_v:.1f}MΩ', N, alignment=AC)
        gnd = _rr(30, 12)
        _cell(ws, r, 14, f'{gnd:.1f}mΩ', N, alignment=AC)
    _cell(ws, 107, 15, 'OK', N, alignment=AC)

    # ─── Stage 29: Post EL Test (rows 113-118) ───
    ws.merge_cells('A113:A118')
    _cell(ws, 113, 1, 29, B, alignment=AC)
    ws.merge_cells('B113:B118')
    _cell(ws, 113, 2, 'Post EL Test', B, alignment=AC)

    _cell(ws, 113, 3, 'Voltage & Current Verification in DC power supply', B, alignment=AV)
    _cell(ws, 113, 4, 'once', N, alignment=AC); _cell(ws, 113, 5, 'Shift', N, alignment=AC)
    ws.merge_cells('F113:G113'); _cell(ws, 113, 6, 'As per WI (GSPL/P/WI/027)', B, alignment=AV)
    ws.merge_cells('H113:O113')
    voltage = _rr(49, 0.5); current_a = _rr(5.4, 0.2)
    _cell(ws, 113, 8, f'{voltage:.2f}V,  {current_a:.3f}A', N, alignment=AM)

    # EL results
    ws.merge_cells('C114:C118')
    _cell(ws, 114, 3, 'EL Inspection and Visual inspection', B, alignment=AV)
    ws.merge_cells('D114:D118'); _cell(ws, 114, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E114:E118'); _cell(ws, 114, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F114:G116'); _cell(ws, 114, 6, 'Post EL Inspection Criteria', B, alignment=AV)

    snos_el = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([114, 115, 116]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_el[idx] if idx < len(snos_el) else '', N, alignment=AM)

    ws.merge_cells('F117:G118'); _cell(ws, 117, 6, 'Post EL Visual Criteria', B, alignment=AV)
    for idx, r in enumerate([117, 118]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_el[3+idx] if 3+idx < len(snos_el) else '', N, alignment=AM)
    _cell(ws, 114, 15, 'OK', N, alignment=AC)

    # ─── Stage 30: RFID (rows 119-120) ───
    ws.merge_cells('A119:A120')
    _cell(ws, 119, 1, 30, B, alignment=AC)
    ws.merge_cells('B119:B120')
    _cell(ws, 119, 2, 'RFID', B, alignment=AC)

    _cell(ws, 119, 3, 'RFID Position', B, alignment=AV)
    _cell(ws, 119, 4, 'once', N, alignment=AC); _cell(ws, 119, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F119:G119'); _cell(ws, 119, 6, 'As per Process card', B, alignment=AV)
    ws.merge_cells('H119:N119'); _cell(ws, 119, 8, 'As per Process Card — Center Position', N, alignment=AM)
    _cell(ws, 119, 15, 'OK', N, alignment=AC)

    _cell(ws, 120, 3, 'Cell & Module Make & Manufacturing Month Verification', B, alignment=AV)
    _cell(ws, 120, 4, 'once', N, alignment=AC); _cell(ws, 120, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F120:G120'); _cell(ws, 120, 6, 'As per BOM and Process card', B, alignment=AV)
    ws.merge_cells('H120:N120'); _cell(ws, 120, 8, 'Verified — As per BOM', N, alignment=AM)
    _cell(ws, 120, 15, 'OK', N, alignment=AC)

    # ─── Stage 31: Final Visual Inspection (rows 121-130) ───
    ws.merge_cells('A121:A130')
    _cell(ws, 121, 1, 31, B, alignment=AC)
    ws.merge_cells('B121:B130')
    _cell(ws, 121, 2, 'Final Visual Inspection', B, alignment=AC)

    ws.merge_cells('C121:C125')
    _cell(ws, 121, 3, 'Visual Inspection', B, alignment=AV)
    ws.merge_cells('D121:D125'); _cell(ws, 121, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E121:E125'); _cell(ws, 121, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F121:G125'); _cell(ws, 121, 6, 'Post lam visual inspection criteria', B, alignment=AV)

    snos_final = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([121, 122, 123, 124, 125]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No.', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_final[idx] if idx < len(snos_final) else '', N, alignment=AM)
    _cell(ws, 121, 15, 'OK', N, alignment=AC)

    # Backlabel (126-130)
    ws.merge_cells('C126:C130')
    _cell(ws, 126, 3, 'Backlabel', B, alignment=AV)
    ws.merge_cells('D126:D130'); _cell(ws, 126, 4, '5 pieces', N, alignment=AC)
    ws.merge_cells('E126:E130'); _cell(ws, 126, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F126:G130'); _cell(ws, 126, 6, 'Air bubble,Tilt,Misprint,folded label not acceptable', B, alignment=AV)

    snos_lbl = _gen_serials(prefix, start, 5)
    for idx, r in enumerate([126, 127, 128, 129, 130]):
        ws.merge_cells(f'H{r}:I{r}'); _cell(ws, r, 8, 'S.No', B, alignment=AC)
        ws.merge_cells(f'J{r}:N{r}')
        _cell(ws, r, 10, snos_lbl[idx] if idx < len(snos_lbl) else '', N, alignment=AM)
    _cell(ws, 126, 15, 'OK', N, alignment=AC)

    # ─── Stage 32: Dimension measurement (rows 131-135) ───
    ws.merge_cells('A131:A135')
    _cell(ws, 131, 1, 32, B, alignment=AC)
    ws.merge_cells('B131:B135')
    _cell(ws, 131, 2, 'Dimension measurement', B, alignment=AC)

    _cell(ws, 131, 3, 'L*W and Module Profile', B, alignment=AV)
    _cell(ws, 131, 4, 'once', N, alignment=AC); _cell(ws, 131, 5, 'per shift', N, alignment=AC)
    ws.merge_cells('F131:G132'); _cell(ws, 131, 6, 'As per Module drawing (±1mm)', B, alignment=AV)
    ws.merge_cells('H131:N131'); _cell(ws, 131, 8, '2382mm x 1134mm x 30mm', N, alignment=AM)
    _cell(ws, 131, 15, 'OK', N, alignment=AC)

    _cell(ws, 132, 3, 'Mounting Hole X & Y Pitch', B, alignment=AV)
    _cell(ws, 132, 4, 'once', N, alignment=AC); _cell(ws, 132, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('H132:N132'); _cell(ws, 132, 8, '1400mm x 1091mm', N, alignment=AM)
    _cell(ws, 132, 15, 'OK', N, alignment=AC)

    _cell(ws, 133, 3, 'Diagonal Difference', B, alignment=AV)
    _cell(ws, 133, 4, 'once', N, alignment=AC); _cell(ws, 133, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F133:G133'); _cell(ws, 133, 6, '≤3mm', B, alignment=AV)
    diag = _rr(2.0, 0.5)
    ws.merge_cells('H133:N133'); _cell(ws, 133, 8, f'{diag:.1f}mm', N, alignment=AM)
    _cell(ws, 133, 15, 'OK', N, alignment=AC)

    _cell(ws, 134, 3, 'Corner Gap', B, alignment=AV)
    _cell(ws, 134, 4, 'once', N, alignment=AC); _cell(ws, 134, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F134:G134'); _cell(ws, 134, 6, 'As per visual Inspection criteria', B, alignment=AV)
    cg = _rr(0.02, 0.01)
    ws.merge_cells('H134:N134'); _cell(ws, 134, 8, f'{cg:.2f}mm', N, alignment=AM)
    _cell(ws, 134, 15, 'OK', N, alignment=AC)

    _cell(ws, 135, 3, 'JB Cable length', B, alignment=AV)
    _cell(ws, 135, 4, 'once', N, alignment=AC); _cell(ws, 135, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F135:G135'); _cell(ws, 135, 6, 'As per Process card', B, alignment=AV)
    ws.merge_cells('H135:N135'); _cell(ws, 135, 8, f'{cable_len}mm', N, alignment=AM)
    _cell(ws, 135, 15, 'OK', N, alignment=AC)

    # ─── Stage 33: Packaging (rows 136-139) ───
    ws.merge_cells('A136:A139')
    _cell(ws, 136, 1, 33, B, alignment=AC)
    ws.merge_cells('B136:B139')
    _cell(ws, 136, 2, 'Packaging', B, alignment=AC)

    _cell(ws, 136, 3, 'Packaging Label', B, alignment=AV)
    _cell(ws, 136, 4, 'once', N, alignment=AC); _cell(ws, 136, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F136:G137'); _cell(ws, 136, 6, 'WI For Packaging', B, alignment=AV)
    ws.merge_cells('H136:N136'); _cell(ws, 136, 8, 'As per WI', N, alignment=AM)
    _cell(ws, 136, 15, 'OK', N, alignment=AC)

    _cell(ws, 137, 3, 'Content in Box', B, alignment=AV)
    _cell(ws, 137, 4, 'once', N, alignment=AC); _cell(ws, 137, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('H137:N137'); _cell(ws, 137, 8, 'Verified', N, alignment=AM)
    _cell(ws, 137, 15, 'OK', N, alignment=AC)

    _cell(ws, 138, 3, 'Box Condition', B, alignment=AV)
    _cell(ws, 138, 4, 'once', N, alignment=AC); _cell(ws, 138, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F138:G138'); _cell(ws, 138, 6, 'Damage,dull printing,wet boxes not allowed', B, alignment=AV)
    ws.merge_cells('H138:N138'); _cell(ws, 138, 8, 'Good Condition — No Damage', N, alignment=AM)
    _cell(ws, 138, 15, 'OK', N, alignment=AC)

    _cell(ws, 139, 3, 'Wooden Pallet dimension', B, alignment=AV)
    _cell(ws, 139, 4, 'once', N, alignment=AC); _cell(ws, 139, 5, 'Per shift', N, alignment=AC)
    ws.merge_cells('F139:G139'); _cell(ws, 139, 6, 'Should not be less than module dimension', B, alignment=AV)
    ws.merge_cells('H139:N139'); _cell(ws, 139, 8, '2386mm x 1019mm x 146mm', N, alignment=AM)
    _cell(ws, 139, 15, 'OK', N, alignment=AC)


def _shift_start(shift):
    return {'A': '06:00', 'B': '14:00', 'C': '22:00'}.get(shift, '06:00')


def _shift_mid(shift):
    return {'A': '10:00', 'B': '18:00', 'C': '02:00'}.get(shift, '10:00')
