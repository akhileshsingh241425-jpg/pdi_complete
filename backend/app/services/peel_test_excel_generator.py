"""
Peel Test Excel Report Generator
Generates Excel reports with multiple sheets for Day/Night shifts and Front/Back sides
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime
import os
import random


def create_sheet_data(ws, stringer_name, side_type, date):
    """Create data for one sheet"""
    # Styles
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=9)
    small_font = Font(name='Arial', size=8, bold=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    # Row 1: Logo, Company name, Document No label, Document No value, Details
    # Column A: Logo (rows 1-3)
    ws.merge_cells('A1:A3')
    # Add logo image
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'frontend', 'src', 'components', 'image.png')
    if os.path.exists(logo_path):
        img = OpenpyxlImage(logo_path)
        img.width = 100
        img.height = 60
        ws.add_image(img, 'A1')
    ws['A1'].border = border_thin
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    
    # Columns B-E: Company name (rows 1-3)
    ws.merge_cells('B1:E3')
    ws['B1'] = 'Gautam Solar Private Limited'
    ws['B1'].font = Font(name='Arial', size=14, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].border = border_thin
    
    # Columns F-G: Document labels
    ws.merge_cells('F1:G1')
    ws['F1'] = 'Document No.'
    ws['F1'].font = Font(name='Arial', size=10, bold=True)
    ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F1'].border = border_thin
    
    ws.merge_cells('F2:G2')
    ws['F2'] = 'Issue Date'
    ws['F2'].font = Font(name='Arial', size=10, bold=True)
    ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F2'].border = border_thin
    
    ws.merge_cells('F3:G3')
    ws['F3'] = 'Rev. No. & Date'
    ws['F3'].font = Font(name='Arial', size=10, bold=True)
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F3'].border = border_thin
    
    # Column H: Document values
    ws['H1'] = 'GSPL/IPQC/S5/009'
    ws['H1'].font = Font(name='Arial', size=10, bold=False)
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].border = border_thin
    
    ws['H2'] = '01/11/2024'
    ws['H2'].font = Font(name='Arial', size=10, bold=False)
    ws['H2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H2'].border = border_thin
    
    ws['H3'] = '0'
    ws['H3'].font = Font(name='Arial', size=10, bold=False)
    ws['H3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H3'].border = border_thin
    
    # Row 4: Title section and Page info
    ws.merge_cells('A4:E4')
    ws['A4'] = 'Type of Document:- Peel Test Report\nRibbon to Cell'
    ws['A4'].font = Font(name='Arial', size=10, bold=True)
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws['A4'].border = border_thin
    ws.row_dimensions[4].height = 30
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Page'
    ws['F4'].font = Font(name='Arial', size=10, bold=True)
    ws['F4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F4'].border = border_thin
    
    ws['H4'] = 'Page 1 of 1'
    ws['H4'].font = Font(name='Arial', size=10, bold=False)
    ws['H4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H4'].border = border_thin
    
    # Row 5: Empty with border
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}5'].border = border_thin
    ws.row_dimensions[5].height = 5
    
    # Row 6: Date, Stringer, Shift info
    date_str = date.strftime('%d/%m/%Y')
    
    ws.merge_cells('A6:H6')
    ws['A6'] = f'DATE:- {date_str}  STRINGER:- {stringer_name} {side_type} side  SHIFT:- Day'
    ws['A6'].font = Font(name='Arial', size=10, bold=True)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A6'].border = border_thin
    ws.row_dimensions[6].height = 20
    
    # Row 7: Empty
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}7'].border = border_thin
    
    # Row 8: Table headers
    headers = [
        'No.',
        'MaxForce\n@ 1st\ninterval\n(N)',
        'MaxForce\n@ 2nd\ninterval\n(N)',
        'MaxForce\n@ 3rd\ninterval\n(N)',
        'MaxForce\n@ 4th\ninterval\n(N)',
        'MaxForce\n@ 5th\ninterval\n(N)',
        'MaxForce\n@ 6th\ninterval\n(N)',
        'MaxForce\n@ 7th\ninterval\n(N)'
    ]
    
    for idx, header in enumerate(headers, start=1):
        col_letter = chr(64 + idx)
        cell = ws[f'{col_letter}8']
        cell.value = header
        cell.font = Font(name='Arial', size=8, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        cell.fill = gray_fill
    
    ws.row_dimensions[8].height = 40
    
    # Generate 16 samples with random data (rows 9-24)
    for i in range(1, 17):
        row_num = 8 + i
        
        ws[f'A{row_num}'] = i
        ws[f'A{row_num}'].font = normal_font
        ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row_num}'].border = border_thin
        
        for interval in range(7):
            col_letter = chr(66 + interval)
            value = round(random.uniform(2.0, 4.0), 3)
            cell = ws[f'{col_letter}{row_num}']
            cell.value = value
            cell.number_format = '0.000'
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        ws.row_dimensions[row_num].height = 18
    
    # Signature section
    sign_row = row_num + 3
    
    # Random names for signatures
    test_names = ['Shubham', 'Mohit', 'Raju', 'Shubham', 'Sureden']
    verify_names = ['Aman', 'Taj']
    approve_names = ['Aman', 'Taj']
    
    test_by = random.choice(test_names)
    verify_by = random.choice(verify_names)
    approve_by = random.choice(approve_names)
    
    # Test Performed By
    ws.merge_cells(f'A{sign_row}:B{sign_row}')
    ws[f'A{sign_row}'] = 'Test Performed By:'
    ws[f'A{sign_row}'].font = Font(name='Arial', size=9, bold=True)
    ws[f'A{sign_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{sign_row}'].border = border_thin
    
    ws.merge_cells(f'C{sign_row}:D{sign_row}')
    ws[f'C{sign_row}'] = test_by
    ws[f'C{sign_row}'].font = Font(name='Arial', size=9)
    ws[f'C{sign_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'C{sign_row}'].border = border_thin
    
    # Verified By
    ws[f'E{sign_row}'] = 'Verified By:'
    ws[f'E{sign_row}'].font = Font(name='Arial', size=9, bold=True)
    ws[f'E{sign_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'E{sign_row}'].border = border_thin
    
    ws[f'F{sign_row}'] = verify_by
    ws[f'F{sign_row}'].font = Font(name='Arial', size=9)
    ws[f'F{sign_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'F{sign_row}'].border = border_thin
    
    # Approved By
    ws[f'G{sign_row}'] = 'Approved By:'
    ws[f'G{sign_row}'].font = Font(name='Arial', size=9, bold=True)
    ws[f'G{sign_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'G{sign_row}'].border = border_thin
    
    ws[f'H{sign_row}'] = approve_by
    ws[f'H{sign_row}'].font = Font(name='Arial', size=9)
    ws[f'H{sign_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'H{sign_row}'].border = border_thin
    
    ws.row_dimensions[sign_row].height = 25

def generate_peel_test_excel(line_number, date=None, output_folder='generated_pdfs'):
    """
    Generate Excel report for 1 line with 12 sheets (3 stringers × 2 sides × 2 positions)
    
    Args:
        line_number: Line number (1, 2, or 3)
        date: Report date (datetime object or string)
        output_folder: Output directory path
    
    Returns:
        str: Path to generated Excel file
    """
    if date is None:
        date = datetime.now()
    elif isinstance(date, str):
        date = datetime.strptime(date, '%Y-%m-%d')
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # 3 stringers per line, each with A and B sides
    stringers = [1, 2, 3]
    sides_suffix = ['A', 'B']
    positions = ['Front', 'Back']
    
    # Create 12 sheets: Stringer 1A-Front, 1A-Back, 1B-Front, 1B-Back, 2A-Front, 2A-Back, etc.
    for stringer_num in stringers:
        for side_suffix in sides_suffix:
            for position in positions:
                stringer_name = f'{stringer_num}{side_suffix}'
                sheet_name = f'{stringer_name}-{position}'
                ws = wb.create_sheet(title=sheet_name)
                create_sheet_data(ws, stringer_name, position, date)
    
    # Save file
    filename = f'PeelTest_Line{line_number}_Day_{date.strftime("%Y%m%d")}.xlsx'
    filepath = os.path.join(output_folder, filename)
    wb.save(filepath)
    
    return filepath
