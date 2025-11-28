from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import json

def generate_production_excel(company, production_data, rejections, start_date, end_date, 
                              cells_received_qty=0, cells_received_mw=0, report_options=None):
    """
    Generate colorful Excel report with multiple sheets based on selected options
    """
    if report_options is None:
        report_options = {
            'includeProductionDetails': True,
            'includeCellInventory': True,
            'includeKPIMetrics': True,
            'includeDayWiseSummary': True,
            'includeRejections': True
        }
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Company Info & Summary (Always included)
    create_summary_sheet(wb, company, production_data, start_date, end_date, 
                        cells_received_qty, cells_received_mw)
    
    # Sheet 2: Daily Production Details
    if report_options.get('includeProductionDetails', True):
        create_production_sheet(wb, production_data)
    
    # Sheet 3: Cell Inventory
    if report_options.get('includeCellInventory', True):
        create_inventory_sheet(wb, production_data, cells_received_qty)
    
    # Sheet 4: KPI Metrics
    if report_options.get('includeKPIMetrics', True):
        create_kpi_sheet(wb, production_data, cells_received_qty)
    
    # Sheet 5: Day-wise Rejection Summary
    if report_options.get('includeDayWiseSummary', True) and report_options.get('includeRejections', True):
        create_rejection_summary_sheet(wb, production_data)
    
    # Sheet 6: Detailed Rejections
    if report_options.get('includeRejections', True):
        create_rejection_details_sheet(wb, rejections)
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Production_Report_{company['name']}_{timestamp}.xlsx"
    
    # Create absolute path to generated_pdfs folder
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'generated_pdfs')
    os.makedirs(output_dir, exist_ok=True)
    
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath

def create_summary_sheet(wb, company, production_data, start_date, end_date, 
                        cells_received_qty, cells_received_mw):
    """Sheet 1: Company Info & Summary"""
    ws = wb.create_sheet("Summary", 0)
    
    # Header styling
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "PRODUCTION REPORT"
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Company Info Header
    ws.merge_cells('A3:F3')
    ws['A3'] = "COMPANY INFORMATION"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    
    # Company Details
    info_data = [
        ['Company Name:', company.get('name', 'N/A'), 'Report Period:', f"{start_date} to {end_date}"],
        ['Address:', company.get('address', 'N/A'), 'Generated On:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ['Contact:', company.get('contact', 'N/A'), 'Module Type:', company.get('module_type', 'N/A')],
    ]
    
    label_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11)
    
    row = 4
    for info_row in info_data:
        ws[f'A{row}'] = info_row[0]
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = label_fill
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = info_row[1]
        ws[f'B{row}'].font = value_font
        
        ws[f'D{row}'] = info_row[2]
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].fill = label_fill
        
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'] = info_row[3]
        ws[f'E{row}'].font = value_font
        
        row += 1
    
    # Production Summary Header
    ws.merge_cells(f'A{row+1}:F{row+1}')
    ws[f'A{row+1}'] = "PRODUCTION SUMMARY"
    ws[f'A{row+1}'].font = header_font
    ws[f'A{row+1}'].fill = header_fill
    ws[f'A{row+1}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row+1].height = 25
    
    # Calculate totals
    total_day = sum(p.get('day_production', 0) for p in production_data)
    total_night = sum(p.get('night_production', 0) for p in production_data)
    total_prod = total_day + total_night
    total_power = total_prod * company.get('module_wattage', 0) / 1000
    
    row += 2
    summary_data = [
        ['Total Production:', f"{total_prod:,} Modules", 'Day Production:', f"{total_day:,} Modules"],
        ['Total Power:', f"{total_power:,.2f} kW", 'Night Production:', f"{total_night:,} Modules"],
        ['Production Days:', f"{len(production_data)} Days", 'Avg Daily:', f"{total_prod//len(production_data) if production_data else 0:,} Modules"],
    ]
    
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    for summary_row in summary_data:
        ws[f'A{row}'] = summary_row[0]
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = green_fill
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = summary_row[1]
        ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = summary_row[2]
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].fill = green_fill
        
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'] = summary_row[3]
        ws[f'E{row}'].font = Font(name='Calibri', size=12, bold=True)
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_cells in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.border = thin_border

def create_production_sheet(wb, production_data):
    """Sheet 2: Daily Production Details"""
    ws = wb.create_sheet("Production Details")
    
    # Header
    headers = ['Date', 'Day of Week', 'Day Shift', 'Night Shift', 'Total Production', 
               'Cells Used', 'Cell Rej %', 'Cells Rejected', 'Module Rej %', 'Modules Rejected']
    
    header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 25
    
    # Data rows
    data_font = Font(name='Calibri', size=10)
    light_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    for idx, prod in enumerate(production_data, 2):
        day_prod = prod.get('day_production', 0)
        night_prod = prod.get('night_production', 0)
        total = day_prod + night_prod
        cells_used = total * 132
        
        row_data = [
            prod.get('date', ''),
            prod.get('day_of_week', ''),
            day_prod,
            night_prod,
            total,
            cells_used,
            prod.get('cell_rejection_percent', 0),
            prod.get('cells_rejected', 0),
            prod.get('module_rejection_percent', 0),
            prod.get('modules_rejected', 0)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if idx % 2 == 0:
                cell.fill = light_fill
            
            # Number formatting
            if col in [3, 4, 5, 6, 8, 10]:  # Production and rejection numbers
                cell.number_format = '#,##0'
            elif col in [7, 9]:  # Percentages
                cell.number_format = '0.00%'
    
    # Set column widths
    widths = [12, 14, 12, 12, 15, 12, 12, 14, 12, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(production_data)+1, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border

def create_inventory_sheet(wb, production_data, cells_received_qty):
    """Sheet 3: Cell Inventory"""
    ws = wb.create_sheet("Cell Inventory")
    
    # Convert MW to cells if needed
    if cells_received_qty < 10000:
        cells_received_qty = int(cells_received_qty * 103219.2)
    
    # Calculate values
    total_production = sum(p.get('day_production', 0) + p.get('night_production', 0) for p in production_data)
    cells_used = total_production * 132
    cells_rejected = sum(p.get('cells_rejected', 0) for p in production_data)
    cells_remaining = cells_received_qty - cells_used - cells_rejected
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "CELL INVENTORY TRACKING"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    header_fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    
    headers = ['Category', 'Quantity (Cells)', 'Percentage', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[3].height = 25
    
    # Data
    inventory_data = [
        ['Cells Received', cells_received_qty, 100.0, '✓ Initial Stock'],
        ['Cells Used', cells_used, (cells_used/cells_received_qty*100) if cells_received_qty > 0 else 0, '→ Production'],
        ['Cells Rejected', cells_rejected, (cells_rejected/cells_received_qty*100) if cells_received_qty > 0 else 0, '✗ Defective'],
        ['Cells Remaining', cells_remaining, (cells_remaining/cells_received_qty*100) if cells_received_qty > 0 else 0, '◉ Available'],
    ]
    
    fills = [
        PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid"),
        PatternFill(start_color="C5E1A5", end_color="C5E1A5", fill_type="solid"),
    ]
    
    for idx, (data, fill) in enumerate(zip(inventory_data, fills), 4):
        ws.cell(row=idx, column=1, value=data[0]).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=idx, column=2, value=data[1]).number_format = '#,##0'
        ws.cell(row=idx, column=3, value=data[2]/100).number_format = '0.00%'
        ws.cell(row=idx, column=4, value=data[3]).font = Font(name='Calibri', size=10)
        
        for col in range(1, 5):
            cell = ws.cell(row=idx, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

def create_kpi_sheet(wb, production_data, cells_received_qty):
    """Sheet 4: KPI Metrics"""
    ws = wb.create_sheet("KPI Metrics")
    
    # Calculate KPIs
    total_production = sum(p.get('day_production', 0) + p.get('night_production', 0) for p in production_data)
    total_modules_rejected = sum(p.get('modules_rejected', 0) for p in production_data)
    total_cells_rejected = sum(p.get('cells_rejected', 0) for p in production_data)
    cells_used = total_production * 132
    
    module_rej_pct = (total_modules_rejected / total_production * 100) if total_production > 0 else 0
    cell_rej_pct = (total_cells_rejected / cells_used * 100) if cells_used > 0 else 0
    efficiency = ((total_production - total_modules_rejected) / total_production * 100) if total_production > 0 else 0
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "KEY PERFORMANCE INDICATORS (KPI)"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # KPI Data
    kpis = [
        ['Production KPI', total_production, 'Modules', '#4CAF50', '✓ Good'],
        ['Power Generated', total_production * 545 / 1000, 'kW', '#2196F3', '⚡ Energy'],
        ['Module Rejection %', module_rej_pct, '%', '#FF9800' if module_rej_pct < 1 else '#F44336', '⚠ Quality'],
        ['Cell Rejection %', cell_rej_pct, '%', '#FF9800' if cell_rej_pct < 1 else '#F44336', '⚠ Defects'],
        ['Production Efficiency', efficiency, '%', '#4CAF50', '✓ Performance'],
    ]
    
    row = 3
    for kpi in kpis:
        # KPI Name
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = kpi[0]
        cell.font = Font(name='Calibri', size=13, bold=True)
        cell.fill = PatternFill(start_color=kpi[3][1:], end_color=kpi[3][1:], fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Value
        ws[f'C{row}'] = kpi[1]
        ws[f'C{row}'].font = Font(name='Calibri', size=14, bold=True)
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        if kpi[2] == '%':
            ws[f'C{row}'].number_format = '0.00"%"'
        else:
            ws[f'C{row}'].number_format = '#,##0.00'
        
        # Unit
        ws[f'D{row}'] = kpi[2]
        ws[f'D{row}'].font = Font(name='Calibri', size=11)
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Status
        ws[f'E{row}'] = kpi[4]
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[row].height = 30
        row += 2
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # Add borders
    thin_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    for r in range(3, row, 2):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = thin_border

def create_rejection_summary_sheet(wb, production_data):
    """Sheet 5: Day-wise Rejection Summary"""
    ws = wb.create_sheet("Rejection Summary")
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DAY-WISE REJECTION SUMMARY"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = ['Date', 'Day', 'Modules Rejected', 'Rejection %', 'Status']
    header_fill = PatternFill(start_color="F48FB1", end_color="F48FB1", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[3].height = 25
    
    # Data
    for idx, prod in enumerate(production_data, 4):
        total = prod.get('day_production', 0) + prod.get('night_production', 0)
        rejected = prod.get('modules_rejected', 0)
        rej_pct = (rejected / total * 100) if total > 0 else 0
        
        # Status indicator
        if rej_pct < 0.5:
            status = '✓ Low'
            status_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif rej_pct < 1:
            status = '⚠ Medium'
            status_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        elif rej_pct < 2:
            status = '⚠ High'
            status_fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
        else:
            status = '✗ Critical'
            status_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        ws.cell(row=idx, column=1, value=prod.get('date', '')).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=prod.get('day_of_week', '')).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=3, value=rejected).number_format = '#,##0'
        ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=4, value=rej_pct/100).number_format = '0.00%'
        ws.cell(row=idx, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=5, value=status).fill = status_fill
        ws.cell(row=idx, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=5).font = Font(name='Calibri', size=10, bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=3, max_row=len(production_data)+3, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

def create_rejection_details_sheet(wb, rejections):
    """Sheet 6: Detailed Rejections"""
    ws = wb.create_sheet("Rejection Details")
    
    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "DETAILED REJECTION RECORDS"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = ['No', 'Date', 'Serial Number', 'Defect Reason', 'Stage', 'Defect Type', 'Remarks']
    header_fill = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[3].height = 25
    
    # Data
    light_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    for idx, rej in enumerate(rejections, 4):
        defect_type = rej.get('defect_type', 'Minor')
        type_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid") if defect_type == 'Major' else PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        
        ws.cell(row=idx, column=1, value=idx-3).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=rej.get('date', '')).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=3, value=rej.get('serial', '')).alignment = Alignment(horizontal='left')
        ws.cell(row=idx, column=4, value=rej.get('reason', '')).alignment = Alignment(horizontal='left')
        ws.cell(row=idx, column=5, value=rej.get('stage', '')).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=6, value=defect_type).fill = type_fill
        ws.cell(row=idx, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=6).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=idx, column=7, value=rej.get('remarks', '')).alignment = Alignment(horizontal='left')
        
        if idx % 2 == 0:
            for col in [1, 2, 3, 4, 5, 7]:
                ws.cell(row=idx, column=col).fill = light_fill
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=3, max_row=len(rejections)+3, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

def generate_ipqc_excel(ipqc_data, bom_data, metadata):
    """
    Generate IPQC report in Excel format - exactly same as PDF
    
    Args:
        ipqc_data: List of stages with checkpoints
        bom_data: Bill of Materials data
        metadata: Report metadata (date, shift, customer, etc.)
    
    Returns:
        str: Path to generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IPQC Report"
    
    # Define styles
    title_font = Font(name='Calibri', size=14, bold=True)
    header_font = Font(name='Calibri', size=10, bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    label_font = Font(name='Calibri', size=9, bold=True)
    value_font = Font(name='Calibri', size=9)
    
    stage_font = Font(name='Calibri', size=9, bold=True)
    stage_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    checkpoint_font = Font(name='Calibri', size=8)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Company Header (3 columns like PDF)
    ws.merge_cells('A1:B1')
    ws['A1'] = "GAUTAM\nSOLAR"
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].border = thin_border
    
    ws.merge_cells('C1:E1')
    ws['C1'] = "Gautam Solar Private Limited\nIPQC Check Sheet"
    ws['C1'].font = Font(name='Calibri', size=12, bold=True)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['C1'].border = thin_border
    
    ws['F1'] = "Document No."
    ws['F1'].font = label_font
    ws['F1'].alignment = Alignment(horizontal='left', vertical='top')
    ws['F1'].border = thin_border
    
    ws['G1'] = metadata.get('customer_id', 'GSPL/IPQC/IPC/003')
    ws['G1'].font = value_font
    ws['G1'].alignment = Alignment(horizontal='left', vertical='top')
    ws['G1'].border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Row 2: Date, Time, Shift, PO
    ws['A2'] = f"Date: {metadata.get('date', '')}"
    ws['A2'].font = value_font
    ws['A2'].border = thin_border
    
    ws['B2'] = "Time:"
    ws['B2'].font = value_font
    ws['B2'].border = thin_border
    
    ws['C2'] = f"Shift: {metadata.get('shift', '')}"
    ws['C2'].font = value_font
    ws['C2'].border = thin_border
    
    ws.merge_cells('D2:G2')
    ws['D2'] = f"Po.no.: {metadata.get('po_number', '')}"
    ws['D2'].font = value_font
    ws['D2'].border = thin_border
    
    ws.row_dimensions[2].height = 18
    
    # Row 3: Column headers (EXACTLY same as PDF - 7 columns)
    row = 3
    headers = ['Sr.No.', 'Stage', 'Check point', 'Quantum of Check\nSample Size | Frequency', 'Shift\nAcceptance Criteria', 'Monitoring Result', 'Remarks,If any']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = 30
    
    # Data rows - Stage by stage (EXACTLY like PDF - 7 columns with merge)
    row += 1
    for stage in ipqc_data:
        stage_no = stage.get('sr_no', '')
        stage_name = stage.get('stage', '')
        checkpoints = stage.get('checkpoints', [])
        
        if not checkpoints:
            continue
            
        start_row = row
        
        # Each checkpoint as a row
        for idx, checkpoint in enumerate(checkpoints):
            # Sr.No (only in first checkpoint)
            if idx == 0:
                ws[f'A{row}'] = stage_no
            else:
                ws[f'A{row}'] = ''
            
            # Stage will be merged later
            ws[f'B{row}'] = stage_name if idx == 0 else ''
            
            # Checkpoint
            ws[f'C{row}'] = checkpoint.get('checkpoint', '')
            
            # Sample Size and Frequency
            sample_freq = f"{checkpoint.get('sample_size', '')}\n{checkpoint.get('frequency', '')}"
            ws[f'D{row}'] = sample_freq
            
            # Acceptance Criteria
            ws[f'E{row}'] = checkpoint.get('acceptance_criteria', '')
            
            # Monitoring Result
            monitoring = checkpoint.get('monitoring_result', '')
            if isinstance(monitoring, list):
                monitoring = '\n'.join(monitoring)
            ws[f'F{row}'] = monitoring
            
            # Remarks
            ws[f'G{row}'] = checkpoint.get('remarks', '')
            
            # Apply styling to all cells in row
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = checkpoint_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row += 1
        
        # Merge Stage column (B) for all checkpoints of this stage
        if len(checkpoints) > 1:
            ws.merge_cells(f'B{start_row}:B{row-1}')
        
        # Apply stage styling to merged cell
        ws[f'B{start_row}'].font = stage_font
        ws[f'B{start_row}'].fill = stage_fill
        ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Set column widths (matching PDF proportions)
    ws.column_dimensions['A'].width = 8      # Sr.No
    ws.column_dimensions['B'].width = 20     # Stage
    ws.column_dimensions['C'].width = 30     # Check point
    ws.column_dimensions['D'].width = 20     # Quantum of Check
    ws.column_dimensions['E'].width = 30     # Acceptance Criteria
    ws.column_dimensions['F'].width = 35     # Monitoring Result
    ws.column_dimensions['G'].width = 18     # Remarks
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    customer_name = metadata.get('customer_id', 'Unknown').replace('/', '_')
    filename = f"IPQC_Report_{customer_name}_{timestamp}.xlsx"
    
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'generated_pdfs')
    os.makedirs(output_dir, exist_ok=True)
    
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath

def create_ipqc_metadata_sheet(wb, metadata, bom_data):
    """Sheet 1: Metadata & BOM Information"""
    ws = wb.create_sheet("Report Info", 0)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "IPQC INSPECTION REPORT"
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Report Metadata Header
    ws.merge_cells('A3:F3')
    ws['A3'] = "REPORT METADATA"
    ws['A3'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    
    # Metadata
    label_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11)
    
    metadata_info = [
        ['Date:', metadata.get('date', 'N/A'), 'Shift:', metadata.get('shift', 'N/A')],
        ['Customer ID:', metadata.get('customer_id', 'N/A'), 'PO Number:', metadata.get('po_number', 'N/A')],
        ['Serial Prefix:', metadata.get('serial_prefix', 'N/A'), 'Serial Start:', metadata.get('serial_start', 'N/A')],
        ['Module Count:', metadata.get('module_count', 'N/A'), 'Cell Manufacturer:', metadata.get('cell_manufacturer', 'N/A')],
        ['Cell Efficiency:', f"{metadata.get('cell_efficiency', 'N/A')}%", 'JB Cable Length:', f"{metadata.get('jb_cable_length', 'N/A')} mm"],
        ['Golden Module:', metadata.get('golden_module_number', 'N/A'), 'Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    row = 4
    for info in metadata_info:
        ws[f'A{row}'] = info[0]
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = label_fill
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = info[1]
        ws[f'B{row}'].font = value_font
        
        ws[f'D{row}'] = info[2]
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].fill = label_fill
        
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'] = info[3]
        ws[f'E{row}'].font = value_font
        
        row += 1
    
    # BOM Header
    ws.merge_cells(f'A{row+1}:F{row+1}')
    ws[f'A{row+1}'] = "BILL OF MATERIALS (BOM)"
    ws[f'A{row+1}'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    ws[f'A{row+1}'].fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    ws[f'A{row+1}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row+1].height = 25
    
    row += 2
    
    # BOM Data
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    bom_items = [
        ['Customer Name:', bom_data.get('customer_name', 'N/A'), 'Module Type:', bom_data.get('module_type', 'N/A')],
        ['Cell Type:', bom_data.get('cell_type', 'N/A'), 'Cell Size:', bom_data.get('cell_size', 'N/A')],
        ['Glass Type:', bom_data.get('glass_type', 'N/A'), 'Glass Thickness:', bom_data.get('glass_thickness', 'N/A')],
        ['EVA Type:', bom_data.get('eva_type', 'N/A'), 'Backsheet Type:', bom_data.get('backsheet_type', 'N/A')],
        ['Frame Type:', bom_data.get('frame_type', 'N/A'), 'JB Type:', bom_data.get('jb_type', 'N/A')],
    ]
    
    for bom_info in bom_items:
        ws[f'A{row}'] = bom_info[0]
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = green_fill
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = bom_info[1]
        ws[f'B{row}'].font = value_font
        
        ws[f'D{row}'] = bom_info[2]
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].fill = green_fill
        
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'] = bom_info[3]
        ws[f'E{row}'].font = value_font
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_cells in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.border = thin_border

def create_ipqc_checkpoints_sheet(wb, ipqc_data):
    """Sheet 2: All IPQC Checkpoints"""
    ws = wb.create_sheet("IPQC Checkpoints")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "IPQC INSPECTION CHECKPOINTS"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['Stage No', 'Stage Name', 'Checkpoint', 'Specification', 'Method', 'Acceptance', 'Sample', 'Monitoring Result']
    header_fill = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[3].height = 30
    
    # Data
    data_font = Font(name='Calibri', size=10)
    light_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    stage_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    
    row = 4
    for stage in ipqc_data:
        stage_no = stage.get('sr_no', '')
        stage_name = stage.get('stage', '')
        checkpoints = stage.get('checkpoints', [])
        
        for idx, checkpoint in enumerate(checkpoints):
            # Stage number and name only on first checkpoint
            if idx == 0:
                ws.cell(row=row, column=1, value=stage_no).fill = stage_fill
                ws.cell(row=row, column=2, value=stage_name).fill = stage_fill
            else:
                ws.cell(row=row, column=1, value='').fill = light_fill if row % 2 == 0 else PatternFill()
                ws.cell(row=row, column=2, value='').fill = light_fill if row % 2 == 0 else PatternFill()
            
            # Checkpoint data
            ws.cell(row=row, column=3, value=checkpoint.get('checkpoint', ''))
            ws.cell(row=row, column=4, value=checkpoint.get('specification', ''))
            ws.cell(row=row, column=5, value=checkpoint.get('method', ''))
            ws.cell(row=row, column=6, value=checkpoint.get('acceptance_criteria', ''))
            ws.cell(row=row, column=7, value=checkpoint.get('sample_size', ''))
            ws.cell(row=row, column=8, value=checkpoint.get('monitoring_result', ''))
            
            # Styling
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                if row % 2 == 0 and col > 2:
                    cell.fill = light_fill
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 35

def create_ipqc_summary_sheet(wb, ipqc_data):
    """Sheet 3: Stage-wise Summary"""
    ws = wb.create_sheet("Stage Summary")
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "STAGE-WISE SUMMARY"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['Stage No', 'Stage Name', 'Total Checkpoints', 'Critical Points', 'Status']
    header_fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[3].height = 25
    
    # Data
    light_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ok_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    for idx, stage in enumerate(ipqc_data, 4):
        checkpoints = stage.get('checkpoints', [])
        critical_count = sum(1 for cp in checkpoints if 'critical' in cp.get('acceptance_criteria', '').lower())
        
        ws.cell(row=idx, column=1, value=stage.get('sr_no', ''))
        ws.cell(row=idx, column=2, value=stage.get('stage', ''))
        ws.cell(row=idx, column=3, value=len(checkpoints))
        ws.cell(row=idx, column=4, value=critical_count)
        ws.cell(row=idx, column=5, value='✓ Inspected').fill = ok_fill
        
        for col in range(1, 6):
            cell = ws.cell(row=idx, column=col)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if idx % 2 == 0 and col < 5:
                cell.fill = light_fill
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
